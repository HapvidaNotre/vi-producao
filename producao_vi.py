import streamlit as st
import time
import base64
from datetime import datetime, timezone, timedelta
from pathlib import Path
import csv, io
import streamlit.components.v1 as components
import pandas as pd
import re
import requests
import json

# Fuso horário de Brasília (UTC-3)
_TZ_BR = timezone(timedelta(hours=-3))
def now_br():
    """Retorna datetime atual no fuso de Brasília."""
    return datetime.now(_TZ_BR)

st.set_page_config(
    page_title="Vi Lingerie - Producao",
    page_icon="👗",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────
OPERADORES  = ["Lucivanio","Enagio","Daniel","Italo","Cildenir","Samya","Neide","Eduardo","Talyson"]
ETAPAS      = ["Separacao","Mesa_Embalagem","Conferencia"]
ETAPAS_LBL  = ["Separação de Pedidos","Mesa de Embalagem","Conferência de Pedidos"]
ADMIN_SENHA = "vi2026"

# ─────────────────────────────────────
#  SUPABASE CONFIG
# ─────────────────────────────────────
try:
    SB_URL = st.secrets["SUPABASE_URL"]
    SB_KEY = st.secrets["SUPABASE_KEY"]
except:
    SB_URL = "https://uiybrhaqtcwejtbbctge.supabase.co"
    SB_KEY = "sb_publishable_oIpsQMjK3QGL6IPkccJHqQ_OQCfrJf0"

def _sb_headers():
    return {
        "apikey": SB_KEY,
        "Authorization": f"Bearer {SB_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

def _get(table, params="", paginar=False):
    """
    Busca registros do Supabase.
    paginar=True: faz múltiplas requisições para retornar TODOS os registros
    (o Supabase limita 1000 por request por padrão).
    """
    headers = {**_sb_headers(), "Prefer": "count=none"}
    if not paginar:
        r = requests.get(f"{SB_URL}/rest/v1/{table}?{params}",
                         headers=headers, timeout=10)
        return r.json() if r.ok else []
    # Paginação automática — retorna todos os registros sem corte
    PAGE = 1000
    todos = []
    offset = 0
    while True:
        sep = "&" if params else ""
        url = f"{SB_URL}/rest/v1/{table}?{params}{sep}limit={PAGE}&offset={offset}"
        r = requests.get(url, headers=headers, timeout=15)
        if not r.ok:
            break
        lote = r.json()
        if not isinstance(lote, list):
            break
        todos.extend(lote)
        if len(lote) < PAGE:
            break   # última página
        offset += PAGE
    return todos

def _post(table, data):
    r = requests.post(f"{SB_URL}/rest/v1/{table}", headers=_sb_headers(),
                      data=json.dumps(data), timeout=10)
    return r.ok

def _patch(table, match_params, data):
    r = requests.patch(f"{SB_URL}/rest/v1/{table}?{match_params}",
                       headers={**_sb_headers(), "Prefer": "return=minimal"},
                       data=json.dumps(data), timeout=10)
    return r.ok

def _delete(table, params):
    r = requests.delete(f"{SB_URL}/rest/v1/{table}?{params}",
                        headers=_sb_headers(), timeout=10)
    return r.ok

def _upsert(table, data, on_conflict):
    h = {**_sb_headers(), "Prefer": f"resolution=merge-duplicates,return=minimal"}
    r = requests.post(f"{SB_URL}/rest/v1/{table}?on_conflict={on_conflict}",
                      headers=h, data=json.dumps(data), timeout=10)
    if not r.ok:
        # Grava o erro no session_state para diagnóstico — visível na tela admin
        try:
            _err_detail = r.json()
        except Exception:
            _err_detail = r.text
        import streamlit as _st
        _st.session_state["_ultimo_erro_supabase"] = {
            "status": r.status_code,
            "table": table,
            "detail": _err_detail,
        }
    return r.ok

# ─────────────────────────────────────
#  DATABASE — Supabase REST API
# ─────────────────────────────────────
def init_db():
    pass  # Tables created via supabase_setup.sql

# ─── Planilha / Pedidos base ───
@st.cache_data(ttl=15, show_spinner=False)
def buscar_pedidos_base():
    rows = _get("pedidos_base", "select=numero,cliente,produto,status&order=numero.asc", paginar=True)
    if isinstance(rows, list):
        return [(r["numero"], r.get("cliente",""), r.get("produto",""), r.get("status","aberto"))
                for r in rows]
    return []

def status_pedido(numero):
    rows = _get("pedidos_base", f"numero=eq.{numero}&select=status")
    if not rows: return "nao_encontrado"
    return rows[0].get("status", "nao_encontrado")

def cadastrar_pedido_avulso(numero, cliente="", produto="", est_alocado=None, vr_alocado=None):
    now_str = now_br().strftime("%d/%m/%Y %H:%M")
    payload = {"numero": numero, "cliente": cliente, "produto": produto,
               "status": "aberto", "importado_em": now_str}
    if est_alocado is not None:
        payload["est_alocado"] = est_alocado
    if vr_alocado is not None:
        payload["vr_alocado"] = vr_alocado
    ok = _upsert("pedidos_base", payload, "numero")
    buscar_pedidos_base.clear()
    return ok

def marcar_concluido(numero):
    """
    Marca o pedido como concluído no banco APENAS se ele era AF (percentual < 99.9).
    Pedidos que já vinham com 100% na planilha permanecem intocados no Sistema A
    (já estavam na aba Tratados). Só pedidos AF que passaram pelas 3 etapas
    do Sistema B precisam ter o status atualizado.
    """
    try:
        rows = _get("pedidos_base", f"numero=eq.{numero}&select=percentual")
        if isinstance(rows, list) and rows:
            pct = float(rows[0].get("percentual") or 0)
            if pct >= 99.9:
                return  # Já era 100% — permanece como está no Sistema A
    except Exception:
        pass
    _patch("pedidos_base", f"numero=eq.{numero}", {"status": "concluido"})

def verificar_etapa_registro(pedido, etapa_idx):
    rows = _get("registros",
        f"pedido=eq.{pedido}&etapa_idx=eq.{etapa_idx}&select=operador&order=id.desc&limit=1")
    if rows: return True, rows[0].get("operador")
    return False, None

def pedido_em_andamento(pedido, etapa_idx):
    rows = _get("sessoes_ativas",
        f"pedido=eq.{pedido}&etapa_idx=eq.{etapa_idx}&select=operador,iniciado_em")
    if rows:
        r = rows[0]
        ini_ts = int(r.get("iniciado_em", 0))
        # iniciado_em == 0 significa pausado — não conta como em andamento
        if ini_ts == 0:
            return False, None
        if int(time.time()) - ini_ts < 14400:
            return True, r.get("operador")
    return False, None

def registrar_sessao_ativa(pedido, etapa_idx, operador):
    _upsert("sessoes_ativas",
            {"pedido": pedido, "etapa_idx": etapa_idx,
             "operador": operador, "iniciado_em": int(time.time())},
            "pedido,etapa_idx")
    buscar_todas_sessoes_ativas.clear()

def remover_sessao_ativa(pedido, etapa_idx):
    _delete("sessoes_ativas", f"pedido=eq.{pedido}&etapa_idx=eq.{etapa_idx}")
    buscar_todas_sessoes_ativas.clear()

def pausar_indefinidamente(pedido, etapa_idx, operador, tempo_acumulado):
    """
    Salva o tempo acumulado na sessão ativa (campo tempo_pausado).
    O cronômetro fica 'congelado' por tempo indeterminado — iniciado_em=0
    sinaliza que está pausado. A sessão NÃO expira automaticamente.
    """
    _upsert("sessoes_ativas", {
        "pedido":        pedido,
        "etapa_idx":     etapa_idx,
        "operador":      operador,
        "iniciado_em":   0,               # 0 = pausado por tempo indeterminado
        "tempo_pausado": tempo_acumulado, # segundos já trabalhados
    }, "pedido,etapa_idx")
    buscar_todas_sessoes_ativas.clear()

# Mantém alias antigo para compatibilidade
def pausar_para_amanha(pedido, etapa_idx, operador, tempo_acumulado):
    pausar_indefinidamente(pedido, etapa_idx, operador, tempo_acumulado)

def registrar_pausa_log(pedido, etapa_idx, operador, tempo_pausado_s, motivo=""):
    """Salva um registro permanente da pausa na tabela pausas_log."""
    payload = {
        "pedido":          str(pedido),
        "etapa_idx":       int(etapa_idx),
        "operador":        operador,
        "pausado_em":      now_br().strftime("%d/%m/%Y %H:%M"),
        "tempo_pausado_s": int(tempo_pausado_s),
        "motivo":          motivo.strip() if motivo else "",
    }
    _post("pausas_log", payload)

@st.cache_data(ttl=10, show_spinner=False)
def buscar_pausas_log():
    rows = _get("pausas_log", "select=*&order=id.desc", paginar=True)
    if isinstance(rows, list):
        return [(r.get("id"), r.get("pedido"), r.get("operador"),
                 r.get("etapa_idx"), r.get("pausado_em"),
                 r.get("tempo_pausado_s"), r.get("motivo", ""))
                for r in rows]
    return []

@st.cache_data(ttl=8, show_spinner=False)
def buscar_pedidos_pausados():
    """
    Retorna as sessões atualmente em pausa (iniciado_em == 0).
    Para cada uma, busca a data/hora da última entrada no pausas_log.
    Retorna lista de dicts: pedido, operador, etapa_idx, pausado_em, tempo_pausado_s, motivo.
    """
    rows = _get("sessoes_ativas",
                "select=pedido,operador,etapa_idx,tempo_pausado&iniciado_em=eq.0")
    if not isinstance(rows, list):
        return []

    # Busca logs de pausa para obter data/hora de cada pausa ativa
    pausas_log_rows = _get("pausas_log",
                           "select=pedido,etapa_idx,pausado_em,motivo&order=id.desc",
                           paginar=True)
    # Monta índice: (pedido, etapa_idx) -> último registro de pausa
    pausa_ts_map = {}
    if isinstance(pausas_log_rows, list):
        for p in pausas_log_rows:
            chave = (str(p.get("pedido","")), int(p.get("etapa_idx", 0)))
            if chave not in pausa_ts_map:   # order desc — pega o mais recente
                pausa_ts_map[chave] = {
                    "pausado_em": p.get("pausado_em", "—"),
                    "motivo":     p.get("motivo", ""),
                }

    resultado = []
    for r in rows:
        ped = str(r.get("pedido", ""))
        eta = int(r.get("etapa_idx", 0))
        chave = (ped, eta)
        info_log = pausa_ts_map.get(chave, {})
        resultado.append({
            "pedido":        ped,
            "operador":      r.get("operador", ""),
            "etapa_idx":     eta,
            "tempo_pausado": int(r.get("tempo_pausado") or 0),
            "pausado_em":    info_log.get("pausado_em", "—"),
            "motivo":        info_log.get("motivo", ""),
        })
    return resultado

def buscar_tempo_pausado(pedido, etapa_idx):
    """Retorna os segundos pausados salvos na sessão ativa, ou 0."""
    rows = _get("sessoes_ativas",
        f"pedido=eq.{pedido}&etapa_idx=eq.{etapa_idx}&select=tempo_pausado")
    if rows and isinstance(rows, list) and rows[0].get("tempo_pausado"):
        return int(rows[0]["tempo_pausado"])
    return 0

@st.cache_data(ttl=10, show_spinner=False)
def buscar_pedidos_por_etapa(etapa_idx):
    # Retorna TODOS os pedidos (abertos e concluídos) sem filtrar por etapa.
    # O operador pode selecionar qualquer pedido independente do status ou progresso.
    pedidos_rows = _get("pedidos_base",
        "select=numero,cliente&order=numero.asc",
        paginar=True)
    if not isinstance(pedidos_rows, list):
        return []
    return [(p["numero"], p.get("cliente", "")) for p in pedidos_rows]

@st.cache_data(ttl=5, show_spinner=False)
def buscar_todas_sessoes_ativas():
    rows = _get("sessoes_ativas", "select=*&order=iniciado_em.asc", paginar=True)
    return rows if isinstance(rows, list) else []

def buscar_status_completo_pedido(numero):
    """
    Retorna dict com o status detalhado de cada etapa do pedido:
    {
      "base_status": "aberto"|"concluido"|"nao_encontrado",
      "etapas": [
        {"idx":0, "label":"Separação", "feita":True/False, "em_andamento":False,
         "operador":"...", "tempo":120, "data":"..."},
        ...
      ]
    }
    """
    # Status base
    base_rows = _get("pedidos_base", f"numero=eq.{numero}&select=status,cliente")
    if not base_rows:
        return {"base_status": "nao_encontrado", "cliente": "", "etapas": []}
    base_status = base_rows[0].get("status", "aberto")
    cliente     = base_rows[0].get("cliente", "")

    # Registros finalizados por etapa
    reg_rows = _get("registros",
        f"pedido=eq.{numero}&select=etapa_idx,operador,tempo_segundos,data&order=id.asc")
    regs_por_etapa = {}
    if isinstance(reg_rows, list):
        for r in reg_rows:
            ei = r.get("etapa_idx")
            if ei is not None:
                regs_por_etapa[int(ei)] = r

    # Sessões ativas (em andamento agora)
    sess_rows = _get("sessoes_ativas",
        f"pedido=eq.{numero}&select=etapa_idx,operador,iniciado_em")
    sess_por_etapa = {}
    if isinstance(sess_rows, list):
        for r in sess_rows:
            ei = r.get("etapa_idx")
            if ei is not None:
                sess_por_etapa[int(ei)] = r

    etapas = []
    for idx, lbl in enumerate(ETAPAS_LBL):
        reg  = regs_por_etapa.get(idx)
        sess = sess_por_etapa.get(idx)
        etapas.append({
            "idx":          idx,
            "label":        lbl,
            "feita":        reg is not None,
            "em_andamento": sess is not None,
            "operador":     (reg or sess or {}).get("operador", ""),
            "tempo":        (reg or {}).get("tempo_segundos"),
            "data":         (reg or {}).get("data", ""),
            "iniciado_em":  (sess or {}).get("iniciado_em"),
        })

    # Conclusão real = etapa 2 registrada pelos operadores.
    # Ignora o status do banco (que pode ter vindo como "concluido" pela planilha 100%).
    etapa2_feita = any(e["idx"] == 2 and e.get("feita") for e in etapas)
    base_status  = "concluido" if etapa2_feita else "aberto"

    return {
        "base_status": base_status,
        "cliente":     cliente,
        "etapas":      etapas,
    }

def finalizar_pip(pedido, etapa_idx, operador, iniciado_em):
    tempo = max(int(time.time()) - int(iniciado_em), 1)
    # Busca qtd de peças do pedido para registrar no histórico
    _qtd_pip = None
    try:
        _r = _get("pedidos_base", f"numero=eq.{pedido}&select=est_alocado")
        if isinstance(_r, list) and _r and _r[0].get("est_alocado"):
            _qtd_pip = int(float(_r[0]["est_alocado"]))
    except Exception:
        pass
    salvar(pedido, operador, ETAPAS[etapa_idx], etapa_idx, tempo, _qtd_pip)
    remover_sessao_ativa(pedido, etapa_idx)
    if etapa_idx == 2:
        marcar_concluido(pedido)

def salvar(pedido, operador, etapa, etapa_idx, tempo, qtd_pecas=None):
    fim_dt    = now_br()
    inicio_dt = fim_dt - timedelta(seconds=tempo)
    payload = {
        "pedido": pedido, "operador": operador, "etapa": etapa,
        "etapa_idx": etapa_idx, "tempo_segundos": tempo,
        "data":    fim_dt.strftime("%d/%m/%Y %H:%M"),
        "inicio":  inicio_dt.strftime("%d/%m/%Y %H:%M"),
        "qtd_pecas": qtd_pecas if qtd_pecas else None,
    }
    ok = _post("registros", payload)
    if not ok:
        # Tenta uma segunda vez após 1s para cobrir falhas transitórias de rede
        import time as _t; _t.sleep(1)
        _post("registros", payload)

def buscar():
    rows = _get("registros", "select=*&order=id.desc", paginar=True)
    if isinstance(rows, list):
        return [(r.get("id"), r.get("pedido"), r.get("operador"), r.get("etapa"),
                 r.get("etapa_idx"), r.get("tempo_segundos"), r.get("data"),
                 r.get("inicio"), r.get("qtd_pecas"))
                for r in rows]
    return []

def limpar():
    _delete("registros", "id=gte.0")

def limpar_sessoes_ativas():
    """Remove todas as sessões ativas buscando e deletando uma a uma."""
    rows = _get("sessoes_ativas", "select=pedido,etapa_idx")
    if isinstance(rows, list):
        for r in rows:
            ped = r.get("pedido", "")
            eta = r.get("etapa_idx", 0)
            if ped != "":
                _delete("sessoes_ativas", f"pedido=eq.{ped}&etapa_idx=eq.{eta}")
    # fallback: deleta por campo operador sempre preenchido
    _delete("sessoes_ativas", "operador=neq.___x___")

def buscar_pedidos_avulsos():
    rows = _get(
        "pedidos_base",
        "select=numero,status,importado_em"
        "&cliente=eq."
        "&percentual=is.null"
        "&order=importado_em.desc"
    )
    if isinstance(rows, list):
        return [(r["numero"], r.get("status","aberto"), r.get("importado_em",""))
                for r in rows]
    return []

def excluir_pedido_avulso(numero):
    _delete("sessoes_ativas", f"pedido=eq.{numero}")
    _delete("registros",      f"pedido=eq.{numero}")
    _delete("pedidos_base",   f"numero=eq.{numero}")

init_db()

# ── Limpeza automática de sessões expiradas (>12h) ──
def _limpar_sessoes_expiradas():
    """Remove sessões com mais de 12h buscando e deletando individualmente.
    Sessões pausadas (iniciado_em == 0) NÃO são removidas — ficam indefinidamente."""
    limite = int(time.time()) - 43200
    rows = _get("sessoes_ativas", "select=pedido,etapa_idx,iniciado_em")
    if isinstance(rows, list):
        for r in rows:
            try:
                ini = int(r.get("iniciado_em", 0))
                if ini == 0:
                    continue  # Pausado — não expira automaticamente
                if ini < limite:
                    ped = r.get("pedido", "")
                    eta = r.get("etapa_idx", 0)
                    if ped:
                        _delete("sessoes_ativas", f"pedido=eq.{ped}&etapa_idx=eq.{eta}")
            except Exception:
                pass

# Roda limpeza de sessões expiradas no máximo a cada 30 min por sessão
_agora = int(time.time())
if _agora - st.session_state.get("_ultima_limpeza", 0) > 1800:
    _limpar_sessoes_expiradas()
    st.session_state["_ultima_limpeza"] = _agora

# ─────────────────────────────────────
#  QUERY PARAM — PiP FINALIZAR
# ─────────────────────────────────────
_qp = st.query_params
_pip_action = _qp.get("pip_action", "")

if _pip_action == "finalizar":
    try:
        _ped = _qp.get("pedido", "")
        _eta = int(_qp.get("etapa", 0))
        _op  = _qp.get("operador", "")
        _ini = int(_qp.get("iniciado_em", 0))
        if _ped and _op and _ini:
            finalizar_pip(_ped, _eta, _op, _ini)
    except Exception:
        pass
    st.query_params.clear()
    st.rerun()

elif _pip_action == "fechar":
    # Fecha o PiP sem salvar o tempo — apenas remove a sessão ativa
    try:
        _ped = _qp.get("pedido", "")
        _eta = int(_qp.get("etapa", 0))
        if _ped:
            remover_sessao_ativa(_ped, _eta)
    except Exception:
        pass
    st.query_params.clear()
    st.rerun()

# ─────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────
def fmt(s):
    if s is None: return "—"
    s = int(s)
    if s < 60: return f"{s}s"
    m, sec = divmod(s, 60)
    if m < 60: return f"{m}m {sec:02d}s"
    h, mi = divmod(m, 60)
    return f"{h}h {mi:02d}m"

@st.cache_resource
def logo_b64():
    p = Path(__file__).parent / "logo_vi.png"
    return base64.b64encode(p.read_bytes()).decode() if p.exists() else None

def media(lst):
    lst_clean = [x for x in lst if x is not None]
    return int(sum(lst_clean) / len(lst_clean)) if lst_clean else 0

def get_elapsed():
    if st.session_state.get("rodando") and st.session_state.get("inicio"):
        return st.session_state.acum + int(time.time() - st.session_state.inicio)
    return st.session_state.get("acum", 0)

# ─────────────────────────────────────
#  SESSION STATE
# ─────────────────────────────────────
for k, v in {
    "tela":"home", "operador":None, "pedido":None,
    "etapa_idx":0, "rodando":False, "inicio":None,
    "acum":0, "modal":None,
    "pedido_prox":None, "etapa_prox":None,
    "erro_pedido":False, "erro_senha":False,
    "pedido_status":None, "pedido_confirm":False,
    "etapa_escolhida":None, "duplicata_info":None,
    "pedido_validado":False,
    "op_filtro_andamento": "Todos",
}.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ─────────────────────────────────────
#  CSS
# ─────────────────────────────────────
b64 = logo_b64()
logo_src = f"data:image/png;base64,{b64}" if b64 else ""

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Nunito:wght@600;700;800;900&family=DM+Mono:wght@400;500&display=swap');

html, body, [data-testid="stAppViewContainer"] {{
    background: #F7F5F2 !important;
    font-family: 'Nunito', sans-serif !important;
}}
[data-testid="stHeader"], [data-testid="stSidebar"], footer, #MainMenu {{ display:none !important; }}
.block-container {{ padding-top:1.8rem !important; padding-bottom:2rem !important; max-width:600px !important; margin: 0 auto !important; }}

.logo-box {{ text-align:center; margin-bottom:1.6rem; }}
.logo-box img {{ height:56px; object-fit:contain; }}

.section-label {{
    font-size:11px; font-weight:800; letter-spacing:2.5px; text-transform:uppercase;
    color:#9C9490; margin-bottom:1.2rem; text-align:center;
}}

/* ── STEPPER ── */
.stepper {{ display:flex; align-items:flex-start; margin-bottom:1.6rem; }}
.step {{ flex:1; display:flex; flex-direction:column; align-items:center; gap:6px; }}
.sdot {{
    width:32px; height:32px; border-radius:50%;
    border:2px solid #E0DBD4; background:#EDE9E4;
    display:flex; align-items:center; justify-content:center;
    font-size:13px; font-weight:800; color:#A09890;
}}
.sdot.active {{ background:#C8566A; border-color:#C8566A; color:#fff; box-shadow:0 0 0 5px rgba(200,86,106,0.14); }}
.sdot.done   {{ background:#4A7C59; border-color:#4A7C59; color:#fff; }}
.slbl {{ font-size:10px; font-weight:800; letter-spacing:1px; text-transform:uppercase; color:#A09890; text-align:center; }}
.slbl.active {{ color:#C8566A; }} .slbl.done {{ color:#4A7C59; }}
.sline {{ flex:1; height:2px; background:#E0DBD4; margin-top:14px; }}
.sline.done {{ background:#4A7C59; }}

/* ── CARD ── */
.vi-card {{
    background:#fff; border:1.5px solid #E8E3DC; border-radius:16px;
    padding:28px; box-shadow:0 4px 20px rgba(0,0,0,0.06); margin-bottom:1rem;
}}

/* ── INPUT ── */
.stTextInput > div > div > input {{
    border: 2px solid #E0DBD4 !important;
    border-radius: 12px !important;
    background: #FFFFFF !important;
    font-family: 'Nunito', sans-serif !important;
    font-size: 16px !important;
    font-weight: 700 !important;
    padding: 16px 18px !important;
    color: #1A1714 !important;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06) !important;
    transition: border-color .2s, box-shadow .2s !important;
    height: 56px !important;
}}
.stTextInput > div > div > input:focus {{
    border-color: #C8566A !important;
    box-shadow: 0 0 0 4px rgba(200,86,106,0.12) !important;
}}
.stTextInput > div > div > input::placeholder {{
    color: #C0BAB4 !important; font-weight: 600 !important;
}}
label {{
    font-family: 'Nunito', sans-serif !important;
    font-size: 11px !important; font-weight: 800 !important;
    color: #9C9490 !important; letter-spacing: 1.5px !important;
    text-transform: uppercase !important;
}}

/* ── BUTTONS ── */
.stButton > button {{
    font-family: 'Nunito', sans-serif !important;
    font-weight: 800 !important;
    border-radius: 12px !important;
    transition: all .18s ease !important;
    height: 54px !important;
    font-size: 15px !important;
    letter-spacing: .5px !important;
}}
.btn-iniciar > button {{
    background: linear-gradient(135deg, #2C6E49, #1E4D35) !important;
    color: #fff !important; border: none !important;
    box-shadow: 0 5px 0 rgba(20,50,30,0.45), 0 8px 20px rgba(44,110,73,0.30) !important;
}}
.btn-iniciar > button:hover {{
    background: linear-gradient(135deg, #357a54, #266040) !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 0 rgba(20,50,30,0.40), 0 14px 28px rgba(44,110,73,0.35) !important;
}}
.btn-iniciar > button:active {{
    transform: translateY(2px) !important;
    box-shadow: 0 2px 0 rgba(20,50,30,0.45), 0 3px 8px rgba(44,110,73,0.20) !important;
}}
.btn-voltar > button {{
    background: #FFFFFF !important;
    color: #5C5450 !important;
    border: 2px solid #DDD8D2 !important;
    box-shadow: 0 3px 0 rgba(0,0,0,0.10), 0 4px 12px rgba(0,0,0,0.07) !important;
}}
.btn-voltar > button:hover {{
    border-color: #C8566A !important; color: #C8566A !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 5px 0 rgba(0,0,0,0.08), 0 8px 16px rgba(200,86,106,0.12) !important;
}}
.btn-finalizar > button {{
    background: linear-gradient(135deg, #C8566A, #9E3F52) !important;
    color: #fff !important; border: none !important;
    box-shadow: 0 5px 0 rgba(100,20,35,0.45), 0 8px 20px rgba(200,86,106,0.32) !important;
}}
.btn-finalizar > button:hover {{
    background: linear-gradient(135deg, #d9617a, #b04560) !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 0 rgba(100,20,35,0.40), 0 14px 28px rgba(200,86,106,0.38) !important;
}}
.btn-finalizar > button:active {{
    transform: translateY(2px) !important;
    box-shadow: 0 2px 0 rgba(100,20,35,0.45) !important;
}}
.btn-primary > button {{
    background: linear-gradient(135deg, #C8566A, #9E3F52) !important;
    color: #fff !important; border: none !important;
    box-shadow: 0 5px 0 rgba(100,20,35,0.40), 0 8px 18px rgba(200,86,106,0.28) !important;
}}
.btn-primary > button:hover {{
    background: linear-gradient(135deg, #d9617a, #b04560) !important;
    transform: translateY(-2px) !important;
}}
.btn-outline > button {{
    background: #fff !important; color: #5C5450 !important;
    border: 2px solid #DDD8D2 !important;
    box-shadow: 0 3px 0 rgba(0,0,0,0.08) !important;
}}
.btn-outline > button:hover {{ border-color:#C8566A !important; color:#C8566A !important; }}
.btn-sm > button {{ height:40px !important; font-size:12px !important; }}

/* ── TIMER ── */
.timer-wrap {{
    background: linear-gradient(135deg, #fff, #fdf5f7);
    border: 2px solid #F0E0E4;
    border-radius: 16px;
    padding: 20px;
    text-align: center;
    margin-bottom: 20px;
    box-shadow: 0 4px 20px rgba(200,86,106,0.08);
}}
.timer-num {{
    font-family:'DM Mono',monospace; font-size:62px; font-weight:500;
    color:#C8566A; letter-spacing:-2px; line-height:1;
}}
.pedido-lbl {{ font-size:10px; font-weight:800; color:#9C9490; text-align:center; letter-spacing:2px; text-transform:uppercase; margin-bottom:4px; }}
.pedido-num {{ font-family:'DM Mono',monospace; font-size:24px; font-weight:500; color:#1A1714; text-align:center; margin-bottom:6px; }}

/* ── ADMIN ── */
.stat-box {{ background:#fff; border:1.5px solid #E8E3DC; border-radius:14px; padding:18px; text-align:center; box-shadow:0 2px 12px rgba(0,0,0,0.05); }}
.stat-num {{ font-family:'DM Mono',monospace; font-size:32px; font-weight:500; color:#C8566A; }}
.stat-lbl {{ font-size:11px; font-weight:800; color:#9C9490; letter-spacing:.8px; margin-top:5px; text-transform:uppercase; }}
table {{ width:100%; border-collapse:collapse; font-size:13px; }}
th {{ text-align:left; padding:10px 12px; font-size:10px; font-weight:800; letter-spacing:1.2px; text-transform:uppercase; color:#9C9490; border-bottom:2px solid #EDE9E4; }}
td {{ padding:11px 12px; border-bottom:1px solid #F2EEE9; color:#2C2826; font-weight:600; }}
.tag {{ display:inline-block; padding:3px 10px; border-radius:100px; font-size:10px; font-weight:800; letter-spacing:.5px; }}
.tag-sep {{ background:#EBF0FB; color:#3B5EC6; }}
.tag-conf {{ background:#FBF2E6; color:#C47B2A; }}
.tag-emb  {{ background:#E8F2EC; color:#4A7C59; }}

/* ── EXPANDER — força texto escuro independente do tema ── */
[data-testid="stExpander"] details summary p {{
    color: #1A1714 !important;
    font-weight: 800 !important;
    font-size: 13px !important;
}}
[data-testid="stExpander"] details summary svg {{
    fill: #5C5450 !important;
}}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────
#  BOTÃO DE RECUPERAÇÃO DE ERRO
#  Aparece fixo no canto sempre que o Streamlit exibe tela de erro.
#  Monitora via JS a presença do bloco de erro e revela o botão.
# ─────────────────────────────────────
st.markdown("""
<style>
#vi-recovery-btn {
    display: none;
    position: fixed;
    bottom: 24px;
    left: 50%;
    transform: translateX(-50%);
    z-index: 99999;
    background: linear-gradient(135deg, #C8566A, #a03050);
    color: #fff;
    font-family: 'Nunito', sans-serif;
    font-size: 15px;
    font-weight: 900;
    padding: 14px 32px;
    border-radius: 50px;
    border: none;
    cursor: pointer;
    box-shadow: 0 6px 24px rgba(200,86,106,0.45);
    letter-spacing: 0.3px;
    white-space: nowrap;
}
#vi-recovery-btn:hover {
    filter: brightness(1.1);
    transform: translateX(-50%) translateY(-2px);
}
</style>
<button id="vi-recovery-btn" onclick="window.location.href=window.location.pathname">
  ↩ Voltar ao Lobby
</button>
<script>
(function() {
    function checkError() {
        // Streamlit exibe erro em [data-testid="stException"] ou div com class que contém "stException"
        var errEl = document.querySelector('[data-testid="stException"]') ||
                    document.querySelector('.stException') ||
                    document.querySelector('[class*="exception"]') ||
                    document.querySelector('[class*="Exception"]');
        var btn = document.getElementById('vi-recovery-btn');
        if (btn) {
            btn.style.display = errEl ? 'block' : 'none';
        }
    }
    // Verifica imediatamente e a cada 800ms
    checkError();
    setInterval(checkError, 800);
})();
</script>
""", unsafe_allow_html=True)

# ─────────────────────────────────────
#  RENDER HELPERS
# ─────────────────────────────────────
def render_logo():
    if logo_src:
        st.markdown(f'<div class="logo-box"><img src="{logo_src}"></div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="logo-box" style="font-size:22px;font-weight:700;color:#C8566A;">Vi Lingerie</div>', unsafe_allow_html=True)

def render_stepper(idx):
    html = '<div class="stepper">'
    for i, lbl in enumerate(ETAPAS_LBL):
        dc = "done" if i < idx else ("active" if i == idx else "")
        html += f'<div class="step"><div class="sdot {dc}">{i+1}</div><div class="slbl {dc}">{lbl}</div></div>'
        if i < 2:
            html += f'<div class="sline {"done" if i < idx else ""}"></div>'
    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)

def render_avatar_grid(on_click_key="home"):
    import streamlit.components.v1 as _cv1

    COLORS = [
        "#C8566A","#3B7DD8","#4A7C59","#E07B3A",
        "#7C5CBF","#B85C38","#2E9E8F","#8E6BBF","#C8566A",
    ]

    selecionado = st.session_state.get("operador")
    if selecionado and selecionado in OPERADORES:
        idx_op = OPERADORES.index(selecionado)
        cor    = COLORS[idx_op % len(COLORS)]
        ini    = (selecionado[0]+selecionado[1]).upper()
    else:
        cor, ini = "#9C9490", "?"

    _cv1.html(f"""<!DOCTYPE html><html><head>
<link href="https://fonts.googleapis.com/css2?family=Nunito:wght@700;800;900&display=swap" rel="stylesheet">
<style>
*{{margin:0;padding:0;box-sizing:border-box;}}
body{{background:transparent;font-family:'Nunito',sans-serif;overflow:visible;}}
.card{{
    display:flex;align-items:center;gap:14px;
    background:#fff;border:2px solid #E8E2DC;
    border-radius:16px 16px 0 0;
    padding:14px 18px 12px;
    box-shadow:0 2px 12px rgba(0,0,0,.06);
    border-bottom: none;
}}
.av{{width:44px;height:44px;border-radius:50%;flex-shrink:0;
    background:linear-gradient(145deg,{cor},{cor}bb);
    display:flex;align-items:center;justify-content:center;
    font-size:15px;font-weight:900;color:#fff;
    box-shadow:0 3px 10px rgba(0,0,0,.20);}}
.label{{font-size:10px;font-weight:800;letter-spacing:1.8px;
    text-transform:uppercase;color:#9C9490;margin-bottom:3px;}}
.value{{font-size:16px;font-weight:900;color:#1A1714;}}
.hint{{font-size:11px;color:#9C9490;margin-top:2px;font-weight:600;}}
.arrow{{margin-left:auto;font-size:18px;color:#C8566A;}}
</style></head>
<body>
<div class="card">
    <div class="av">{ini}</div>
    <div>
        <div class="label">Operador</div>
        <div class="value">{"Nenhum selecionado" if not selecionado else selecionado}</div>
        <div class="hint">{"Escolha abaixo ↓" if not selecionado else "✓ Selecionado"}</div>
    </div>
    <div class="arrow">⌄</div>
</div>
</body></html>""", height=96, scrolling=False)

    st.markdown("""
    <style>
    div[data-testid="stSelectbox"] label { display:none !important; }
    div[data-testid="stSelectbox"] > div > div {
        border: 2px solid #E8E2DC !important;
        border-radius: 16px !important;
        background: #fff !important;
        min-height: 56px !important;
        padding: 0 16px !important;
        font-family: 'Nunito', sans-serif !important;
        font-size: 16px !important;
        font-weight: 800 !important;
        color: #1A1714 !important;
        box-shadow: 0 4px 16px rgba(0,0,0,.07) !important;
        display: flex !important;
        align-items: center !important;
    }
    div[data-testid="stSelectbox"] > div > div > div {
        white-space: nowrap !important;
        overflow: visible !important;
        text-overflow: unset !important;
        padding: 0 !important;
        min-height: unset !important;
        line-height: 1.3 !important;
    }
    div[data-testid="stSelectbox"] > div > div:focus-within {
        border-color: #C8566A !important;
        box-shadow: 0 4px 16px rgba(200,86,106,.15) !important;
    }
    div[data-testid="stSelectbox"] ul {
        background: #fff !important;
        border: 2px solid #C8566A !important;
        border-radius: 12px !important;
        padding: 6px !important;
        box-shadow: 0 16px 40px rgba(0,0,0,.14) !important;
        font-family: 'Nunito', sans-serif !important;
    }
    div[data-testid="stSelectbox"] ul li {
        border-radius: 8px !important;
        font-weight: 800 !important;
        font-size: 15px !important;
        padding: 12px 16px !important;
        white-space: nowrap !important;
    }
    div[data-testid="stSelectbox"] ul li:hover {
        background: #FFF0F2 !important;
        color: #C8566A !important;
    }
    </style>
    """, unsafe_allow_html=True)

    opcoes  = ["— Selecione o operador —"] + OPERADORES
    idx_cur = 0
    if selecionado and selecionado in OPERADORES:
        idx_cur = OPERADORES.index(selecionado) + 1

    escolha = st.selectbox("op", opcoes,
                           index=idx_cur,
                           key=f"op_sel_{on_click_key}",
                           label_visibility="collapsed")

    if escolha and escolha != "— Selecione o operador —":
        if st.session_state.get("operador") != escolha:
            st.session_state.operador = escolha
            st.rerun()


# ─────────────────────────────────────
#  PiP — JANELA FLUTUANTE
# ─────────────────────────────────────
ETAPA_CORES = ["#C8566A", "#3B7DD8", "#4A7C59"]
ETAPA_ICONS = ["📦", "🗃️", "✅"]

def render_pip():
    """Renderiza janelas PiP flutuantes — só aparece para sessões com cronômetro ativo."""
    sessoes = buscar_todas_sessoes_ativas()
    if not sessoes:
        return

    cards_html = ""
    cards_js   = ""

    for s in sessoes:
        ped     = s.get("pedido", "")
        op      = s.get("operador", "")
        eta_idx = int(s.get("etapa_idx", 0))
        ini     = int(s.get("iniciado_em", 0))
        cor     = ETAPA_CORES[eta_idx]
        icon    = ETAPA_ICONS[eta_idx]
        lbl     = ETAPAS_LBL[eta_idx]
        uid     = f"{ped}_{eta_idx}"

        cards_html += f"""
        <div class="pip-card" id="pip-card-{uid}" style="border-top:3px solid {cor};">
          <!-- Header -->
          <div class="pip-drag-handle" id="pip-handle-{uid}">
            <span style="font-size:11px;opacity:0.5;flex-shrink:0;">⠿</span>
            <span style="font-size:10px;font-weight:800;letter-spacing:1px;opacity:0.7;flex:1;text-align:center;text-transform:uppercase;">{icon} {lbl}</span>
            <span class="pip-minimize" onclick="togglePip('{uid}')" title="Minimizar">─</span>
            <span class="pip-close-btn" onclick="pedirFechamento('{uid}','{ped}',{eta_idx},'{op}',{ini})" title="Fechar PiP">✕</span>
          </div>
          <!-- Body normal -->
          <div class="pip-body" id="pip-body-{uid}">
            <div style="font-size:11px;opacity:0.6;margin-bottom:2px;">Pedido <strong style="opacity:1;color:#fff;">{ped}</strong> · {op}</div>
            <div class="pip-timer" id="pip-timer-{uid}">00:00:00</div>
            <button class="pip-btn-fin" onclick="finalizarPip('{ped}',{eta_idx},'{op}',{ini})">&#9632; FINALIZAR</button>
          </div>
          <!-- Popup de confirmação (oculto por padrão) -->
          <div class="pip-confirm" id="pip-confirm-{uid}">
            <div class="pip-confirm-icon">⚠️</div>
            <div class="pip-confirm-msg">Fechar o PiP <strong>descarta o tempo</strong> sem salvar.<br>Tem certeza?</div>
            <div class="pip-confirm-btns">
              <button class="pip-confirm-yes" onclick="fecharPip('{uid}','{ped}',{eta_idx})">Sim, fechar</button>
              <button class="pip-confirm-no"  onclick="cancelarFechamento('{uid}')">Cancelar</button>
            </div>
          </div>
        </div>"""

        cards_js += f"startTimer('{uid}', {ini});\n"

    components.html(f"""
    <script>
    (function() {{
        var pd = window.parent.document;

        if (pd.getElementById('pip-container')) {{
            pd.getElementById('pip-container').remove();
        }}
        if (pd.getElementById('pip-styles')) {{
            pd.getElementById('pip-styles').remove();
        }}

        var style = pd.createElement('style');
        style.id = 'pip-styles';
        style.textContent = `
            #pip-container {{
                position: fixed;
                bottom: 24px;
                right: 24px;
                z-index: 99999;
                display: flex;
                flex-direction: column;
                gap: 10px;
                pointer-events: none;
            }}
            .pip-card {{
                pointer-events: all;
                background: rgba(26,23,20,0.94);
                border-radius: 14px;
                box-shadow: 0 8px 32px rgba(0,0,0,0.45);
                overflow: hidden;
                resize: both;
                min-width: 210px;
                min-height: 110px;
                width: 260px;
                backdrop-filter: blur(8px);
            }}
            .pip-drag-handle {{
                padding: 8px 12px;
                background: rgba(255,255,255,0.06);
                cursor: grab;
                display: flex;
                align-items: center;
                gap: 6px;
                color: rgba(255,255,255,0.65);
                user-select: none;
                font-family: 'Nunito', sans-serif;
            }}
            .pip-drag-handle:active {{ cursor: grabbing; }}
            .pip-minimize {{
                cursor: pointer;
                font-size: 14px;
                padding: 0 4px;
                opacity: 0.6;
            }}
            .pip-minimize:hover {{ opacity: 1; }}
            .pip-body {{
                padding: 10px 14px 14px;
                font-family: 'Nunito', sans-serif;
                color: #fff;
            }}
            .pip-timer {{
                font-family: 'DM Mono', 'Courier New', monospace;
                font-size: 34px;
                font-weight: 500;
                color: #fff;
                letter-spacing: -1px;
                margin: 6px 0 10px;
                line-height: 1;
            }}
            .pip-btn-fin {{
                width: 100%;
                background: #C8566A;
                color: #fff;
                border: none;
                border-radius: 8px;
                padding: 9px 0;
                font-family: 'Nunito', sans-serif;
                font-size: 13px;
                font-weight: 800;
                cursor: pointer;
                letter-spacing: 0.5px;
            }}
            .pip-btn-fin:hover {{ background: #a83050; }}

            /* ── Botão X fechar ── */
            .pip-close-btn {{
                cursor: pointer;
                font-size: 13px;
                padding: 0 2px 0 6px;
                opacity: 0.45;
                color: #fff;
                flex-shrink: 0;
                line-height: 1;
                transition: opacity .15s, color .15s;
            }}
            .pip-close-btn:hover {{ opacity: 1; color: #FF6B6B; }}

            /* ── Popup de confirmação ── */
            .pip-confirm {{
                display: none;
                padding: 16px 14px 14px;
                font-family: 'Nunito', sans-serif;
                color: #fff;
                text-align: center;
                animation: fadeIn .15s ease;
            }}
            @keyframes fadeIn {{ from{{opacity:0;transform:translateY(4px)}} to{{opacity:1;transform:translateY(0)}} }}
            .pip-confirm-icon {{ font-size: 26px; margin-bottom: 8px; }}
            .pip-confirm-msg {{
                font-size: 12px; font-weight: 600; color: rgba(255,255,255,0.8);
                line-height: 1.55; margin-bottom: 14px;
            }}
            .pip-confirm-msg strong {{ color: #FF6B6B; }}
            .pip-confirm-btns {{ display: flex; gap: 8px; }}
            .pip-confirm-yes {{
                flex: 1; background: rgba(200,86,106,0.25); color: #FF8FA3;
                border: 1.5px solid rgba(200,86,106,0.5); border-radius: 8px;
                padding: 8px 0; font-family: 'Nunito', sans-serif;
                font-size: 12px; font-weight: 800; cursor: pointer;
                transition: background .15s;
            }}
            .pip-confirm-yes:hover {{ background: #C8566A; color: #fff; border-color: #C8566A; }}
            .pip-confirm-no {{
                flex: 1; background: rgba(255,255,255,0.07); color: rgba(255,255,255,0.65);
                border: 1.5px solid rgba(255,255,255,0.15); border-radius: 8px;
                padding: 8px 0; font-family: 'Nunito', sans-serif;
                font-size: 12px; font-weight: 800; cursor: pointer;
                transition: background .15s;
            }}
            .pip-confirm-no:hover {{ background: rgba(255,255,255,0.14); color: #fff; }}
        `;
        pd.head.appendChild(style);

        var container = pd.createElement('div');
        container.id = 'pip-container';
        container.innerHTML = `{cards_html}`;
        pd.body.appendChild(container);

        window.parent.startTimer = function(uid, iniciado_em) {{
            var el = pd.getElementById('pip-timer-' + uid);
            if (!el) return;
            function update() {{
                var elapsed = Math.floor(Date.now() / 1000) - iniciado_em;
                var h = Math.floor(elapsed / 3600);
                var m = Math.floor((elapsed % 3600) / 60);
                var s = elapsed % 60;
                el.textContent =
                    String(h).padStart(2,'0') + ':' +
                    String(m).padStart(2,'0') + ':' +
                    String(s).padStart(2,'0');
            }}
            update();
            setInterval(update, 1000);
        }};

        window.parent.finalizarPip = function(pedido, etapa, operador, iniciado_em) {{
            var url = new URL(window.parent.location.href);
            url.searchParams.set('pip_action', 'finalizar');
            url.searchParams.set('pedido', pedido);
            url.searchParams.set('etapa', etapa);
            url.searchParams.set('operador', operador);
            url.searchParams.set('iniciado_em', iniciado_em);
            window.parent.location.href = url.toString();
        }};

        window.parent.togglePip = function(uid) {{
            var body = pd.getElementById('pip-body-' + uid);
            if (!body) return;
            body.style.display = body.style.display === 'none' ? 'block' : 'none';
        }};

        /* Abre o popup de confirmação de fechamento */
        window.parent.pedirFechamento = function(uid, pedido, etapa, operador, iniciado_em) {{
            var body    = pd.getElementById('pip-body-' + uid);
            var confirm = pd.getElementById('pip-confirm-' + uid);
            if (!body || !confirm) return;
            body.style.display    = 'none';
            confirm.style.display = 'block';
        }};

        /* Usuário confirmou — remove do DOM + chama endpoint para deletar sessão */
        window.parent.fecharPip = function(uid, pedido, etapa) {{
            var card = pd.getElementById('pip-card-' + uid);
            if (card) {{
                card.style.transition = 'opacity .2s, transform .2s';
                card.style.opacity = '0';
                card.style.transform = 'scale(0.92)';
                setTimeout(function() {{ if (card) card.remove(); }}, 220);
            }}
            // Redireciona com ação de fechar (sem salvar tempo)
            var url = new URL(window.parent.location.href);
            url.searchParams.set('pip_action', 'fechar');
            url.searchParams.set('pedido', pedido);
            url.searchParams.set('etapa', etapa);
            window.parent.location.href = url.toString();
        }};

        /* Usuário cancelou — volta ao body normal */
        window.parent.cancelarFechamento = function(uid) {{
            var body    = pd.getElementById('pip-body-' + uid);
            var confirm = pd.getElementById('pip-confirm-' + uid);
            if (!body || !confirm) return;
            confirm.style.display = 'none';
            body.style.display    = 'block';
        }};

        pd.querySelectorAll('.pip-drag-handle').forEach(function(handle) {{
            var card = handle.closest('.pip-card');
            var dragging = false, sx, sy, ox, oy;
            handle.addEventListener('mousedown', function(e) {{
                dragging = true;
                sx = e.clientX; sy = e.clientY;
                var rect = card.getBoundingClientRect();
                ox = rect.left; oy = rect.top;
                card.style.position = 'fixed';
                card.style.left = ox + 'px';
                card.style.top  = oy + 'px';
                card.style.margin = '0';
                e.preventDefault();
            }});
            pd.addEventListener('mousemove', function(e) {{
                if (!dragging) return;
                card.style.left = (ox + e.clientX - sx) + 'px';
                card.style.top  = (oy + e.clientY - sy) + 'px';
            }});
            pd.addEventListener('mouseup', function() {{ dragging = false; }});
        }});

        pd.querySelectorAll('.pip-drag-handle').forEach(function(handle) {{
            var card = handle.closest('.pip-card');
            var ox, oy, sx, sy;
            handle.addEventListener('touchstart', function(e) {{
                var t = e.touches[0];
                sx = t.clientX; sy = t.clientY;
                var rect = card.getBoundingClientRect();
                ox = rect.left; oy = rect.top;
                card.style.position = 'fixed';
                card.style.left = ox + 'px';
                card.style.top  = oy + 'px';
                card.style.margin = '0';
            }}, {{passive:true}});
            handle.addEventListener('touchmove', function(e) {{
                var t = e.touches[0];
                card.style.left = (ox + t.clientX - sx) + 'px';
                card.style.top  = (oy + t.clientY - sy) + 'px';
                e.preventDefault();
            }}, {{passive:false}});
        }});

        {cards_js}
    }})();
    </script>
    """, height=0, scrolling=False)


# ─────────────────────────────────────
#  TELA: HOME
# ─────────────────────────────────────
def _render_status_pedido(num, status, etapa_idx):
    num    = st.session_state.get("_pedido_validando", "")
    status = st.session_state.get("_status_cache") or buscar_status_completo_pedido(num)
    render_stepper(etapa_idx)

    base_st   = status["base_status"]
    cliente   = status.get("cliente", "")
    etapas    = status["etapas"]
    etapa_lbl = ETAPAS_LBL[etapa_idx]
    # Guarda defensiva: etapa_idx pode vir corrompido do session_state (ex: troca de planilha)
    if not etapas or etapa_idx < 0 or etapa_idx >= len(etapas):
        st.session_state["_status_cache"] = None
        st.session_state.etapa_idx        = 0
        st.session_state.pedido           = None
        st.session_state.pedido_validado  = False
        st.session_state.pedido_status    = None
        st.warning("⚠️ Pedido não encontrado ou dados desatualizados. Tente buscar novamente.")
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
        if st.button("← Voltar ao Lobby", use_container_width=True, key="_recovery_voltar"):
            st.session_state.tela          = "home"
            st.session_state.etapa_escolhida = None
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
        st.stop()
    etapa_info = etapas[etapa_idx]  # dados da etapa que foi selecionada

    COR = ["#C8566A", "#3B7DD8", "#4A7C59"]

    # ── Header do pedido ─────────────────────────────────────────────────
    cor_h = COR[etapa_idx]
    import streamlit.components.v1 as _cv1
    _cv1.html(f"""
    <style>*{{margin:0;padding:0;box-sizing:border-box;}}
    body{{font-family:'Nunito',sans-serif;background:transparent;}}</style>
    <div style="background:linear-gradient(135deg,{cor_h},{cor_h}bb);
                border-radius:14px;padding:14px 18px;display:flex;
                align-items:center;gap:14px;">
      <div>
        <div style="font-size:9px;font-weight:800;letter-spacing:2px;
             color:rgba(255,255,255,0.55);text-transform:uppercase;">Pedido</div>
        <div style="font-size:22px;font-weight:900;color:#fff;
             font-family:monospace;">#{num}</div>
        {(f'<div style="font-size:12px;color:rgba(255,255,255,0.7);margin-top:1px;">{cliente}</div>') if cliente else ""}
      </div>
      <div style="margin-left:auto;text-align:right;">
        <div style="font-size:9px;font-weight:800;letter-spacing:2px;
             color:rgba(255,255,255,0.55);text-transform:uppercase;">Etapa atual</div>
        <div style="font-size:14px;font-weight:800;color:#fff;">{etapa_lbl}</div>
      </div>
    </div>""", height=74, scrolling=False)

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    # ── CASO 1: Pedido não encontrado ────────────────────────────────────
    if base_st == "nao_encontrado":
        _cv1.html("""
        <div style="background:#FFFBEB;border:2px solid #F59E0B;border-radius:14px;
                    padding:20px;text-align:center;font-family:sans-serif;">
          <div style="font-size:28px;margin-bottom:8px;">❓</div>
          <div style="font-size:14px;font-weight:800;color:#92400E;">Pedido não encontrado na base</div>
          <div style="font-size:12px;color:#B45309;margin-top:4px;font-weight:600;">
            Deseja cadastrá-lo como pedido avulso?</div>
        </div>""", height=110, scrolling=False)
        ca, cb = st.columns(2)
        with ca:
            st.markdown('<div class="btn-iniciar">', unsafe_allow_html=True)
            if st.button("✓ Cadastrar e Continuar", use_container_width=True):
                cadastrar_pedido_avulso(num)
                buscar_pedidos_por_etapa.clear()
                st.session_state.pedido          = num
                st.session_state.pedido_status   = None
                st.session_state.pedido_validado = True
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
        with cb:
            st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
            if st.button("✕ Cancelar", use_container_width=True):
                st.session_state.pedido_status = None; st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
        return

    # ── Timeline resumida das 3 etapas ───────────────────────────────────
    tl_html = '<div style="display:flex;gap:0;margin-bottom:12px;">'
    for e in etapas:
        if e["feita"]:
            bg_c, txt_c, ic = "#4A7C59", "#fff", "✓"
        elif e["em_andamento"]:
            bg_c, txt_c, ic = "#E07B3A", "#fff", "⏱"
        else:
            bg_c, txt_c, ic = "#EDE9E4", "#9C9490", str(e["idx"]+1)
        brd = "3px solid #1A1714" if e["idx"] == etapa_idx else "none"
        tl_html += f'''<div style="flex:1;background:{bg_c};padding:8px 4px;
            text-align:center;outline:{brd};position:relative;">
          <div style="font-size:16px;">{ic}</div>
          <div style="font-size:9px;font-weight:800;color:{txt_c};
               opacity:0.85;margin-top:2px;letter-spacing:.5px;">{e["label"].split()[0].upper()}</div>
        </div>'''
    tl_html += '</div>'
    _cv1.html(f"""<style>*{{margin:0;padding:0;box-sizing:border-box;}}
    body{{font-family:sans-serif;background:transparent;}}</style>
    <div style="border-radius:12px;overflow:hidden;border:1.5px solid #EDE9E4;">
      {tl_html}
    </div>""", height=68, scrolling=False)

    # ── CASO 2: Pedido concluído pelos operadores (3 etapas feitas) ─────
    if base_st == "concluido":
        _cv1.html("""
        <div style="background:#F0F7F3;border:2px solid #4A7C59;border-radius:14px;
                    padding:20px;text-align:center;font-family:sans-serif;">
          <div style="font-size:28px;margin-bottom:8px;">🎉</div>
          <div style="font-size:14px;font-weight:800;color:#2d5a3d;">Pedido concluído pela produção!</div>
          <div style="font-size:12px;color:#4A7C59;margin-top:4px;font-weight:600;">
            As 3 etapas do Sistema B foram finalizadas pelos operadores.</div>
        </div>""", height=110, scrolling=False)
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
        if st.button("← Buscar outro pedido", use_container_width=True):
            st.session_state.pedido_status = None; st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
        return

    # ── CASO 3: Esta etapa está EM ANDAMENTO / pausada ─────────────────
    if etapa_info["em_andamento"]:
        op_and      = etapa_info["operador"]
        ini_ts      = etapa_info["iniciado_em"]
        # Soma tempo_pausado (caso operador tenha saído e voltado)
        tp          = buscar_tempo_pausado(num, etapa_idx)
        desde_ini   = max(int(time.time()) - int(ini_ts), 0) if ini_ts else 0
        total_s     = tp + desde_ini
        h_t, r_t    = divmod(total_s, 3600); m_t, s_t = divmod(r_t, 60)
        elapsed_str = f"{h_t:02d}:{m_t:02d}:{s_t:02d}"
        retomando   = tp > 0   # veio de uma pausa anterior
        aviso_retomada = (
            f'<div style="font-size:11px;color:#B45309;font-weight:600;'
            f'background:rgba(245,158,11,0.1);border-radius:6px;padding:6px;'
            f'text-align:center;margin-top:8px;">'
            f'⏸ Tempo anterior salvo: {fmt(tp)} — continuando de onde parou.</div>'
        ) if retomando else ""
        _cv1.html(f"""
        <style>*{{margin:0;padding:0;box-sizing:border-box;}}
        body{{font-family:'Nunito',sans-serif;background:transparent;}}</style>
        <div style="background:#FFF8F0;border:2px solid #E07B3A;border-radius:14px;padding:18px 20px;">
          <div style="display:flex;align-items:center;gap:10px;margin-bottom:10px;">
            <div style="font-size:24px;">⏱</div>
            <div>
              <div style="font-size:14px;font-weight:900;color:#9A3412;">Em andamento</div>
              <div style="font-size:12px;color:#C2410C;font-weight:600;">
                Etapa: <strong>{etapa_lbl}</strong> &nbsp;·&nbsp; Operador: <strong>{op_and}</strong></div>
            </div>
            <div style="margin-left:auto;font-family:monospace;font-size:26px;
                 font-weight:500;color:#E07B3A;">{elapsed_str}</div>
          </div>
          {aviso_retomada}
          <div style="font-size:13px;color:#7C2D12;font-weight:700;text-align:center;
               background:rgba(224,123,58,0.08);border-radius:8px;padding:8px;margin-top:8px;">
            Deseja retomar o cronômetro ou finalizar este pedido?
          </div>
        </div>""", height=150 if retomando else 130, scrolling=False)
        ca, cb, cc = st.columns(3)
        with ca:
            st.markdown('<div class="btn-iniciar">', unsafe_allow_html=True)
            if st.button("▶ Retomar", use_container_width=True, key="caso3_retomar"):
                # Restaura o estado local com o tempo acumulado
                st.session_state.pedido          = num
                st.session_state.operador        = op_and
                st.session_state.pedido_validado = True
                st.session_state.pedido_status   = None
                st.session_state.rodando         = True
                st.session_state.inicio          = time.time()
                st.session_state.acum            = tp   # retoma do tempo pausado
                # Reativa a sessão no banco (atualiza iniciado_em para sair do estado pausado)
                _patch("sessoes_ativas",
                       f"pedido=eq.{num}&etapa_idx=eq.{eta_idx}",
                       {"iniciado_em": int(time.time())})
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
        with cb:
            st.markdown('<div class="btn-finalizar">', unsafe_allow_html=True)
            if st.button("■ Finalizar", use_container_width=True, key="caso3_finalizar"):
                # Usa o tempo total (pausado + desde último início)
                tempo = max(total_s, 1)
                salvar(num, op_and, ETAPAS[etapa_idx], etapa_idx, tempo)
                remover_sessao_ativa(num, etapa_idx)
                if etapa_idx == 2:
                    marcar_concluido(num)
                st.toast(f"✅ {op_and} · Pedido #{num} finalizado em {fmt(tempo)}!", icon="🎉")
                st.session_state.pedido_status = None
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
        with cc:
            st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
            if st.button("✕ Cancelar", use_container_width=True, key="caso3_cancelar"):
                st.session_state.pedido_status = None; st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
        return

    # ── CASO 4: Esta etapa já foi finalizada ─────────────────────────────
    if etapa_info["feita"]:
        op_fin   = etapa_info["operador"]
        tempo_fin = fmt(etapa_info["tempo"]) if etapa_info["tempo"] else "—"
        data_fin  = etapa_info["data"] or ""
        prox_idx  = etapa_idx + 1
        tem_prox  = prox_idx < len(ETAPAS_LBL)
        prox_lbl  = ETAPAS_LBL[prox_idx] if tem_prox else ""
        _cv1.html(f"""
        <style>*{{margin:0;padding:0;box-sizing:border-box;}}
        body{{font-family:'Nunito',sans-serif;background:transparent;}}</style>
        <div style="background:#F0F7F3;border:2px solid #4A7C59;border-radius:14px;padding:18px 20px;">
          <div style="display:flex;align-items:center;gap:10px;margin-bottom:10px;">
            <div style="font-size:24px;">✅</div>
            <div>
              <div style="font-size:14px;font-weight:900;color:#2d5a3d;">Etapa já finalizada</div>
              <div style="font-size:12px;color:#4A7C59;font-weight:600;">
                <strong>{etapa_lbl}</strong> &nbsp;·&nbsp; por <strong>{op_fin}</strong>
                &nbsp;·&nbsp; {tempo_fin}</div>
            </div>
            <div style="margin-left:auto;font-size:11px;color:#9C9490;text-align:right;">{data_fin}</div>
          </div>
          {(f'<div style="font-size:13px;color:#2d5a3d;font-weight:700;text-align:center;background:rgba(74,124,89,0.08);border-radius:8px;padding:8px;">Deseja ir para a próxima etapa: <strong>{prox_lbl}</strong>?</div>') if tem_prox else
           '<div style="font-size:13px;color:#2d5a3d;font-weight:700;text-align:center;background:rgba(74,124,89,0.08);border-radius:8px;padding:8px;">Este pedido já passou por todas as etapas.</div>'}
        </div>""", height=130, scrolling=False)
        if tem_prox:
            ca, cb = st.columns(2)
            with ca:
                st.markdown('<div class="btn-iniciar">', unsafe_allow_html=True)
                if st.button(f"▶  Sim, ir para {prox_lbl}", use_container_width=True):
                    st.session_state.etapa_escolhida = prox_idx
                    st.session_state.pedido          = num
                    st.session_state.pedido_status   = None
                    st.session_state.pedido_validado = True
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)
            with cb:
                st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
                if st.button("✕ Não", use_container_width=True):
                    st.session_state.pedido_status = None; st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
            if st.button("← Buscar outro pedido", use_container_width=True):
                st.session_state.pedido_status = None; st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
        return

    # ── CASO 5: Etapa anterior sem registro — avisa mas permite continuar ──
    if etapa_idx > 0 and not etapas[etapa_idx - 1]["feita"] and not etapas[etapa_idx - 1]["em_andamento"]:
        et_ant = ETAPAS_LBL[etapa_idx - 1]
        _cv1.html(f"""
        <div style="background:#FFFBEB;border:2px solid #F59E0B;border-radius:14px;
                    padding:18px 20px;font-family:sans-serif;">
          <div style="display:flex;align-items:center;gap:10px;margin-bottom:8px;">
            <div style="font-size:24px;flex-shrink:0;">⚠️</div>
            <div>
              <div style="font-size:13px;font-weight:800;color:#92400E;">Etapa anterior sem registro</div>
              <div style="font-size:11px;color:#B45309;font-weight:600;margin-top:2px;">
                <strong>{et_ant}</strong> não consta como finalizada no sistema.</div>
            </div>
          </div>
          <div style="font-size:11px;color:#92400E;font-weight:600;
            background:rgba(245,158,11,0.1);border-radius:8px;padding:8px;text-align:center;">
            Se a etapa já foi realizada fisicamente, você pode continuar mesmo assim.
          </div>
        </div>""", height=138, scrolling=False)
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        ca5, cb5 = st.columns(2)
        with ca5:
            st.markdown('<div class="btn-iniciar">', unsafe_allow_html=True)
            if st.button("▶  Continuar mesmo assim", use_container_width=True, key="caso5_continuar"):
                st.session_state.pedido          = num
                st.session_state.pedido_status   = None
                st.session_state.pedido_validado = True
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
        with cb5:
            st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
            if st.button("← Voltar", use_container_width=True, key="caso5_voltar"):
                st.session_state.pedido_status = None; st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
        return

    # ── CASO 6: Pronto para iniciar ──────────────────────────────────────
    _cv1.html(f"""
    <div style="background:#F0F7F3;border:2px solid #4A7C59;border-radius:14px;
                padding:16px 20px;text-align:center;font-family:sans-serif;">
      <div style="font-size:24px;margin-bottom:6px;">✅</div>
      <div style="font-size:14px;font-weight:800;color:#2d5a3d;">
        Pedido pronto para <strong>{etapa_lbl}</strong></div>
      <div style="font-size:12px;color:#4A7C59;margin-top:4px;font-weight:600;">
        Selecione o operador e inicie o processo.</div>
    </div>""", height=100, scrolling=False)
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    ca, cb = st.columns(2)
    with ca:
        st.markdown('<div class="btn-iniciar">', unsafe_allow_html=True)
        if st.button("▶  Continuar", use_container_width=True):
            st.session_state.pedido          = num
            st.session_state.pedido_status   = None
            st.session_state.pedido_validado = True
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    with cb:
        st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
        if st.button("← Voltar", use_container_width=True):
            st.session_state.pedido_status = None; st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    return


def tela_home():
    _auto_refresh_watcher()
    render_logo()

    if st.session_state.etapa_escolhida is None:
        st.markdown("""
        <div style="display:flex;align-items:center;gap:10px;margin:0 0 20px;">
            <div style="flex:1;height:1px;background:#EDE9E4;"></div>
            <div style="font-size:11px;font-weight:800;letter-spacing:2.5px;text-transform:uppercase;
                color:#9C9490;white-space:nowrap;">Em qual etapa vai trabalhar?</div>
            <div style="flex:1;height:1px;background:#EDE9E4;"></div>
        </div>
        """, unsafe_allow_html=True)

        ETAPA_CFG = [
            {"icon":"📦","color":"#C8566A","bg":"#FFF0F2","shadow":"rgba(200,86,106,0.22)",
             "desc":"Separar peças conforme localização no pedido","step":"01"},
            {"icon":"🗃️","color":"#3B7DD8","bg":"#F0F5FF","shadow":"rgba(59,125,216,0.22)",
             "desc":"Embalar conforme observação do pedido","step":"02"},
            {"icon":"✅","color":"#4A7C59","bg":"#F0F7F3","shadow":"rgba(74,124,89,0.22)",
             "desc":"Conferência via código de barras","step":"03"},
        ]
        st.markdown("""
        <style>
        .etapa-card > div[data-testid="stButton"] > button {
            background: #FFFFFF !important;
            border: 1.5px solid #EDE9E4 !important;
            border-radius: 18px !important;
            height: 86px !important;
            width: 100% !important;
            text-align: left !important;
            padding: 0 20px !important;
            font-family: 'Nunito', sans-serif !important;
            font-size: 16px !important;
            font-weight: 800 !important;
            color: #1A1714 !important;
            box-shadow: 0 2px 10px rgba(0,0,0,0.05) !important;
            transition: all 0.18s ease !important;
        }
        .etapa-card > div[data-testid="stButton"] > button:active {
            transform: translateY(1px) !important;
        }
        </style>""", unsafe_allow_html=True)

        for i, lbl in enumerate(ETAPAS_LBL):
            cfg = ETAPA_CFG[i]
            st.markdown(f"""
            <style>
            .etapa-card-{i} > div[data-testid="stButton"] > button {{
                border-left: 5px solid {cfg["color"]} !important;
            }}
            .etapa-card-{i} > div[data-testid="stButton"] > button:hover {{
                background: {cfg["bg"]} !important;
                border-color: {cfg["color"]} !important;
                box-shadow: 0 8px 24px {cfg["shadow"]}, 0 2px 6px rgba(0,0,0,0.05) !important;
                transform: translateY(-3px) !important;
            }}
            </style>
            <div style="display:flex;align-items:center;gap:8px;margin-bottom:4px;padding-left:4px;">
                <span style="background:{cfg["color"]};color:#fff;font-size:9px;font-weight:900;
                    letter-spacing:1.5px;padding:2px 8px;border-radius:20px;text-transform:uppercase;">
                    Etapa {cfg["step"]}
                </span>
                <span style="font-size:11px;color:#9C9490;font-weight:600;">{cfg["desc"]}</span>
            </div>""", unsafe_allow_html=True)
            st.markdown(f'<div class="etapa-card etapa-card-{i}">', unsafe_allow_html=True)
            if st.button(f'{cfg["icon"]}  {lbl}', use_container_width=True, key=f"etapa_btn_{i}"):
                st.session_state.etapa_escolhida = i
                st.session_state.operador = None
                st.session_state.pedido   = None
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
            if i < 2: st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Botão de Operações em Andamento ──────────────────────────────────
        # Conta sessões ativas (em andamento) e pausadas separadamente
        sessoes_ativas_agora = buscar_todas_sessoes_ativas()
        n_ativas  = len([s for s in sessoes_ativas_agora if int(s.get("iniciado_em", 0)) > 0])
        n_pausadas = len([s for s in sessoes_ativas_agora if int(s.get("iniciado_em", 0)) == 0])
        n_total = n_ativas + n_pausadas
        # Só exibe o botão se houver alguma sessão (ativa ou pausada)
        if n_total > 0:
            if n_pausadas > 0 and n_ativas == 0:
                badge_txt = f"{n_pausadas} pausado{'s' if n_pausadas > 1 else ''}"
                badge_cor = "#7C3AED"
            elif n_pausadas > 0:
                badge_txt = f"{n_ativas} em andamento · {n_pausadas} pausado{'s' if n_pausadas > 1 else ''}"
                badge_cor = "#C8566A"
            else:
                badge_txt = f"{n_ativas} em andamento"
                badge_cor = "#C8566A"
            st.markdown(f"""
            <style>
            .btn-andamento > button {{
                background: linear-gradient(135deg, #1c1917, #2d2925) !important;
                color: #fff !important; border: none !important;
                border-radius: 14px !important; height: 64px !important;
                font-size: 15px !important; font-weight: 800 !important;
                box-shadow: 0 5px 0 rgba(0,0,0,0.40), 0 8px 20px rgba(0,0,0,0.20) !important;
                position: relative !important;
            }}
            .btn-andamento > button:hover {{
                background: linear-gradient(135deg, #292524, #3d3530) !important;
                transform: translateY(-2px) !important;
            }}
            </style>
            <div style="position:relative;margin-bottom:8px;">
              <div style="position:absolute;top:-10px;right:12px;z-index:10;
                background:{badge_cor};color:#fff;font-size:11px;font-weight:900;
                padding:3px 10px;border-radius:20px;border:2px solid #F7F5F2;
                letter-spacing:.5px;">
                {badge_txt}
              </div>
            </div>""", unsafe_allow_html=True)
            st.markdown('<div class="btn-andamento">', unsafe_allow_html=True)
            if st.button("⏱  Ver Operações em Andamento", use_container_width=True):
                st.session_state.tela = "operacoes"; st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
            st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

        _, col_c, _ = st.columns([2, 1, 2])
        with col_c:
            st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
            if st.button("⚙ Admin", use_container_width=True):
                st.session_state.tela = "admin_login"; st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
        return

    etapa_idx = st.session_state.etapa_escolhida
    # Valida etapa_idx — pode vir corrompido se o session_state ficou de uma sessão anterior
    if etapa_idx is None or not isinstance(etapa_idx, int) or etapa_idx < 0 or etapa_idx >= len(ETAPAS_LBL):
        st.session_state.etapa_escolhida = None
        st.rerun()
        return
    etapa_lbl = ETAPAS_LBL[etapa_idx]

    # ── PASSO 2: Digitar Pedido + BUSCAR ─────────────────────────────────────
    if not st.session_state.pedido_validado:

        # ── Se status já foi carregado, renderiza painel e sai ──────────────
        if st.session_state.pedido_status in ("mostrar_status", "mostrar_status_ok"):
            num    = st.session_state.get("_pedido_validando", "")
            # Garante que os dados estão carregados
            if st.session_state.pedido_status == "mostrar_status":
                st.session_state["_status_cache"] = buscar_status_completo_pedido(num)
                st.session_state.pedido_status    = "mostrar_status_ok"
            status = st.session_state.get("_status_cache") or buscar_status_completo_pedido(num)
            _render_status_pedido(num, status, etapa_idx)
            return

        render_stepper(etapa_idx)
        st.markdown(
            f'<div style="background:#F5E8EB;border-left:4px solid #C8566A;border-radius:0 10px 10px 0;'
            f'padding:10px 16px;margin-bottom:20px;font-size:14px;font-weight:700;color:#1A1714;">'
            f'Etapa selecionada: <strong>{etapa_lbl}</strong></div>',
            unsafe_allow_html=True
        )
        st.markdown("""
        <div style="display:flex;align-items:center;gap:10px;margin-bottom:16px;">
            <div style="flex:1;height:1px;background:#EDE9E4;"></div>
            <div style="font-size:11px;font-weight:800;letter-spacing:2.5px;text-transform:uppercase;
                color:#9C9490;white-space:nowrap;">Nº do Pedido</div>
            <div style="flex:1;height:1px;background:#EDE9E4;"></div>
        </div>
        """, unsafe_allow_html=True)

        pedidos_etapa   = buscar_pedidos_por_etapa(etapa_idx)
        pedidos_abertos = [p[0] for p in pedidos_etapa]
        pedidos_info    = {p[0]: p[1] for p in pedidos_etapa}
        has_base        = len(pedidos_etapa) > 0

        st.markdown("""
        <style>
        div[data-testid="stTextInput"] label { display:none !important; }
        div[data-testid="stSelectbox"] label { display:none !important; }
        </style>""", unsafe_allow_html=True)

        pedido_inp = ""

        if has_base and pedidos_abertos:
            def fmt_op_ped(n):
                cli = pedidos_info.get(n, "")
                return f"{n}  —  {cli}" if cli else n
            opcoes_disp = ["— Selecione ou digite —"] + [fmt_op_ped(n) for n in sorted(pedidos_abertos)]
            opcoes_map  = {"— Selecione ou digite —": ""} | {fmt_op_ped(n): n for n in sorted(pedidos_abertos)}
            _, col_sel, _ = st.columns([0.3, 4, 0.3])
            with col_sel:
                sel = st.selectbox("_sel", opcoes_disp, key="home_pedido_sel")
            pedido_inp = opcoes_map.get(sel, "")
            if pedido_inp:
                cli_nome = pedidos_info.get(pedido_inp, "")
                if cli_nome:
                    import streamlit.components.v1 as _cv1
                    _cv1.html(
                        f'<div style="background:#F0F7F3;border:1.5px solid #4A7C59;border-radius:10px;'
                        f'padding:10px 16px;font-family:sans-serif;display:flex;align-items:center;gap:10px;margin:6px 0;">'
                        f'<div style="font-size:18px;">🛍</div>'
                        f'<div><div style="font-size:9px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:#4A7C59;margin-bottom:2px;">Cliente</div>'
                        f'<div style="font-size:13px;font-weight:800;color:#1A1714;">{cli_nome}</div></div></div>',
                        height=58, scrolling=False
                    )
            with st.expander("✏ Digitar número manualmente"):
                manual = st.text_input("Número manual", placeholder="Ex: 49735", key="home_pedido_manual")
                if manual.strip(): pedido_inp = manual.strip()
        else:
            _, col_inp, _ = st.columns([0.3, 4, 0.3])
            with col_inp:
                pedido_inp = st.text_input("_ped", placeholder="Ex: 49735", key="home_pedido_txt")
            if not pedidos_abertos:
                import streamlit.components.v1 as _cv1
                _cv1.html(
                    f'<div style="background:#FEF3C7;border:1px solid #F59E0B;border-radius:10px;'
                    f'padding:10px 14px;font-family:sans-serif;font-size:12px;font-weight:700;color:#92400E;text-align:center;">'
                    f'⚠ Nenhum pedido aguardando {etapa_lbl} no momento.</div>',
                    height=50, scrolling=False
                )

        if st.session_state.pedido_status == "concluido":
            import streamlit.components.v1 as _cv1
            _cv1.html(
                '<div style="background:#F0F7F3;border:2px solid #4A7C59;border-radius:14px;'
                'padding:16px 20px;font-family:sans-serif;text-align:center;margin:8px 0;">'
                '<div style="font-size:22px;margin-bottom:6px;">🎉</div>'
                '<div style="font-size:14px;font-weight:800;color:#2d5a3d;margin-bottom:4px;">Pedido Concluído pela Produção</div>'
                '<div style="font-size:12px;color:#4A7C59;font-weight:600;">As 3 etapas foram finalizadas pelos operadores.</div>'
                '</div>',
                height=110, scrolling=False
            )
            if st.button("← Buscar outro pedido"):
                st.session_state.pedido_status = None; st.rerun()
            return

        if st.session_state.pedido_status == "nao_encontrado":
            num_pend = st.session_state.get("_pedido_validando", "")
            import streamlit.components.v1 as _cv1
            _cv1.html(
                f'<div style="background:#FFFBEB;border:2px solid #FCD34D;border-radius:14px;'
                f'padding:16px 20px;font-family:sans-serif;text-align:center;margin:8px 0;">'
                f'<div style="font-size:22px;margin-bottom:6px;">❓</div>'
                f'<div style="font-size:14px;font-weight:800;color:#92400E;margin-bottom:4px;">Pedido Não Encontrado</div>'
                f'<div style="font-size:12px;color:#B45309;font-weight:600;">'
                f'Pedido <b>{num_pend}</b> não está na base. Deseja cadastrá-lo?</div>'
                f'</div>',
                height=115, scrolling=False
            )
            cc1, cc2 = st.columns(2)
            with cc1:
                st.markdown('<div class="btn-iniciar">', unsafe_allow_html=True)
                if st.button("✓ Cadastrar", use_container_width=True):
                    cadastrar_pedido_avulso(num_pend)
                    buscar_pedidos_por_etapa.clear()
                    st.session_state.pedido_status  = None
                    st.session_state.pedido         = num_pend
                    st.session_state.pedido_validado = True
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)
            with cc2:
                st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
                if st.button("✕ Cancelar", use_container_width=True):
                    st.session_state.pedido_status = None; st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)
            return

        if st.session_state.duplicata_info:
            info   = st.session_state.duplicata_info
            op_ant = info["operador_anterior"]
            etl    = ETAPAS_LBL[etapa_idx]
            msg_tp = "está sendo processado agora" if info["em_andamento"] else "já passou pela fase de"
            import streamlit.components.v1 as _cv1
            _cv1.html(
                f'<div style="background:#FFF7ED;border:2px solid #F97316;border-radius:14px;'
                f'padding:18px 22px;font-family:sans-serif;margin:8px 0;">'
                f'<div style="font-size:22px;text-align:center;margin-bottom:8px;">⚠️</div>'
                f'<div style="font-size:14px;font-weight:800;color:#9A3412;text-align:center;margin-bottom:8px;">Atenção — Pedido em Conflito</div>'
                f'<div style="font-size:13px;color:#7C2D12;font-weight:600;text-align:center;line-height:1.6;">'
                f'Pedido {info["pedido"]} {msg_tp} <b>{etl}</b>'
                f'{(" (por " + op_ant + ")") if op_ant else ""}.<br><br>'
                f'Deseja mesmo assim prosseguir?</div>'
                f'</div>',
                height=175, scrolling=False
            )
            cc1, cc2 = st.columns(2)
            with cc1:
                st.markdown('<div class="btn-iniciar">', unsafe_allow_html=True)
                if st.button("✓ Sim, prosseguir", use_container_width=True):
                    st.session_state.pedido          = info["pedido"]
                    st.session_state.duplicata_info  = None
                    st.session_state.pedido_validado = True
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)
            with cc2:
                st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
                if st.button("✕ Não", use_container_width=True):
                    st.session_state.duplicata_info = None; st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)
            return

        st.markdown("<br>", unsafe_allow_html=True)
        c1, _, c2 = st.columns([3, 0.3, 1.5])
        with c1:
            st.markdown('<div class="btn-iniciar">', unsafe_allow_html=True)
            if st.button("🔍  BUSCAR", use_container_width=True, key="home_buscar"):
                num = pedido_inp.strip() if isinstance(pedido_inp, str) else ""
                if not num:
                    st.session_state.erro_pedido = True; st.rerun()
                st.session_state.erro_pedido       = False
                st.session_state._pedido_validando = num
                st.session_state.pedido_status     = "mostrar_status"
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
        with c2:
            st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
            if st.button("← Voltar", use_container_width=True, key="home_ped_voltar"):
                st.session_state.etapa_escolhida = None
                st.session_state.pedido_status   = None; st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        if st.session_state.erro_pedido:
            st.markdown(
                '<div style="text-align:center;color:#C8566A;font-size:13px;font-weight:800;margin-top:6px;">'
                '⚠ Selecione ou digite o número do pedido.</div>',
                unsafe_allow_html=True
            )
        return


    # ── PASSO 3: Selecionar Operador ─────────────────────────────────────────
    pedido_val   = st.session_state.pedido
    pedidos_etapa = buscar_pedidos_por_etapa(etapa_idx)
    pedidos_info  = {p[0]: p[1] for p in pedidos_etapa}
    if pedido_val not in pedidos_info:
        _rows = _get("pedidos_base", f"numero=eq.{pedido_val}&select=cliente")
        if _rows: pedidos_info[pedido_val] = _rows[0].get("cliente", "")
    cli_nome = pedidos_info.get(pedido_val, "")

    render_stepper(etapa_idx)

    import streamlit.components.v1 as _cv1
    _cv1.html(
        f'<div style="background:#F0F7F3;border:1.5px solid #4A7C59;border-radius:12px;'
        f'padding:12px 18px;font-family:sans-serif;display:flex;align-items:center;gap:14px;margin-bottom:4px;">'
        f'<div style="width:38px;height:38px;border-radius:50%;background:#4A7C59;flex-shrink:0;'
        f'display:flex;align-items:center;justify-content:center;font-size:18px;">✓</div>'
        f'<div>'
        f'<div style="font-size:9px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:#4A7C59;margin-bottom:2px;">Pedido Encontrado</div>'
        f'<div style="font-size:16px;font-weight:900;color:#1A1714;font-family:monospace;">{pedido_val}'
        f'{"  · " + cli_nome if cli_nome else ""}</div>'
        f'<div style="font-size:11px;color:#9C9490;margin-top:2px;">Etapa: {etapa_lbl}</div>'
        f'</div></div>',
        height=80, scrolling=False
    )

    st.markdown("<br style='line-height:0.3'>", unsafe_allow_html=True)
    st.markdown("""
    <div style="display:flex;align-items:center;gap:10px;margin-bottom:16px;">
        <div style="flex:1;height:1px;background:#EDE9E4;"></div>
        <div style="font-size:11px;font-weight:800;letter-spacing:2.5px;text-transform:uppercase;
            color:#9C9490;white-space:nowrap;">Identificar Operador</div>
        <div style="flex:1;height:1px;background:#EDE9E4;"></div>
    </div>
    """, unsafe_allow_html=True)

    render_avatar_grid(on_click_key="op_step3")

    if st.session_state.operador:
        st.markdown("<br style='line-height:0.3'>", unsafe_allow_html=True)
        _, col_ini, _ = st.columns([0.3, 4, 0.3])
        with col_ini:
            st.markdown('<div class="btn-iniciar">', unsafe_allow_html=True)
            if st.button("▶  INICIAR OPERAÇÃO", use_container_width=True, key="home_iniciar_op"):
                # ✅ CORREÇÃO: NÃO registra sessão ativa aqui.
                # A sessão só é registrada quando o operador clicar em
                # "INICIAR CRONÔMETRO" na tela de produção.
                # Isso garante que o PiP só aparece com o cronômetro rodando.
                st.session_state.etapa_idx       = etapa_idx
                st.session_state.rodando         = False
                st.session_state.inicio          = None
                st.session_state.acum            = 0
                st.session_state.modal           = None
                st.session_state.tela            = "producao"
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    _, col_v, _ = st.columns([2, 2, 2])
    with col_v:
        st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
        if st.button("← Voltar", use_container_width=True, key="home_op_voltar"):
            st.session_state.pedido_validado = False
            st.session_state.pedido          = None
            st.session_state.operador        = None; st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)


# ─────────────────────────────────────
#  TELA: PRODUÇÃO
# ─────────────────────────────────────
def tela_producao():
    _auto_refresh_watcher()
    render_logo()
    render_stepper(st.session_state.etapa_idx)

    op        = st.session_state.operador
    etapa_idx = st.session_state.etapa_idx
    etapa_lbl = ETAPAS_LBL[etapa_idx]

    # Busca Est. Alocado (qtd) e Vr. Alocado (valor) do pedido — etapa 0
    _est_alocado_banco = None
    _vr_alocado_banco  = None
    if etapa_idx == 0:
        _info_ped = _get("pedidos_base",
            f"numero=eq.{st.session_state.pedido or ''}&select=est_alocado,vr_alocado")
        if isinstance(_info_ped, list) and _info_ped:
            _est_alocado_banco = _info_ped[0].get("est_alocado")
            _vr_alocado_banco  = _info_ped[0].get("vr_alocado")

        _qtd_banco = int(float(_est_alocado_banco)) if _est_alocado_banco else 0

        # Inicializa rastreando o pedido atual — reseta ao trocar de pedido
        _ped_key = f"ped_qtd_pedido_ref"
        if st.session_state.get(_ped_key) != st.session_state.pedido:
            # Novo pedido — reinicia confirmação e preenche com valor do banco
            st.session_state[_ped_key]          = st.session_state.pedido
            st.session_state.ped_qtd_confirmada = None
            st.session_state.ped_qtd_valor      = _qtd_banco
            st.session_state.qtd_pecas_prefill  = None

        # Garante que ped_qtd_valor nunca fique zerado se o banco tiver valor
        if not st.session_state.get("ped_qtd_valor") and _qtd_banco:
            st.session_state.ped_qtd_valor = _qtd_banco

    st.markdown("<br style='line-height:0.5'>", unsafe_allow_html=True)

    # ── Card info + botão INICIAR CRONÔMETRO ──
    if not st.session_state.rodando and st.session_state.acum == 0 and not st.session_state.modal:
        pedido_val = st.session_state.pedido or ""
        initial    = op[0].upper()

        components.html(f"""
        <!DOCTYPE html><html><head>
        <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@800;900&display=swap" rel="stylesheet">
        <style>* {{margin:0;padding:0;box-sizing:border-box;}}</style>
        </head><body style="background:transparent;font-family:Nunito,sans-serif;">
        <div style="background:linear-gradient(135deg,#C8566A 0%,#9E3F52 100%);border-radius:20px;
            box-shadow:0 8px 0 rgba(100,20,35,0.28),0 16px 36px rgba(200,86,106,0.28);
            overflow:hidden;position:relative;">
            <div style="position:absolute;right:-30px;top:-30px;width:140px;height:140px;border-radius:50%;background:rgba(255,255,255,0.07);"></div>
            <div style="display:flex;align-items:center;justify-content:space-between;padding:22px 28px;position:relative;z-index:1;">
                <div>
                    <div style="font-size:9px;font-weight:800;letter-spacing:2.5px;color:rgba(255,255,255,0.55);text-transform:uppercase;margin-bottom:4px;">Etapa Atual</div>
                    <div style="font-size:22px;font-weight:900;color:#fff;letter-spacing:-0.3px;">{etapa_lbl}</div>
                </div>
                <div style="width:1.5px;height:48px;background:rgba(255,255,255,0.2);border-radius:2px;margin:0 16px;flex-shrink:0;"></div>
                <div style="text-align:right;">
                    <div style="font-size:9px;font-weight:800;letter-spacing:2.5px;color:rgba(255,255,255,0.55);text-transform:uppercase;margin-bottom:4px;">Operador</div>
                    <div style="display:flex;align-items:center;gap:8px;justify-content:flex-end;">
                        <div style="width:32px;height:32px;border-radius:50%;background:rgba(255,255,255,0.20);border:2px solid rgba(255,255,255,0.35);display:flex;align-items:center;justify-content:center;font-size:14px;font-weight:900;color:#fff;">{initial}</div>
                        <span style="font-size:17px;font-weight:800;color:#fff;">{op}</span>
                    </div>
                </div>
            </div>
            <div style="background:rgba(0,0,0,0.15);padding:12px 28px;display:flex;align-items:center;gap:10px;">
                <div style="font-size:9px;font-weight:800;letter-spacing:2px;color:rgba(255,255,255,0.5);text-transform:uppercase;">Pedido</div>
                <div style="font-family:monospace;font-size:18px;font-weight:800;color:#fff;letter-spacing:1px;">{pedido_val}</div>
            </div>
        </div>
        </body></html>
        """, height=130, scrolling=False)

        st.markdown("<br style='line-height:0.3'>", unsafe_allow_html=True)

        # ── Aviso se há tempo pausado salvo ──────────────────────────
        tempo_pausado_prev = buscar_tempo_pausado(pedido_val, etapa_idx)
        if tempo_pausado_prev > 0:
            h_pv, r_pv = divmod(tempo_pausado_prev, 3600); m_pv, s_pv = divmod(r_pv, 60)
            components.html(f"""<!DOCTYPE html><html><head>
            <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@700;800;900&display=swap" rel="stylesheet">
            </head><body style="background:transparent;font-family:Nunito,sans-serif;margin:0;">
            <div style="background:#FFF7ED;border:2px solid #E07B3A;border-radius:14px;
                        padding:14px 18px;display:flex;align-items:center;gap:14px;">
              <div style="font-size:26px;flex-shrink:0;">🌙</div>
              <div>
                <div style="font-size:13px;font-weight:800;color:#92400E;margin-bottom:2px;">
                  Etapa pausada do dia anterior</div>
                <div style="font-size:12px;font-weight:600;color:#B45309;">
                  Tempo salvo: <strong style="font-family:monospace;">{h_pv:02d}:{m_pv:02d}:{s_pv:02d}</strong>
                  — o cronômetro continuará de onde parou.
                </div>
              </div>
            </div></body></html>""", height=78, scrolling=False)
            st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

        # ── Confirmação de Est. Alocado + Vr. Alocado (etapa 0 — Separação) ──
        if etapa_idx == 0:
            ja_confirmou = st.session_state.ped_qtd_confirmada is not None

            if not ja_confirmou:
                # Formata valor para exibição
                _vr_fmt = f"R$ {float(_vr_alocado_banco):,.2f}".replace(",","X").replace(".",",").replace("X",".") if _vr_alocado_banco else "—"
                _qtd_exibir = st.session_state.ped_qtd_valor

                components.html(f"""<!DOCTYPE html><html><head>
                <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@700;800;900&family=DM+Mono:wght@500&display=swap" rel="stylesheet">
                </head><body style="background:transparent;font-family:Nunito,sans-serif;margin:0;">
                <div style="background:#fff;border:2px solid #3B7DD8;border-radius:16px;
                            padding:16px 20px;box-shadow:0 4px 16px rgba(59,125,216,0.12);">
                  <div style="display:flex;align-items:center;gap:12px;margin-bottom:12px;">
                    <div style="width:42px;height:42px;border-radius:12px;
                      background:linear-gradient(135deg,#3B7DD8,#2563EB);
                      display:flex;align-items:center;justify-content:center;font-size:20px;flex-shrink:0;">📦</div>
                    <div>
                      <div style="font-size:14px;font-weight:900;color:#1A1714;">Confirmar Dados do Pedido</div>
                      <div style="font-size:11px;font-weight:600;color:#9C9490;">Verifique e corrija se necessário antes de iniciar.</div>
                    </div>
                  </div>
                  <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;">
                    <div style="background:#F0F5FF;border-radius:10px;padding:12px 14px;">
                      <div style="font-size:9px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:#3B7DD8;margin-bottom:4px;">Qtd. de Itens</div>
                      <div style="font-family:'DM Mono',monospace;font-size:22px;font-weight:700;color:#1A1714;">{_qtd_exibir}</div>
                      <div style="font-size:10px;color:#9C9490;margin-top:2px;">Est. Alocado</div>
                    </div>
                    <div style="background:#F0F7F3;border-radius:10px;padding:12px 14px;">
                      <div style="font-size:9px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:#4A7C59;margin-bottom:4px;">Valor do Pedido</div>
                      <div style="font-family:'DM Mono',monospace;font-size:18px;font-weight:700;color:#1A1714;">{_vr_fmt}</div>
                      <div style="font-size:10px;color:#9C9490;margin-top:2px;">Vr. Alocado</div>
                    </div>
                  </div>
                </div></body></html>""", height=168, scrolling=False)

                st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
                st.markdown("""
                <div style="font-size:10px;font-weight:800;letter-spacing:1.5px;color:#9C9490;
                  text-transform:uppercase;text-align:center;margin-bottom:6px;">
                  Corrigir quantidade se necessário:
                </div>""", unsafe_allow_html=True)
                st.markdown("""
                <style>
                div[data-testid="stNumberInput"] label { display:none !important; }
                div[data-testid="stNumberInput"] input {
                    border: 2px solid #3B7DD8 !important; border-radius: 12px !important;
                    background: #fff !important; font-family: 'Nunito', sans-serif !important;
                    font-size: 22px !important; font-weight: 800 !important;
                    padding: 14px 18px !important; height: 58px !important;
                    text-align: center !important; color: #1A1714 !important;
                }
                div[data-testid="stNumberInput"] input:focus {
                    border-color: #2563EB !important;
                    box-shadow: 0 0 0 4px rgba(59,125,216,0.12) !important;
                }
                </style>""", unsafe_allow_html=True)

                _, c_num, _ = st.columns([0.3, 4, 0.3])
                with c_num:
                    qtd_input = st.number_input("_qtd_conf", min_value=0,
                        value=st.session_state.ped_qtd_valor,
                        step=1, key="ped_qtd_input",
                        label_visibility="collapsed")
                    # Salva no session_state a cada interação
                    st.session_state.ped_qtd_valor = int(qtd_input)

                st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
                _, c_conf, _ = st.columns([0.3, 4, 0.3])
                with c_conf:
                    st.markdown("""
                    <style>
                    .btn-conf-qtd > button {
                        background: linear-gradient(135deg,#3B7DD8,#2563EB) !important;
                        color:#fff !important; border:none !important;
                        border-radius:12px !important; height:52px !important;
                        font-size:14px !important; font-weight:800 !important;
                        box-shadow: 0 4px 0 rgba(30,60,140,0.40) !important;
                    }
                    .btn-conf-qtd > button:hover { transform:translateY(-1px) !important; }
                    </style>""", unsafe_allow_html=True)
                    st.markdown('<div class="btn-conf-qtd">', unsafe_allow_html=True)
                    # Label fixo — evita recriação do botão pelo Streamlit a cada digitação
                    if st.button("✓  Confirmar quantidade",
                                 use_container_width=True, key="btn_conf_qtd"):
                        qtd_confirmada = st.session_state.ped_qtd_valor
                        _patch("pedidos_base",
                               f"numero=eq.{pedido_val}",
                               {"est_alocado": qtd_confirmada})
                        st.session_state.ped_qtd_confirmada = qtd_confirmada
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)

                st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
                st.markdown("""
                <div style="background:#FFF8E6;border:1.5px solid #F59E0B;border-radius:10px;
                            padding:10px 16px;font-size:12px;font-weight:700;color:#92400E;
                            text-align:center;">
                    ⚠️ Confirme os dados do pedido para liberar o cronômetro.
                </div>""", unsafe_allow_html=True)
                st.stop()

            else:
                # Badge resumido após confirmação
                _vr_fmt2 = f"R$ {float(_vr_alocado_banco):,.2f}".replace(",","X").replace(".",",").replace("X",".") if _vr_alocado_banco else "—"
                components.html(f"""<!DOCTYPE html><html><head>
                <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@700;800;900&family=DM+Mono:wght@500&display=swap" rel="stylesheet">
                </head><body style="background:transparent;font-family:Nunito,sans-serif;margin:0;">
                <div style="background:#EBF5FF;border:1.5px solid #3B7DD8;border-radius:10px;
                            padding:10px 16px;display:flex;align-items:center;gap:12px;">
                  <span style="font-size:18px;">✅</span>
                  <div style="font-size:12px;font-weight:700;color:#1E40AF;flex:1;">
                    <strong style="font-size:14px;">{st.session_state.ped_qtd_confirmada} itens</strong>
                    <span style="color:#9C9490;margin:0 6px;">·</span>
                    <span style="font-family:'DM Mono',monospace;">{_vr_fmt2}</span>
                  </div>
                </div></body></html>""", height=52, scrolling=False)
                st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        _, c1, _ = st.columns([0.3, 4, 0.3])
        with c1:
            st.markdown('<div class="btn-iniciar">', unsafe_allow_html=True)
            if st.button("▶  INICIAR CRONÔMETRO", use_container_width=True):
                # ✅ SESSÃO REGISTRADA AQUI — único momento correto.
                # Verifica se há tempo pausado salvo (pedido retomado após pausa)
                tempo_pausado = buscar_tempo_pausado(pedido_val, etapa_idx)
                st.session_state.rodando = True
                st.session_state.inicio  = time.time()
                st.session_state.acum    = tempo_pausado   # retoma do ponto pausado (0 se novo)
                # Upsert com tempo_pausado preservado para manter acumulado correto
                _upsert("sessoes_ativas",
                        {"pedido": pedido_val, "etapa_idx": etapa_idx,
                         "operador": op, "iniciado_em": int(time.time()),
                         "tempo_pausado": tempo_pausado},
                        "pedido,etapa_idx")
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        _, c2, _ = st.columns([0.3, 4, 0.3])
        with c2:
            st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
            if st.button("← Voltar ao Menu", use_container_width=True):
                # Sem sessão ativa registrada, nada para remover.
                # Limpa estado e volta ao lobby.
                st.session_state.pedido          = None
                st.session_state.pedido_status   = None
                st.session_state.pedido_validado = False
                st.session_state.operador        = None
                st.session_state.etapa_escolhida = None
                st.session_state.tela            = "home"
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

    # ── Timer rodando ──
    elif st.session_state.rodando:
        elapsed = get_elapsed()
        h, rem = divmod(elapsed, 3600); m, s = divmod(rem, 60)
        pedido_val = st.session_state.pedido
        initial = op[0].upper()
        timer_str = f"{h:02d}:{m:02d}:{s:02d}"
        components.html(f"""
        <!DOCTYPE html><html><head>
        <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@800;900&family=DM+Mono:wght@500&display=swap" rel="stylesheet">
        <style>* {{margin:0;padding:0;box-sizing:border-box;}}</style>
        </head><body style="background:transparent;font-family:Nunito,sans-serif;">
        <div style="background:linear-gradient(135deg,#C8566A 0%,#9E3F52 100%);border-radius:20px;padding:0;box-shadow:0 8px 0 rgba(100,20,35,0.28),0 16px 36px rgba(200,86,106,0.28);overflow:hidden;position:relative;">
            <div style="position:absolute;right:-30px;top:-30px;width:140px;height:140px;border-radius:50%;background:rgba(255,255,255,0.06);"></div>
            <div style="position:absolute;left:-20px;bottom:-40px;width:110px;height:110px;border-radius:50%;background:rgba(255,255,255,0.04);"></div>
            <div style="display:flex;align-items:center;justify-content:space-between;padding:16px 28px 0;position:relative;z-index:1;">
                <div>
                    <div style="font-size:9px;font-weight:800;letter-spacing:2px;color:rgba(255,255,255,0.55);text-transform:uppercase;margin-bottom:2px;">Pedido</div>
                    <div style="font-family:'DM Mono',monospace;font-size:18px;font-weight:500;color:#fff;">{pedido_val}</div>
                </div>
                <div style="width:1.5px;height:36px;background:rgba(255,255,255,0.2);border-radius:2px;margin:0 16px;flex-shrink:0;"></div>
                <div style="text-align:right;">
                    <div style="font-size:9px;font-weight:800;letter-spacing:2px;color:rgba(255,255,255,0.55);text-transform:uppercase;margin-bottom:2px;">Etapa · Operador</div>
                    <div style="font-size:15px;font-weight:800;color:#fff;">{etapa_lbl} · {op}</div>
                </div>
            </div>
            <div style="text-align:center;padding:18px 24px 24px;position:relative;z-index:1;">
                <div style="font-family:'DM Mono',monospace;font-size:70px;font-weight:500;color:#fff;letter-spacing:-3px;line-height:1;">{timer_str}</div>
                <div style="font-size:10px;font-weight:800;color:rgba(255,255,255,0.45);letter-spacing:2px;text-transform:uppercase;margin-top:8px;">cronômetro em execução</div>
            </div>
        </div>
        </body></html>
        """, height=200, scrolling=False)

        # ── Botões: Voltar ao painel / Finalizar ──
        col_back, col_fin2 = st.columns([1, 2])
        with col_back:
            st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
            if st.button("⊞  Painel", use_container_width=True,
                         help="Volta ao painel sem perder o cronômetro"):
                # Mantém sessão ativa — aparece no painel de operações
                st.session_state.rodando         = False
                st.session_state.inicio          = None
                st.session_state.acum            = 0
                st.session_state.pedido          = None
                st.session_state.pedido_validado = False
                st.session_state.operador        = None
                st.session_state.etapa_escolhida = None
                st.session_state.tela            = "operacoes"
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        _, col_fin, _ = st.columns([0.5, 5, 0.5])
        with col_fin:
            st.markdown("""
            <style>
            div[data-testid="stNumberInput"] label { display:none !important; }
            div[data-testid="stNumberInput"] input {
                border: 2px solid #E0DBD4 !important; border-radius: 12px !important;
                background: #fff !important; font-family: 'Nunito', sans-serif !important;
                font-size: 16px !important; font-weight: 700 !important;
                padding: 14px 18px !important; height: 52px !important;
                text-align: center !important;
            }
            div[data-testid="stNumberInput"] input:focus {
                border-color: #C8566A !important;
                box-shadow: 0 0 0 4px rgba(200,86,106,0.10) !important;
            }
            </style>""", unsafe_allow_html=True)
            st.markdown(
                '<div style="font-size:10px;font-weight:800;letter-spacing:2px;color:#9C9490;'
                'text-transform:uppercase;margin-bottom:6px;text-align:center;">Qtd de Peças</div>',
                unsafe_allow_html=True
            )
            # Pré-preencher com quantidade confirmada no início (etapa 0)
            # ou buscar do banco para as demais etapas
            _qtd_default = 0
            if etapa_idx == 0:
                _qtd_default = st.session_state.get("ped_qtd_confirmada") or                                st.session_state.get("ped_qtd_valor") or 0
            else:
                if "qtd_pecas_prefill" not in st.session_state:
                    _rows_est = _get("pedidos_base",
                        f"numero=eq.{st.session_state.pedido}&select=est_alocado")
                    if isinstance(_rows_est, list) and _rows_est:
                        _v = _rows_est[0].get("est_alocado")
                        st.session_state.qtd_pecas_prefill = int(float(_v)) if _v else 0
                    else:
                        st.session_state.qtd_pecas_prefill = 0
                _qtd_default = st.session_state.qtd_pecas_prefill

            qtd_pecas_val = st.number_input("_qtd_pecas", min_value=0,
                                             value=int(_qtd_default),
                                             step=1, key="qtd_pecas_input",
                                             label_visibility="collapsed")
            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
            st.markdown('<div class="btn-finalizar">', unsafe_allow_html=True)
            if st.button("■  FINALIZAR ETAPA", use_container_width=True):
                tempo = get_elapsed()
                st.session_state.acum    = tempo
                st.session_state.rodando = False
                st.session_state.inicio  = None
                # Usa o valor do input (já pré-preenchido); se zero, usa o default
                qtd = int(qtd_pecas_val) if qtd_pecas_val and int(qtd_pecas_val) > 0 else (int(_qtd_default) if _qtd_default else None)
                salvar(st.session_state.pedido, op, ETAPAS[etapa_idx], etapa_idx, tempo, qtd)
                remover_sessao_ativa(st.session_state.pedido, etapa_idx)
                if etapa_idx == 2:
                    marcar_concluido(st.session_state.pedido)
                    st.session_state.modal = "concluido"
                else:
                    st.session_state.modal = "proxima"
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
        _, col_menu, _ = st.columns([0.5, 5, 0.5])
        with col_menu:
            st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
            if st.button("← Voltar ao Menu", use_container_width=True, key="voltar_menu_rodando"):
                # Sessão ativa permanece intacta no Supabase — cronômetro continua
                # contando pelo iniciado_em. O operador finaliza pelo painel
                # "Ver Operações em Andamento" quando quiser.
                # NÃO chama remover_sessao_ativa nem pausar_para_amanha.
                st.session_state.rodando         = False
                st.session_state.inicio          = None
                st.session_state.acum            = 0
                st.session_state.pedido          = None
                st.session_state.pedido_status   = None
                st.session_state.pedido_validado = False
                st.session_state.operador        = None
                st.session_state.etapa_escolhida = None
                st.session_state.tela            = "home"
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        # ── Botão Pausar para Amanhã ──────────────────────────────────
        st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
        _, col_pausa, _ = st.columns([0.5, 5, 0.5])
        with col_pausa:
            if "pausa_modo" not in st.session_state:
                st.session_state.pausa_modo  = False  # False = normal, True = pedindo senha
            if "pausa_erro" not in st.session_state:
                st.session_state.pausa_erro  = False

            if not st.session_state.pausa_modo:
                # Botão laranja "Pausar para amanhã"
                st.markdown("""
                <style>
                .btn-pausar > button {
                    background: linear-gradient(135deg,#E07B3A,#B85C20) !important;
                    color: #fff !important; border: none !important;
                    border-radius: 12px !important; height: 48px !important;
                    font-size: 13px !important; font-weight: 800 !important;
                    box-shadow: 0 4px 0 rgba(120,50,10,0.40) !important;
                }
                .btn-pausar > button:hover { transform:translateY(-1px) !important; }
                </style>""", unsafe_allow_html=True)
                st.markdown('<div class="btn-pausar">', unsafe_allow_html=True)
                if st.button("🌙  Pausar para Amanhã", use_container_width=True, key="btn_pausar_amanha"):
                    st.session_state.pausa_modo = True
                    st.session_state.pausa_erro = False
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

            else:
                # Card de confirmação com senha admin
                tempo_atual = get_elapsed()
                h_p, r_p   = divmod(tempo_atual, 3600); m_p, s_p = divmod(r_p, 60)
                tempo_str_p = f"{h_p:02d}:{m_p:02d}:{s_p:02d}"

                components.html(f"""<!DOCTYPE html><html><head>
                <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@700;800;900&family=DM+Mono:wght@500&display=swap" rel="stylesheet">
                </head><body style="background:transparent;font-family:Nunito,sans-serif;margin:0;">
                <div style="background:#fff;border:2px solid #E07B3A;border-radius:16px;
                            padding:18px 20px;box-shadow:0 4px 20px rgba(224,123,58,0.18);">
                  <div style="display:flex;align-items:center;gap:12px;margin-bottom:12px;">
                    <div style="width:40px;height:40px;border-radius:12px;
                      background:linear-gradient(135deg,#E07B3A,#B85C20);
                      display:flex;align-items:center;justify-content:center;font-size:20px;flex-shrink:0;">🌙</div>
                    <div>
                      <div style="font-size:14px;font-weight:900;color:#1A1714;">Pausar para Amanhã</div>
                      <div style="font-size:11px;font-weight:600;color:#9C9490;">
                        Tempo atual: <span style="font-family:'DM Mono',monospace;color:#E07B3A;font-weight:700;">{tempo_str_p}</span>
                        será salvo e retomado depois.
                      </div>
                    </div>
                  </div>
                  <div style="font-size:11px;font-weight:800;letter-spacing:1px;color:#9C9490;
                    text-transform:uppercase;margin-bottom:6px;">Digite a senha do administrador:</div>
                </div>
                </body></html>""", height=130, scrolling=False)

                st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
                if st.session_state.pausa_erro:
                    st.markdown("""
                    <div style="background:#FEF2F2;border:1.5px solid #FECACA;border-radius:10px;
                                padding:8px 14px;font-size:12px;font-weight:800;color:#991B1B;
                                text-align:center;margin-bottom:6px;">
                        ❌ Senha incorreta. Tente novamente.
                    </div>""", unsafe_allow_html=True)

                senha_input = st.text_input("_senha_pausa", type="password",
                                             placeholder="Senha admin...",
                                             label_visibility="collapsed",
                                             key="pausa_senha_input")

                motivo_input = st.text_input("_motivo_pausa",
                                              placeholder="Motivo da pausa (opcional)...",
                                              label_visibility="collapsed",
                                              key="pausa_motivo_input")

                st.markdown("""
                <style>
                .btn-pausar-conf > button {
                    background: linear-gradient(135deg,#E07B3A,#B85C20) !important;
                    color:#fff !important; border:none !important;
                    border-radius:10px !important; height:44px !important;
                    font-size:13px !important; font-weight:800 !important;
                }
                .btn-voltar > button { height:44px !important; }
                </style>""", unsafe_allow_html=True)
                c_conf, c_canc = st.columns(2)
                with c_conf:
                    st.markdown('<div class="btn-pausar-conf">', unsafe_allow_html=True)
                    if st.button("🌙 Confirmar Pausa", use_container_width=True, key="btn_pausa_confirmar"):
                        if senha_input.strip() == ADMIN_SENHA:
                            tempo_salvar = get_elapsed()
                            pausar_para_amanha(
                                st.session_state.pedido,
                                etapa_idx,
                                op,
                                tempo_salvar
                            )
                            registrar_pausa_log(
                                st.session_state.pedido,
                                etapa_idx,
                                op,
                                tempo_salvar,
                                motivo_input
                            )
                            buscar_pausas_log.clear()
                            # Limpa o estado local sem remover a sessão do banco
                            st.session_state.rodando         = False
                            st.session_state.inicio          = None
                            st.session_state.acum            = 0
                            st.session_state.pausa_modo      = False
                            st.session_state.pausa_erro      = False
                            st.session_state.pedido          = None
                            st.session_state.pedido_status   = None
                            st.session_state.pedido_validado = False
                            st.session_state.operador        = None
                            st.session_state.etapa_escolhida = None
                            st.session_state.modal           = "pausado"
                            st.rerun()
                        else:
                            st.session_state.pausa_erro = True
                            st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
                with c_canc:
                    st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
                    if st.button("✕ Cancelar", use_container_width=True, key="btn_pausa_cancelar"):
                        st.session_state.pausa_modo = False
                        st.session_state.pausa_erro = False
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)

        time.sleep(1); st.rerun()

    # ── Modal: pausado para amanhã ──
    elif st.session_state.modal == "pausado":
        pedido_val = st.session_state.pedido or ""
        components.html(f"""<!DOCTYPE html><html><head>
        <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@600;700;800;900&family=DM+Mono:wght@500&display=swap" rel="stylesheet">
        <style>*{{margin:0;padding:0;box-sizing:border-box;}}</style>
        </head><body style="background:transparent;font-family:Nunito,sans-serif;">
        <div style="background:#fff;border-radius:20px;overflow:hidden;
                    box-shadow:0 4px 24px rgba(0,0,0,0.08);border:1.5px solid #EDE9E4;">
          <div style="background:linear-gradient(135deg,#E07B3A,#B85C20);
                      padding:28px;text-align:center;position:relative;overflow:hidden;">
            <div style="position:absolute;right:-20px;top:-20px;width:100px;height:100px;
                        border-radius:50%;background:rgba(255,255,255,0.07);"></div>
            <div style="width:58px;height:58px;background:rgba(255,255,255,0.18);border-radius:50%;
                        display:flex;align-items:center;justify-content:center;
                        margin:0 auto 14px;border:2px solid rgba(255,255,255,0.35);position:relative;z-index:1;">
              <span style="font-size:28px;">🌙</span>
            </div>
            <div style="font-size:20px;font-weight:900;color:#fff;margin-bottom:4px;position:relative;z-index:1;">
              Etapa Pausada!</div>
            <div style="font-size:12px;font-weight:700;color:rgba(255,255,255,0.65);
                        letter-spacing:1px;position:relative;z-index:1;">
              O tempo foi salvo. Retome amanhã do ponto em que parou.</div>
          </div>
          <div style="padding:20px 24px;text-align:center;">
            <div style="font-size:11px;font-weight:800;letter-spacing:2px;color:#9C9490;
                        text-transform:uppercase;margin-bottom:8px;">Pedido {pedido_val}</div>
            <div style="font-size:13px;font-weight:600;color:#5C5450;line-height:1.7;">
              Para continuar, basta selecionar o mesmo pedido<br>
              e o sistema retomará o cronômetro de onde parou.
            </div>
          </div>
        </div>
        </body></html>""", height=290, scrolling=False)
        st.markdown("<br style='line-height:0.2'>", unsafe_allow_html=True)
        st.markdown('<div class="btn-iniciar">', unsafe_allow_html=True)
        if st.button("▶  Ir ao Menu Principal", use_container_width=True, key="pausado_home"):
            st.session_state.modal           = None
            st.session_state.pedido          = None
            st.session_state.pedido_validado = False
            st.session_state.etapa_idx       = 0
            st.session_state.acum            = 0
            st.session_state.tela            = "home"
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    elif st.session_state.modal == "proxima":
        next_lbl   = ETAPAS_LBL[etapa_idx + 1]
        tempo_fmt  = fmt(st.session_state.acum)
        pedido_val = st.session_state.pedido
        components.html(f"""
        <!DOCTYPE html><html><head>
        <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@800;900&family=DM+Mono:wght@500&display=swap" rel="stylesheet">
        <style>* {{margin:0;padding:0;box-sizing:border-box;}}</style>
        </head><body style="background:transparent;font-family:Nunito,sans-serif;">
        <div style="background:#fff;border-radius:20px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.08);border:1.5px solid #EDE9E4;">
            <div style="background:linear-gradient(135deg,#4A7C59,#2d5c3e);padding:24px;text-align:center;">
                <div style="width:56px;height:56px;background:rgba(255,255,255,0.2);border-radius:50%;display:flex;align-items:center;justify-content:center;margin:0 auto 12px;border:2px solid rgba(255,255,255,0.35);">
                    <svg width="28" height="28" viewBox="0 0 24 24" fill="none" stroke="white" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"></polyline></svg>
                </div>
                <div style="font-size:20px;font-weight:900;color:#fff;margin-bottom:4px;">✅ Etapa Concluída!</div>
                <div style="font-size:12px;font-weight:700;color:rgba(255,255,255,0.65);letter-spacing:1px;">{etapa_lbl} · {tempo_fmt}</div>
            </div>
            <div style="padding:20px 24px;text-align:center;">
                <div style="font-size:11px;font-weight:800;letter-spacing:2px;color:#9C9490;text-transform:uppercase;margin-bottom:6px;">Pedido {pedido_val}</div>
                <div style="font-size:14px;font-weight:600;color:#5C5450;line-height:1.6;">
                    Próxima etapa: <strong style="color:#1A1714;">{next_lbl}</strong><br>
                    <span style="font-size:12px;color:#9C9490;">O próximo operador iniciará no menu principal.</span>
                </div>
            </div>
        </div>
        </body></html>
        """, height=280, scrolling=False)
        st.markdown("<br style='line-height:0.2'>", unsafe_allow_html=True)
        st.markdown('<div class="btn-iniciar">', unsafe_allow_html=True)
        if st.button("▶  Próximo Pedido", use_container_width=True, key="proxima_home"):
            st.session_state.modal          = None
            st.session_state.pedido         = None
            st.session_state.pedido_validado = False
            st.session_state.etapa_idx      = 0
            st.session_state.acum           = 0
            st.session_state.tela           = "home"
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Modal: pedido concluído ──
    elif st.session_state.modal == "concluido":
        pedido_val = st.session_state.pedido
        tempo_fmt = fmt(st.session_state.acum)
        components.html(f"""
        <!DOCTYPE html><html><head>
        <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@600;700;800;900&display=swap" rel="stylesheet">
        <style>* {{margin:0;padding:0;box-sizing:border-box;}}</style>
        </head><body style="background:transparent;font-family:Nunito,sans-serif;">
        <div style="background:#fff;border-radius:20px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.08);border:1.5px solid #EDE9E4;">
            <div style="background:linear-gradient(135deg,#C8566A 0%,#9E3F52 100%);padding:28px;text-align:center;position:relative;overflow:hidden;">
                <div style="position:absolute;right:-20px;top:-20px;width:100px;height:100px;border-radius:50%;background:rgba(255,255,255,0.07);"></div>
                <div style="position:absolute;left:-10px;bottom:-30px;width:80px;height:80px;border-radius:50%;background:rgba(255,255,255,0.05);"></div>
                <div style="width:60px;height:60px;background:rgba(255,255,255,0.18);border-radius:50%;display:flex;align-items:center;justify-content:center;margin:0 auto 14px;border:2px solid rgba(255,255,255,0.35);position:relative;z-index:1;">
                    <span style="font-size:28px;">🎉</span>
                </div>
                <div style="font-size:22px;font-weight:900;color:#fff;margin-bottom:4px;position:relative;z-index:1;">Pedido Concluído!</div>
                <div style="font-size:12px;font-weight:700;color:rgba(255,255,255,0.65);letter-spacing:1px;position:relative;z-index:1;">Todas as etapas finalizadas</div>
            </div>
            <div style="padding:22px 28px;">
                <div style="display:flex;justify-content:space-between;align-items:center;background:#F7F5F2;border-radius:12px;padding:14px 18px;margin-bottom:10px;">
                    <span style="font-size:11px;font-weight:800;letter-spacing:1.5px;color:#9C9490;text-transform:uppercase;">Pedido</span>
                    <span style="font-family:monospace;font-size:17px;font-weight:800;color:#1A1714;">{pedido_val}</span>
                </div>
                <div style="display:flex;justify-content:space-between;align-items:center;background:#F0F7F3;border-radius:12px;padding:14px 18px;">
                    <span style="font-size:11px;font-weight:800;letter-spacing:1.5px;color:#4A7C59;text-transform:uppercase;">Conferência</span>
                    <span style="font-family:monospace;font-size:17px;font-weight:800;color:#4A7C59;">{tempo_fmt}</span>
                </div>
            </div>
        </div>
        </body></html>
        """, height=300, scrolling=False)
        st.markdown("<br style='line-height:0.2'>", unsafe_allow_html=True)
        st.markdown('<div class="btn-iniciar">', unsafe_allow_html=True)
        if st.button("▶  Próximo Pedido", use_container_width=True):
            st.session_state.modal          = None
            st.session_state.pedido         = None
            st.session_state.pedido_validado = False
            st.session_state.etapa_idx      = 0
            st.session_state.acum           = 0
            st.session_state.tela           = "home"
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)


# ─────────────────────────────────────
#  TELA: ADMIN LOGIN
# ─────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
#  AUTO-REFRESH — detecta novos pedidos no banco e força rerun a cada 30s
#  Só ativo nas telas onde o operador escolhe pedido (home) e operações (gestor).
# ─────────────────────────────────────────────────────────────────────────────
def _hash_pedidos_base():
    """Retorna uma string de fingerprint dos pedidos abertos atuais."""
    try:
        rows = _get("pedidos_base",
                    "select=numero,status,importado_em&order=numero.asc",
                    paginar=True)
        if not isinstance(rows, list):
            return ""
        return "|".join(
            f"{r.get('numero','')}/{r.get('status','')}/{r.get('importado_em','')}"
            for r in rows
        )
    except Exception:
        return ""

@st.fragment(run_every=15)
def _auto_refresh_watcher():
    """
    Fragment silencioso: re-executa a cada 30s.
    Se detectar mudança nos pedidos_base, dispara st.rerun() na página inteira,
    garantindo que o Sistema B reflita imediatamente qualquer upload do Sistema A.
    """
    current_hash = _hash_pedidos_base()
    prev_hash    = st.session_state.get("_pedidos_hash", None)

    if prev_hash is None:
        # Primeira execução — apenas grava o hash inicial
        st.session_state["_pedidos_hash"] = current_hash
        return

    if current_hash != prev_hash:
        st.session_state["_pedidos_hash"] = current_hash
        # Limpa caches para garantir que novos pedidos apareçam imediatamente
        buscar_pedidos_base.clear()
        buscar_pedidos_por_etapa.clear()
        st.toast("📋 Pedidos atualizados — recarregando lista...", icon="🔄")
        st.rerun()


def tela_admin_login():
    erro = st.session_state.erro_senha

    # ── Card visual ──────────────────────────────────────────────────────────
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@600;700&family=Nunito:wght@700;800;900&display=swap');
    .adm-wrap{display:flex;align-items:center;justify-content:center;padding:8px;margin-bottom:8px;}
    .adm-card{width:100%;max-width:420px;background:linear-gradient(160deg,#1a1210 0%,#241816 40%,#1a1210 100%);border-radius:28px;overflow:hidden;position:relative;border:1px solid rgba(200,86,106,0.18);box-shadow:0 32px 80px rgba(0,0,0,0.6),0 0 0 1px rgba(255,255,255,0.04) inset;}
    .adm-accent{height:3px;background:linear-gradient(90deg,transparent 0%,#C8566A 30%,#e8889a 50%,#C8566A 70%,transparent 100%);}
    .adm-orb{position:absolute;border-radius:50%;filter:blur(70px);pointer-events:none;}
    .adm-orb1{width:280px;height:280px;background:#8B2035;opacity:.18;top:-100px;right:-80px;animation:adm-drift 8s ease-in-out infinite;}
    .adm-orb2{width:200px;height:200px;background:#C8566A;opacity:.10;bottom:-60px;left:-60px;animation:adm-drift 10s ease-in-out infinite reverse;}
    @keyframes adm-drift{0%,100%{transform:translate(0,0);}50%{transform:translate(12px,-12px);}}
    .adm-inner{position:relative;z-index:1;padding:36px 32px 28px;}
    .adm-logo{width:68px;height:68px;border-radius:20px;margin:0 auto 24px;background:linear-gradient(145deg,#2a1518,#1a0d0f);border:1px solid rgba(200,86,106,0.35);display:flex;align-items:center;justify-content:center;box-shadow:0 0 0 6px rgba(200,86,106,0.06),0 12px 32px rgba(0,0,0,0.5);animation:adm-glow 4s ease-in-out infinite;}
    @keyframes adm-glow{0%,100%{box-shadow:0 0 0 6px rgba(200,86,106,0.06),0 12px 32px rgba(0,0,0,0.5);}50%{box-shadow:0 0 0 10px rgba(200,86,106,0.10),0 12px 40px rgba(200,86,106,0.25);}}
    .adm-eyebrow{font-size:9px;font-weight:800;letter-spacing:4px;text-transform:uppercase;color:rgba(200,86,106,0.7);text-align:center;margin-bottom:7px;}
    .adm-title{font-family:'Cormorant Garamond',serif;font-size:30px;font-weight:700;color:#fff;text-align:center;letter-spacing:-0.5px;line-height:1;margin-bottom:5px;}
    .adm-sub{font-size:11px;color:rgba(255,255,255,0.3);text-align:center;font-weight:700;letter-spacing:1px;margin-bottom:20px;}
    .adm-sep{height:1px;background:linear-gradient(90deg,transparent,rgba(200,86,106,0.3),transparent);margin-bottom:20px;}
    .adm-status{display:flex;align-items:center;justify-content:center;gap:18px;margin-top:6px;}
    .adm-dot{width:6px;height:6px;border-radius:50%;display:inline-block;}
    .adm-dot-on{background:#4ade80;box-shadow:0 0 8px #4ade80;animation:adm-blink 2s infinite;}
    .adm-dot-off{background:#C8566A;animation:adm-blink 2s 1s infinite;}
    @keyframes adm-blink{0%,100%{opacity:1;}50%{opacity:.3;}}
    .adm-slbl{font-size:10px;font-weight:700;color:rgba(255,255,255,0.2);letter-spacing:.5px;}
    </style>
    <div class="adm-wrap"><div class="adm-card">
      <div class="adm-accent"></div>
      <div class="adm-orb adm-orb1"></div><div class="adm-orb adm-orb2"></div>
      <div class="adm-inner">
        <div class="adm-logo">
          <svg width="30" height="30" viewBox="0 0 24 24" fill="none" stroke="#C8566A" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round">
            <rect x="3" y="11" width="18" height="11" rx="3"/>
            <path d="M7 11V7a5 5 0 0 1 10 0v4"/>
          </svg>
        </div>
        <div class="adm-eyebrow">Acesso Restrito</div>
        <div class="adm-title">Painel Admin</div>
        <div class="adm-sub">Vi Lingerie · Sistema de Produção</div>
        <div class="adm-sep"></div>
        <div class="adm-status">
          <div style="display:flex;align-items:center;gap:6px;"><span class="adm-dot adm-dot-on"></span><span class="adm-slbl">Sistema Online</span></div>
          <div style="display:flex;align-items:center;gap:6px;"><span class="adm-dot adm-dot-off"></span><span class="adm-slbl">Autenticação Necessária</span></div>
        </div>
      </div>
    </div></div>
    """, unsafe_allow_html=True)

    # ── Estilo para os widgets nativos do Streamlit ─────────────────────────
    st.markdown("""
    <style>
    /* Input de senha */
    div[data-testid="stTextInput"] input {
        background: rgba(255,255,255,0.04) !important;
        border: 1.5px solid rgba(255,255,255,0.15) !important;
        border-radius: 14px !important;
        color: #fff !important;
        font-size: 15px !important;
        padding: 12px 16px !important;
        letter-spacing: 4px !important;
    }
    div[data-testid="stTextInput"] input:focus {
        border-color: rgba(200,86,106,0.6) !important;
        box-shadow: 0 0 0 4px rgba(200,86,106,0.10) !important;
    }
    div[data-testid="stTextInput"] label { display: none !important; }

    /* Botão Voltar */
    div[data-testid="stHorizontalBlock"] div[data-testid="column"]:first-child button {
        background: rgba(255,255,255,0.06) !important;
        color: rgba(255,255,255,0.55) !important;
        border: 1.5px solid rgba(255,255,255,0.12) !important;
        border-radius: 14px !important;
        height: 50px !important;
        font-size: 13px !important;
        font-weight: 800 !important;
    }
    div[data-testid="stHorizontalBlock"] div[data-testid="column"]:first-child button:hover {
        background: rgba(255,255,255,0.12) !important;
        color: #fff !important;
    }

    /* Botão Acessar Painel */
    div[data-testid="stHorizontalBlock"] div[data-testid="column"]:last-child button {
        background: linear-gradient(135deg,#C8566A 0%,#8B2035 100%) !important;
        color: #fff !important;
        border: none !important;
        border-radius: 14px !important;
        height: 50px !important;
        font-size: 13px !important;
        font-weight: 800 !important;
        box-shadow: 0 4px 0 rgba(80,10,20,0.5), 0 8px 24px rgba(200,86,106,0.25) !important;
    }
    div[data-testid="stHorizontalBlock"] div[data-testid="column"]:last-child button:hover {
        filter: brightness(1.1) !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── Label e input de senha ──────────────────────────────────────────────
    st.markdown(
        "<div style='font-size:9px;font-weight:800;letter-spacing:3px;text-transform:uppercase;"
        "color:rgba(255,255,255,0.3);margin-bottom:4px;'>🔒 &nbsp;SENHA DE ACESSO</div>",
        unsafe_allow_html=True
    )
    senha_input = st.text_input(
        "_senha_admin", placeholder="········",
        type="password", label_visibility="collapsed",
        key="admin_login_senha"
    )

    # ── Mensagem de erro ────────────────────────────────────────────────────
    if erro:
        st.markdown(
            "<div style='background:rgba(200,86,106,0.12);border:1px solid rgba(200,86,106,0.35);"
            "border-radius:10px;padding:8px 14px;font-size:11px;font-weight:800;"
            "color:rgba(200,86,106,0.95);margin-bottom:4px;'>"
            "❌ &nbsp;Senha incorreta. Tente novamente.</div>",
            unsafe_allow_html=True
        )

    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

    # ── Botões nativos Streamlit ────────────────────────────────────────────
    col_v, col_a = st.columns([1, 2])
    with col_v:
        if st.button("‹ Voltar", use_container_width=True, key="adm_login_voltar"):
            st.session_state.erro_senha = False
            st.session_state.tela = "home"
            st.rerun()
    with col_a:
        if st.button("🔒 Acessar Painel", use_container_width=True, key="adm_login_acessar"):
            if senha_input == ADMIN_SENHA:
                st.session_state.erro_senha = False
                st.session_state.tela = "admin"
            else:
                st.session_state.erro_senha = True
            st.rerun()


# ─────────────────────────────────────
#  TELA: ADMIN PANEL
# ─────────────────────────────────────
def gerar_pdf(regs, op_map, ped_comp, ops_ativ, avg):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)

    ROSA   = colors.HexColor("#C8566A")
    ESCURO = colors.HexColor("#1A1714")
    CLARO  = colors.HexColor("#F7F5F2")
    CINZA  = colors.HexColor("#8C8480")
    VERDE  = colors.HexColor("#4A7C59")
    BEGE   = colors.HexColor("#EDE9E4")
    BRANCO = colors.white

    styles = getSampleStyleSheet()

    def sty(name, **kw):
        return ParagraphStyle(name, **kw)

    S_TITLE   = sty("t",  fontName="Helvetica-Bold", fontSize=22, textColor=ESCURO, spaceAfter=2)
    S_SUB     = sty("s",  fontName="Helvetica",      fontSize=10, textColor=CINZA,  spaceAfter=0)
    S_SECTION = sty("sc", fontName="Helvetica-Bold", fontSize=9,  textColor=CINZA,
                    spaceAfter=6, spaceBefore=16, letterSpacing=1.5)
    S_FOOTER  = sty("f",  fontName="Helvetica",      fontSize=8,  textColor=CINZA, alignment=TA_CENTER)

    story = []
    now_str = now_br().strftime("%d/%m/%Y às %H:%M")

    header_data = [[
        Paragraph("<b><font color='#C8566A' size='18'>Vi</font> LINGERIE</b>", styles["Normal"]),
        Paragraph(f"<font color='#8C8480' size='8'>Gerado em {now_str}</font>",
                  ParagraphStyle("r", alignment=TA_RIGHT))
    ]]
    header_tbl = Table(header_data, colWidths=["60%","40%"])
    header_tbl.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
    ]))
    story.append(header_tbl)
    story.append(HRFlowable(width="100%", thickness=2, color=ROSA, spaceAfter=14))
    story.append(Paragraph("Relatório de Produção", S_TITLE))
    story.append(Paragraph("Desempenho de operadores por etapa do processo produtivo", S_SUB))
    story.append(Spacer(1, 18))

    story.append(Paragraph("RESUMO GERAL", S_SECTION))
    kpi_data = [[
        Paragraph(f"<b><font size='22' color='#C8566A'>{len(ped_comp)}</font></b><br/><font size='8' color='#8C8480'>PEDIDOS CONCLUÍDOS</font>", styles["Normal"]),
        Paragraph(f"<b><font size='22' color='#C8566A'>{len(ops_ativ)}</font></b><br/><font size='8' color='#8C8480'>OPERADORES ATIVOS</font>", styles["Normal"]),
        Paragraph(f"<b><font size='22' color='#C8566A'>{avg}m</font></b><br/><font size='8' color='#8C8480'>TEMPO MÉDIO</font>", styles["Normal"]),
        Paragraph(f"<b><font size='22' color='#C8566A'>{len(regs)}</font></b><br/><font size='8' color='#8C8480'>REGISTROS TOTAIS</font>", styles["Normal"]),
    ]]
    kpi_tbl = Table(kpi_data, colWidths=["25%","25%","25%","25%"])
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), CLARO),
        ("TOPPADDING",    (0,0), (-1,-1), 14), ("BOTTOMPADDING", (0,0), (-1,-1), 14),
        ("LEFTPADDING",   (0,0), (-1,-1), 14), ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(kpi_tbl)
    story.append(Spacer(1, 20))

    if op_map:
        story.append(Paragraph("DESEMPENHO POR OPERADOR", S_SECTION))
        op_header = [
            Paragraph("<b>OPERADOR</b>",      ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=8, textColor=BRANCO)),
            Paragraph("<b>PEDIDOS</b>",        ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=8, textColor=BRANCO, alignment=TA_CENTER)),
            Paragraph("<b>PEÇAS</b>",          ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=8, textColor=BRANCO, alignment=TA_CENTER)),
            Paragraph("<b>TEMPO TOTAL</b>",    ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=8, textColor=BRANCO, alignment=TA_CENTER)),
            Paragraph("<b>TEMPO MÉDIO</b>",    ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=8, textColor=BRANCO, alignment=TA_CENTER)),
            Paragraph("<b>EFICIÊNCIA</b>",     ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=8, textColor=BRANCO, alignment=TA_CENTER)),
        ]
        op_rows_pdf = [op_header]
        for i, (op, d) in enumerate(sorted(op_map.items(), key=lambda x: x[1]["pecas"], reverse=True)):
            n_ped    = len(d["p"])
            t_total  = d["tempo_total"]
            t_medio  = media(d["tempos"])
            pcs      = d["pecas"]
            h_trab   = t_total / 3600 if t_total > 0 else 0
            efic_str = f"{round(pcs/h_trab,1)} pcs/h" if h_trab > 0 and pcs > 0 else "—"
            op_rows_pdf.append([
                Paragraph(f"<b>{op}</b>", ParagraphStyle("o", fontName="Helvetica-Bold", fontSize=9, textColor=ESCURO)),
                Paragraph(str(n_ped),     ParagraphStyle("c", fontSize=9, textColor=ESCURO, alignment=TA_CENTER)),
                Paragraph(str(pcs),       ParagraphStyle("c", fontSize=9, textColor=VERDE,  alignment=TA_CENTER, fontName="Helvetica-Bold")),
                Paragraph(fmt(t_total),   ParagraphStyle("c", fontSize=9, textColor=ESCURO, alignment=TA_CENTER)),
                Paragraph(fmt(t_medio),   ParagraphStyle("c", fontSize=9, textColor=ESCURO, alignment=TA_CENTER)),
                Paragraph(efic_str,       ParagraphStyle("c", fontSize=9, textColor=ESCURO, alignment=TA_CENTER)),
            ])
        op_tbl = Table(op_rows_pdf, colWidths=["28%","12%","12%","16%","16%","16%"])
        op_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), ROSA),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [CLARO, BRANCO]),
            ("GRID",          (0,0), (-1,-1), 0.4, BEGE),
            ("TOPPADDING",    (0,0), (-1,-1), 9),  ("BOTTOMPADDING", (0,0), (-1,-1), 9),
            ("LEFTPADDING",   (0,0), (-1,-1), 10), ("RIGHTPADDING",  (0,0), (-1,-1), 10),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(op_tbl)
        story.append(Spacer(1, 20))

    if regs:
        story.append(Paragraph("HISTÓRICO DE PEDIDOS", S_SECTION))
        ETAPA_NOMES = {"Separacao":"Separação","Conferencia":"Conferência","Embalagem":"Embalagem"}
        hist_header = [
            Paragraph("<b>PEDIDO</b>",    ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=8, textColor=BRANCO)),
            Paragraph("<b>OPERADOR</b>",  ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=8, textColor=BRANCO)),
            Paragraph("<b>ETAPA</b>",     ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=8, textColor=BRANCO, alignment=TA_CENTER)),
            Paragraph("<b>TEMPO</b>",     ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=8, textColor=BRANCO, alignment=TA_CENTER)),
            Paragraph("<b>QTD PÇS</b>",   ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=8, textColor=BRANCO, alignment=TA_CENTER)),
            Paragraph("<b>DATA</b>",      ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=8, textColor=BRANCO, alignment=TA_CENTER)),
        ]
        hist_rows = [hist_header]
        for r in regs[:80]:
            qtd_str = str(r[8]) if r[8] is not None else "—"
            hist_rows.append([
                Paragraph(f"<font name='Courier-Bold' size='8'>{r[1]}</font>", styles["Normal"]),
                Paragraph(f"<font size='8'>{r[2]}</font>", styles["Normal"]),
                Paragraph(ETAPA_NOMES.get(r[3], r[3]), styles["Normal"]),
                Paragraph(f"<font name='Courier' size='8'>{fmt(r[5])}</font>", ParagraphStyle("c", fontSize=8, alignment=TA_CENTER)),
                Paragraph(f"<font name='Courier-Bold' size='8'>{qtd_str}</font>", ParagraphStyle("c", fontSize=8, alignment=TA_CENTER)),
                Paragraph(f"<font size='7' color='#8C8480'>{r[6]}</font>", ParagraphStyle("c", fontSize=7, alignment=TA_CENTER)),
            ])
        hist_tbl = Table(hist_rows, colWidths=["16%","20%","20%","14%","12%","18%"])
        hist_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), ESCURO),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [CLARO, BRANCO]),
            ("GRID",          (0,0), (-1,-1), 0.3, BEGE),
            ("TOPPADDING",    (0,0), (-1,-1), 7), ("BOTTOMPADDING", (0,0), (-1,-1), 7),
            ("LEFTPADDING",   (0,0), (-1,-1), 8), ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(hist_tbl)

    story.append(Spacer(1, 24))
    story.append(HRFlowable(width="100%", thickness=0.5, color=BEGE, spaceAfter=8))
    story.append(Paragraph(f"Vi Lingerie · Relatório gerado automaticamente em {now_str} · Sistema de Produção", S_FOOTER))
    doc.build(story)
    return buf.getvalue()


def tela_admin():
    render_logo()

    c1, c2 = st.columns([3, 1])
    with c1:
        st.markdown("""
        <div style="margin-bottom:1.2rem;">
            <div style="font-size:10px;font-weight:800;letter-spacing:2.5px;text-transform:uppercase;color:#9C9490;margin-bottom:4px;">Painel Administrativo</div>
            <div style="font-size:24px;font-weight:900;color:#1A1714;letter-spacing:-0.3px;">Visão Geral da Produção</div>
        </div>
        """, unsafe_allow_html=True)
    with c2:
        st.markdown('<div class="btn-voltar btn-sm">', unsafe_allow_html=True)
        if st.button("← Sair", use_container_width=True):
            st.session_state.tela = "home"; st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Botão de sincronização manual ────────────────────────────────────────
    col_sync1, col_sync2, col_sync3 = st.columns([1, 2, 1])
    with col_sync2:
        st.markdown("""
        <style>
        .btn-sync > button {
            background: linear-gradient(135deg,#4A7C59,#2d5a3d) !important;
            color: #fff !important; border: none !important;
            border-radius: 12px !important; height: 44px !important;
            font-size: 13px !important; font-weight: 800 !important;
            box-shadow: 0 4px 0 rgba(20,60,30,0.35) !important;
            letter-spacing: 0.5px !important;
        }
        .btn-sync > button:hover { filter: brightness(1.08) !important; }
        </style>""", unsafe_allow_html=True)
        st.markdown('<div class="btn-sync">', unsafe_allow_html=True)
        if st.button("🔄  Sincronizar com Sistema A", use_container_width=True, key="adm_btn_sync"):
            novo_hash = _hash_pedidos_base()
            st.session_state["_pedidos_hash"] = novo_hash
            st.toast("✅ Dados sincronizados com o Sistema A!", icon="🔄")
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    st.markdown("<div style='height:2px'></div>", unsafe_allow_html=True)

    pedidos_base_count = len(buscar_pedidos_base())
    if pedidos_base_count > 0:
        # "Concluídos" = pedidos que os operadores finalizaram a etapa 2 (Conferência)
        regs_concl = _get("registros", "select=pedido&etapa_idx=eq.2", paginar=True)
        nums_concluidos_ops = set(
            r.get("pedido") for r in regs_concl
            if isinstance(regs_concl, list) and r.get("pedido")
        ) if isinstance(regs_concl, list) else set()
        concl_c  = len(nums_concluidos_ops)
        abertos_c = pedidos_base_count - concl_c
        st.markdown(f"""
        <div style="display:flex;align-items:center;gap:14px;background:#F0F7F3;
                    border:1.5px solid #4A7C59;border-radius:12px;padding:14px 20px;margin-bottom:1rem;">
            <div style="width:10px;height:10px;border-radius:50%;background:#4A7C59;
                        box-shadow:0 0 8px #4A7C59;flex-shrink:0;"></div>
            <div style="font-size:13px;font-weight:700;color:#2d5a3d;">
                Base sincronizada via <strong>Programa A</strong>:
                <strong style="color:#1A1714;">{pedidos_base_count}</strong> pedidos
                &nbsp;·&nbsp; <span style="color:#4A7C59;">{abertos_c} em aberto</span>
                &nbsp;·&nbsp; <span style="color:#C8566A;">{concl_c} concluídos pelos operadores</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div style="display:flex;align-items:center;gap:14px;background:#FFFBEB;
                    border:1.5px solid #F59E0B;border-radius:12px;padding:14px 20px;margin-bottom:1rem;">
            <div style="width:10px;height:10px;border-radius:50%;background:#F59E0B;flex-shrink:0;"></div>
            <div style="font-size:13px;font-weight:700;color:#92400E;">
                Nenhuma planilha carregada. Acesse o <strong>Programa A</strong> (Dashboard do Gestor) para importar a carteira de pedidos.
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════
    #  BLOCO 1 — PEDIDOS EM ANDAMENTO
    # ══════════════════════════════════════════════════════════════════
    sessoes_agora = buscar_todas_sessoes_ativas()
    # Exclui sessões pausadas (iniciado_em == 0) — elas aparecem no bloco de Pausas
    sessoes_ativas = [s for s in sessoes_agora if int(s.get("iniciado_em", 0)) != 0]
    n_and = len(sessoes_ativas)

    with st.expander(
        f"⏱️ Pedidos em Andamento — {n_and} ativo(s)" if n_and > 0
        else "⏱️ Pedidos em Andamento — nenhum no momento",
        expanded=n_and > 0
    ):
        if not sessoes_ativas:
            st.markdown("""<div style="background:#F0F7F3;border:1.5px solid #4A7C59;border-radius:12px;
                        padding:20px;text-align:center;">
                <div style="font-size:26px;margin-bottom:6px;">✅</div>
                <div style="font-size:13px;font-weight:700;color:#2d5a3d;">
                    Nenhuma operação em andamento no momento.</div>
            </div>""", unsafe_allow_html=True)
        else:
            ETAPA_COR  = ["#C8566A",  "#3B7DD8",  "#4A7C59"]
            ETAPA_BG   = ["#FFF0F2",  "#F0F5FF",  "#F0F7F3"]
            ETAPA_ICON = ["📦",       "🗃️",        "✅"]

            linhas_html = ""
            for s in sessoes_ativas:
                ped      = str(s.get("pedido",""))
                op       = str(s.get("operador",""))
                eta_idx  = int(s.get("etapa_idx", 0))
                ini_ts   = int(s.get("iniciado_em", 0))
                elapsed  = max(int(time.time()) - ini_ts, 0)
                hh, rem  = divmod(elapsed, 3600)
                mm, ss   = divmod(rem, 60)
                tempo_str = f"{hh:02d}:{mm:02d}:{ss:02d}"
                cor  = ETAPA_COR[eta_idx]
                bg   = ETAPA_BG[eta_idx]
                icon = ETAPA_ICON[eta_idx]
                lbl  = ETAPAS_LBL[eta_idx]
                linhas_html += f"""<tr>
                  <td class="td-ped">{ped}</td>
                  <td class="td-op">{op}</td>
                  <td class="td-c">
                    <span style="background:{bg};color:{cor};font-size:10px;font-weight:800;
                      padding:3px 11px;border-radius:20px;white-space:nowrap;">{icon} {lbl}</span>
                  </td>
                  <td class="td-c" style="font-family:monospace;color:{cor};font-size:13px;
                    font-weight:700;">{tempo_str}</td>
                </tr>"""

            altura = 52 + n_and * 47
            st.markdown(f"""<style>
            *{{margin:0;padding:0;box-sizing:border-box;}}
            body{{background:transparent;font-family:Nunito,sans-serif;}}
            .wrap{{background:#fff;border-radius:14px;border:1.5px solid #EDE9E4;
                   overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.05);}}
            table{{width:100%;border-collapse:collapse;}}
            thead tr{{background:#1A1714;}}
            th{{padding:11px 14px;font-size:9px;font-weight:800;letter-spacing:1.8px;
               text-transform:uppercase;color:rgba(255,255,255,0.40);white-space:nowrap;}}
            th.td-c{{text-align:center;}}
            .td-ped{{padding:12px 14px;font-family:monospace;font-size:13px;
                    font-weight:800;color:#1A1714;white-space:nowrap;}}
            .td-op{{padding:12px 10px;font-size:13px;font-weight:700;color:#1A1714;}}
            .td-c{{padding:12px 10px;text-align:center;}}
            tbody tr{{border-bottom:1px solid #F2EEE9;}}
            tbody tr:last-child{{border-bottom:none;}}
            tbody tr:hover td{{background:#FDFAF9;}}
            </style>
            <div class="wrap">
              <table>
                <thead><tr>
                  <th>Pedido</th>
                  <th>Operador</th>
                  <th class="td-c">Etapa</th>
                  <th class="td-c">Tempo decorrido</th>
                </tr></thead>
                <tbody>{linhas_html}</tbody>
              </table>
            </div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════
    #  BLOCO 1.2 — PEDIDOS EM PAUSA (tempo indeterminado)
    # ══════════════════════════════════════════════════════════════════
    pedidos_pausados = buscar_pedidos_pausados()
    n_paus = len(pedidos_pausados)

    with st.expander(
        f"⏸ Pedidos em Pausa — {n_paus} pedido(s) pausado(s)" if n_paus > 0
        else "⏸ Pedidos em Pausa — nenhum pausado no momento",
        expanded=n_paus > 0
    ):
        if not pedidos_pausados:
            st.markdown("""<div style="background:#FFF8F0;border:1.5px solid #E07B3A33;
                        border-radius:12px;padding:20px;text-align:center;">
                <div style="font-size:26px;margin-bottom:6px;">▶️</div>
                <div style="font-size:13px;font-weight:700;color:#B85C20;">
                    Nenhum pedido pausado no momento.</div>
            </div>""", unsafe_allow_html=True)
        else:
            ETAPA_TAG_P = {
                0: '<span style="background:#EBF0FB;color:#3B5EC6;padding:2px 9px;border-radius:100px;font-size:10px;font-weight:800;">📦 Separação</span>',
                1: '<span style="background:#E8F2EC;color:#4A7C59;padding:2px 9px;border-radius:100px;font-size:10px;font-weight:800;">🗃️ Embalagem</span>',
                2: '<span style="background:#FBF2E6;color:#C47B2A;padding:2px 9px;border-radius:100px;font-size:10px;font-weight:800;">✅ Conferência</span>',
            }
            linhas_paus = ""
            for p in pedidos_pausados:
                eta_tag   = ETAPA_TAG_P.get(p["etapa_idx"], str(p["etapa_idx"]))
                tempo_str = fmt(p["tempo_pausado"]) if p["tempo_pausado"] else "—"
                # Separa data e hora do campo pausado_em (formato "DD/MM/YYYY HH:MM")
                pem = str(p["pausado_em"] or "—")
                if " " in pem:
                    data_paus, hora_paus = pem.split(" ", 1)
                else:
                    data_paus, hora_paus = pem, "—"
                motivo_str = p["motivo"] if p["motivo"] else '<span style="color:#C0BAB4;font-style:italic;">—</span>'
                linhas_paus += f"""<tr>
                  <td class="td-ped">{p['pedido']}</td>
                  <td class="td-op">{p['operador']}</td>
                  <td class="td-c">{eta_tag}</td>
                  <td class="td-c" style="color:#E07B3A;font-family:monospace;font-size:12px;font-weight:700;">{tempo_str}</td>
                  <td class="td-c" style="font-size:12px;font-weight:800;color:#1A1714;">{data_paus}</td>
                  <td class="td-c" style="font-size:13px;font-weight:900;color:#C8566A;font-family:monospace;">{hora_paus}</td>
                  <td class="td-mot">{motivo_str}</td>
                </tr>"""

            altura_p = 54 + n_paus * 48 + 16
            st.markdown(f"""<style>
            .wrap-paus{{background:#fff;border-radius:14px;border:1.5px solid #F5DECB;
                       overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.05);}}
            .wrap-paus table{{width:100%;border-collapse:collapse;}}
            .wrap-paus thead tr{{background:#1A1714;}}
            .wrap-paus th{{padding:11px 10px;font-size:9px;font-weight:800;letter-spacing:1.6px;
               text-transform:uppercase;color:rgba(255,255,255,0.40);white-space:nowrap;text-align:center;}}
            .wrap-paus th.th-l{{text-align:left;padding-left:14px;}}
            .wrap-paus .td-ped{{padding:12px 14px;font-family:monospace;font-size:13px;
                    font-weight:800;color:#1A1714;white-space:nowrap;}}
            .wrap-paus .td-op{{padding:12px 10px;font-size:13px;font-weight:700;color:#1A1714;}}
            .wrap-paus .td-c{{padding:12px 10px;text-align:center;}}
            .wrap-paus .td-mot{{padding:12px 10px;font-size:12px;color:#5C5450;font-weight:600;}}
            .wrap-paus tbody tr{{border-bottom:1px solid #FDF0E8;}}
            .wrap-paus tbody tr:last-child{{border-bottom:none;}}
            .wrap-paus tbody tr:hover td{{background:#FFF8F2;}}
            </style>
            <div class="wrap-paus">
              <table>
                <thead><tr>
                  <th class="th-l">Pedido</th>
                  <th class="th-l">Operador</th>
                  <th>Etapa</th>
                  <th>Tempo acum.</th>
                  <th style="color:#D4A45A;">Data da pausa</th>
                  <th style="color:#F4965A;">Hora da pausa</th>
                  <th class="th-l" style="color:rgba(255,255,255,0.30);">Motivo</th>
                </tr></thead>
                <tbody>{linhas_paus}</tbody>
              </table>
            </div>""", unsafe_allow_html=True)

            # Badge de contagem
            st.markdown(f"""
            <div style="margin-top:10px;display:flex;align-items:center;gap:10px;
                        background:#FFF8F0;border:1.5px solid #E07B3A44;border-radius:10px;
                        padding:10px 16px;">
              <span style="font-size:22px;">⏸</span>
              <div>
                <span style="font-size:13px;font-weight:900;color:#B85C20;">
                  {n_paus} pedido(s) aguardando retomada pelo operador
                </span>
                <span style="font-size:11px;font-weight:600;color:#9C9490;margin-left:8px;">
                  · O operador retoma clicando em ▶ Retomar no painel de operações
                </span>
              </div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════
    #  BLOCO 1.5 — RASTREAR PEDIDO POR NÚMERO
    # ══════════════════════════════════════════════════════════════════
    with st.expander("🔍 Rastrear Pedido por Número", expanded=True):

        for _k, _v in [("rastr_num", ""), ("rastr_info", None)]:
            if _k not in st.session_state: st.session_state[_k] = _v

        col_ri, col_rb = st.columns([3, 1])
        with col_ri:
            rastr_input = st.text_input("_rastr_num",
                placeholder="Digite o número do pedido...",
                label_visibility="collapsed", key="rastr_input_num")
        with col_rb:
            st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
            if st.button("🔍  Buscar", use_container_width=True, key="rastr_btn_buscar"):
                num_r = rastr_input.strip()
                if num_r:
                    st.session_state.rastr_num  = num_r
                    st.session_state.rastr_info = buscar_status_completo_pedido(num_r)
                    # Busca também est_alocado e vr_alocado para exibir no rastreio
                    _rastr_ped = _get("pedidos_base",
                        f"numero=eq.{num_r}&select=est_alocado,vr_alocado")
                    if isinstance(_rastr_ped, list) and _rastr_ped:
                        st.session_state.rastr_est  = _rastr_ped[0].get("est_alocado")
                        st.session_state.rastr_vr   = _rastr_ped[0].get("vr_alocado")
                    else:
                        st.session_state.rastr_est  = None
                        st.session_state.rastr_vr   = None
                    st.rerun()

        info_r = st.session_state.rastr_info
        num_r  = st.session_state.rastr_num
        if "rastr_est" not in st.session_state: st.session_state.rastr_est = None
        if "rastr_vr"  not in st.session_state: st.session_state.rastr_vr  = None

        if info_r is not None:
            base_r = info_r.get("base_status", "nao_encontrado")
            cli_r  = info_r.get("cliente", "")

            if base_r == "nao_encontrado":
                st.markdown(f"""<div style="background:#FFFBEB;border:2px solid #F59E0B;border-radius:12px;
                            padding:14px 20px;text-align:center;margin-top:8px;">
                  <div style="font-size:20px;margin-bottom:4px;">❓</div>
                  <div style="font-size:13px;font-weight:800;color:#92400E;">
                    Pedido <span style="font-family:monospace;">#{num_r}</span> não encontrado.</div>
                </div>""", unsafe_allow_html=True)
            else:
                # Monta linha do stepper para cada etapa
                ETAPA_ICON  = ["📦", "🗃️", "✅"]
                ETAPA_COR_R = ["#C8566A", "#3B7DD8", "#4A7C59"]

                etapas_r    = info_r.get("etapas", [])
                etapa_atual = None
                for e in etapas_r:
                    if e.get("em_andamento"):
                        etapa_atual = e
                        break

                # Sessão ativa?
                sess_r = _get("sessoes_ativas",
                    f"pedido=eq.{num_r}&select=etapa_idx,operador,iniciado_em")
                sess_ativa = sess_r[0] if isinstance(sess_r, list) and sess_r else None

                # Monta HTML das etapas — com operador, data e tempo gasto
                etapas_html = ""
                for e in etapas_r:
                    idx_e   = e["idx"]
                    lbl_e   = e["label"]
                    feita   = e.get("feita", False)
                    andando = e.get("em_andamento", False)
                    op_e    = e.get("operador", "")
                    data_e  = e.get("data", "")
                    tempo_e = e.get("tempo")  # segundos gastos (só existe se feita)

                    # Formata tempo se disponível
                    if tempo_e:
                        _th, _tr = divmod(int(tempo_e), 3600); _tm, _ts = divmod(_tr, 60)
                        tempo_fmt = f"{_th:02d}:{_tm:02d}:{_ts:02d}"
                    else:
                        tempo_fmt = None

                    if feita:
                        bg_e  = "#F0F7F3"; brd_e = "#4A7C59"; cor_e = "#2d5a3d"
                        ic_e  = "✅"; st_e = "Concluída"
                        # Linha de detalhe: operador + data + tempo
                        det_partes = []
                        if op_e:    det_partes.append(f'<strong style="color:#2d5a3d">{op_e}</strong>')
                        if data_e:  det_partes.append(data_e)
                        if tempo_fmt: det_partes.append(f'⏱ {tempo_fmt}')
                        det_e = f'<span style="color:#9C9490;font-size:10px;">{" · ".join(det_partes)}</span>' if det_partes else ""
                    elif andando:
                        bg_e  = "#FFF7ED"; brd_e = "#E07B3A"; cor_e = "#92400E"
                        ic_e  = "⏱️"; st_e = "Em andamento"
                        det_e = f'<span style="color:#E07B3A;font-size:10px;font-weight:800;">👷 {op_e}</span>' if op_e else ""
                    else:
                        bg_e  = "#F7F5F2"; brd_e = "#DDD8D2"; cor_e = "#9C9490"
                        ic_e  = ETAPA_ICON[idx_e]; st_e = "Aguardando"
                        det_e = ""

                    # Destaque visual para a etapa onde o pedido está parado
                    borda_extra = "border-left:4px solid " + brd_e + ";" if (andando or (not feita and any(x.get("feita") for x in etapas_r[:idx_e]))) else ""

                    etapas_html += f"""
                    <div style="background:{bg_e};border:1.5px solid {brd_e};{borda_extra}border-radius:12px;
                                padding:12px 16px;margin-bottom:6px;
                                display:flex;align-items:center;gap:12px;">
                      <span style="font-size:20px;flex-shrink:0;">{ic_e}</span>
                      <div style="flex:1;min-width:0;">
                        <div style="font-size:12px;font-weight:900;color:{cor_e};margin-bottom:2px;">{lbl_e}</div>
                        {det_e}
                      </div>
                      <span style="font-size:10px;font-weight:800;color:{cor_e};
                        background:rgba(0,0,0,0.07);padding:3px 12px;border-radius:20px;white-space:nowrap;">{st_e}</span>
                    </div>"""

                # Status geral do pedido
                cor_ped  = "#4A7C59" if base_r == "aberto" else "#C8566A"
                bg_ped   = "#F0F7F3" if base_r == "aberto" else "#FFF0F2"
                lbl_ped  = "Aberto"  if base_r == "aberto" else "Concluído"
                icon_ped = "🟢"      if base_r == "aberto" else "🔴"

                # Monta frase resumo: "Parado em X etapa" ou "Todas as etapas concluídas"
                etapas_feitas = [e for e in etapas_r if e.get("feita")]
                etapas_pend   = [e for e in etapas_r if not e.get("feita") and not e.get("em_andamento")]
                etapa_andando = next((e for e in etapas_r if e.get("em_andamento")), None)

                if base_r == "concluido":
                    resumo_html = '''<div style="background:#F0F7F3;border:1.5px solid #4A7C59;border-radius:10px;
                        padding:10px 16px;margin-bottom:12px;font-size:12px;font-weight:800;color:#2d5a3d;text-align:center;">
                        ✅ Todas as 3 etapas concluídas — pedido finalizado
                    </div>'''
                elif etapa_andando:
                    op_and_res = etapa_andando.get("operador","")
                    resumo_html = f'''<div style="background:#FFF7ED;border:1.5px solid #E07B3A;border-radius:10px;
                        padding:10px 16px;margin-bottom:12px;font-size:12px;font-weight:800;color:#92400E;text-align:center;">
                        ⏱️ Em andamento agora · <strong>{etapa_andando["label"]}</strong>
                        {f" · 👷 {op_and_res}" if op_and_res else ""}
                    </div>'''
                elif etapas_feitas:
                    prox = etapas_pend[0]["label"] if etapas_pend else "—"
                    ult  = etapas_feitas[-1]
                    resumo_html = f'''<div style="background:#F0F5FF;border:1.5px solid #3B7DD8;border-radius:10px;
                        padding:10px 16px;margin-bottom:12px;font-size:12px;font-weight:800;color:#1e3a8a;text-align:center;">
                        🔵 Parado após <strong>{ult["label"]}</strong> · próxima etapa: <strong>{prox}</strong>
                    </div>'''
                else:
                    resumo_html = '''<div style="background:#F7F5F2;border:1.5px solid #DDD8D2;border-radius:10px;
                        padding:10px 16px;margin-bottom:12px;font-size:12px;font-weight:800;color:#9C9490;text-align:center;">
                        ⚪ Nenhuma etapa iniciada ainda
                    </div>'''  

                # Formata est_alocado e vr_alocado para exibição
                _est_r = st.session_state.rastr_est
                _vr_r  = st.session_state.rastr_vr
                _est_html = f'<strong style="font-family:monospace;">{int(float(_est_r))}</strong> itens' if _est_r is not None else "—"
                _vr_html  = f'R$ {float(_vr_r):,.2f}'.replace(",","X").replace(".",",").replace("X",".") if _vr_r else "—"
                dados_ped_html = f'''
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:12px;">
                  <div style="background:#F0F5FF;border-radius:10px;padding:10px 14px;">
                    <div style="font-size:9px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:#3B7DD8;margin-bottom:3px;">Qtd. Itens</div>
                    <div style="font-size:15px;font-weight:800;color:#1A1714;">{_est_html}</div>
                    <div style="font-size:9px;color:#9C9490;margin-top:1px;">Est. Alocado</div>
                  </div>
                  <div style="background:#F0F7F3;border-radius:10px;padding:10px 14px;">
                    <div style="font-size:9px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:#4A7C59;margin-bottom:3px;">Valor</div>
                    <div style="font-size:15px;font-weight:800;color:#1A1714;">{_vr_html}</div>
                    <div style="font-size:9px;color:#9C9490;margin-top:1px;">Vr. Alocado</div>
                  </div>
                </div>'''

                # Aviso se tem sessão ativa agora
                andamento_html = ""
                if sess_ativa:
                    op_and = sess_ativa.get("operador","")
                    eta_and = int(sess_ativa.get("etapa_idx",0))
                    ini_and = int(sess_ativa.get("iniciado_em",0))
                    el_and  = max(int(time.time()) - ini_and, 0)
                    hh_a, rr_a = divmod(el_and, 3600); mm_a, ss_a = divmod(rr_a, 60)
                    andamento_html = f"""
                    <div style="background:#FFF7ED;border:1.5px solid #E07B3A;border-radius:10px;
                                padding:10px 16px;margin-bottom:10px;font-size:12px;font-weight:700;color:#92400E;">
                      ⏱️ <strong>{op_and}</strong> está trabalhando em
                      <strong>{ETAPAS_LBL[eta_and]}</strong>
                      há <strong style="font-family:monospace;">{hh_a:02d}:{mm_a:02d}:{ss_a:02d}</strong>
                    </div>"""

                h_card = 80 + 90 + len(etapas_r) * 62 + (40 if sess_ativa else 0)
                st.markdown(f"""<div style="background:#fff;border:1.5px solid #EDE9E4;border-radius:16px;
                            padding:16px 18px;box-shadow:0 2px 12px rgba(0,0,0,0.05);margin-top:8px;">
                  <div style="display:flex;align-items:center;gap:12px;margin-bottom:14px;flex-wrap:wrap;">
                    <span style="font-family:'DM Mono',monospace;font-size:18px;font-weight:800;
                      color:#1A1714;">#{num_r}</span>
                    <span style="flex:1;font-size:13px;font-weight:700;color:#5C5450;min-width:0;
                      overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">{cli_r}</span>
                    <span style="background:{bg_ped};color:{cor_ped};font-size:11px;font-weight:800;
                      padding:4px 14px;border-radius:20px;flex-shrink:0;">{icon_ped} {lbl_ped}</span>
                  </div>
                  {resumo_html}
                  {dados_ped_html}
                  {andamento_html}
                  {etapas_html}
                </div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════
    #  BLOCO 2 — ALTERAR STATUS DE PEDIDO
    # ══════════════════════════════════════════════════════════════════
    with st.expander("↩️ Voltar Etapa do Pedido", expanded=False):

        st.markdown("""<div style="background:#F0F5FF;border:1.5px solid #3B7DD8;border-radius:10px;
                    padding:11px 16px;font-size:12px;font-weight:700;color:#1e3a6e;">
            ↩️ Apaga o registro da <strong>última etapa concluída</strong> do pedido,
            voltando-o para a etapa anterior. O operador original é mantido.
        </div>""", unsafe_allow_html=True)

        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

        for _k, _v in [
            ("vep_busca_num",  ""),
            ("vep_info",       None),
            ("vep_confirm",    False),
        ]:
            if _k not in st.session_state:
                st.session_state[_k] = _v

        col_vi, col_vb = st.columns([3, 1])
        with col_vi:
            vep_input = st.text_input(
                "Número do pedido", placeholder="Ex: 48944",
                label_visibility="collapsed", key="vep_input_num"
            )
        with col_vb:
            st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
            if st.button("🔍  Buscar", use_container_width=True, key="vep_btn_buscar"):
                n = vep_input.strip()
                if n:
                    st.session_state.vep_busca_num = n
                    st.session_state.vep_info      = buscar_status_completo_pedido(n)
                    st.session_state.vep_confirm   = False
                    st.rerun()

        vep_info = st.session_state.vep_info
        vep_num  = st.session_state.vep_busca_num

        if vep_info is not None:
            vep_base = vep_info.get("base_status", "nao_encontrado")
            vep_cli  = vep_info.get("cliente", "")
            vep_etps = vep_info.get("etapas", [])

            if vep_base == "nao_encontrado":
                st.markdown(f"""<div style="background:#FFFBEB;border:2px solid #F59E0B;border-radius:12px;
                            padding:14px 20px;text-align:center;margin-top:8px;">
                  <div style="font-size:20px;margin-bottom:4px;">❓</div>
                  <div style="font-size:13px;font-weight:800;color:#92400E;">
                    Pedido <span style="font-family:monospace;">#{vep_num}</span> não encontrado.</div>
                </div>""", unsafe_allow_html=True)
            else:
                # Última etapa concluída
                etps_feitas = [e for e in vep_etps if e.get("feita")]

                if not etps_feitas:
                    st.markdown(f"""<div style="background:#F5F5F5;border:1.5px solid #D1D5DB;border-radius:12px;
                                padding:16px 20px;text-align:center;margin-top:8px;">
                      <div style="font-size:20px;margin-bottom:4px;">○</div>
                      <div style="font-size:13px;font-weight:800;color:#6B7280;">
                        Pedido <span style="font-family:monospace;">#{vep_num}</span>
                        não tem nenhuma etapa concluída para voltar.</div>
                    </div>""", unsafe_allow_html=True)
                else:
                    ultima = etps_feitas[-1]
                    ult_idx = ultima["idx"]
                    ult_lbl = ultima["label"]
                    ult_op  = ultima.get("operador", "—")
                    ult_dt  = ultima.get("data", "")

                    # Monta visual das etapas com destaque na que será removida
                    etapas_html = ""
                    for e in vep_etps:
                        if e["idx"] == ult_idx:
                            etapas_html += (
                                f'<div style="background:#FEF2F2;border:2px solid #FCA5A5;'
                                f'border-radius:10px;padding:10px 14px;display:flex;'
                                f'align-items:center;gap:10px;">'
                                f'<div style="font-size:18px;">🗑️</div>'
                                f'<div>'
                                f'<div style="font-size:12px;font-weight:900;color:#991B1B;">{ult_lbl}</div>'
                                f'<div style="font-size:11px;font-weight:600;color:#B91C1C;">'
                                f'por {ult_op}'
                                f'{" · " + ult_dt if ult_dt else ""}</div>'
                                f'</div>'
                                f'<div style="margin-left:auto;font-size:10px;font-weight:800;'
                                f'color:#DC2626;background:#FEE2E2;padding:2px 10px;'
                                f'border-radius:20px;">será removida</div>'
                                f'</div>'
                            )
                        elif e.get("feita"):
                            etapas_html += (
                                f'<div style="background:#F0F7F3;border:1.5px solid #86EFAC;'
                                f'border-radius:10px;padding:10px 14px;display:flex;'
                                f'align-items:center;gap:10px;">'
                                f'<div style="font-size:16px;">✅</div>'
                                f'<div style="font-size:12px;font-weight:800;color:#4A7C59;">{e["label"]}</div>'
                                f'<div style="margin-left:auto;font-size:10px;font-weight:700;color:#4A7C59;">'
                                f'por {e.get("operador","—")}</div>'
                                f'</div>'
                            )
                        else:
                            etapas_html += (
                                f'<div style="background:#F5F5F5;border:1.5px solid #E5E7EB;'
                                f'border-radius:10px;padding:10px 14px;display:flex;'
                                f'align-items:center;gap:10px;opacity:0.5;">'
                                f'<div style="font-size:16px;">○</div>'
                                f'<div style="font-size:12px;font-weight:800;color:#9C9490;">{e["label"]}</div>'
                                f'</div>'
                            )

                    st.markdown(f"""<div style="background:#fff;border:1.5px solid #EDE9E4;border-radius:14px;
                                padding:14px 18px;box-shadow:0 2px 10px rgba(0,0,0,0.05);margin-top:6px;">
                      <div style="display:flex;align-items:center;gap:10px;margin-bottom:12px;">
                        <span style="font-family:monospace;font-size:17px;font-weight:900;color:#1A1714;">#{vep_num}</span>
                        <span style="font-size:13px;font-weight:700;color:#5C5450;flex:1;
                          overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">{vep_cli}</span>
                      </div>
                      <div style="display:flex;flex-direction:column;gap:6px;">
                        {etapas_html}
                      </div>
                    </div>""", unsafe_allow_html=True)

                    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

                    if not st.session_state.vep_confirm:
                        if st.button(
                            f"↩️  Voltar etapa  —  remover {ult_lbl}",
                            use_container_width=True, key="vep_btn_voltar"
                        ):
                            st.session_state.vep_confirm = True
                            st.rerun()
                    else:
                        st.markdown(f"""<div style="background:#FEF2F2;border:2px solid #F87171;border-radius:12px;
                                    padding:12px 18px;text-align:center;">
                          <div style="font-size:13px;font-weight:800;color:#991B1B;margin-bottom:4px;">
                            Confirmar remoção da etapa?</div>
                          <div style="font-size:12px;color:#B91C1C;font-weight:700;">
                            <span style="font-family:monospace;">#{vep_num}</span>
                            &nbsp;·&nbsp; {ult_lbl} &nbsp;·&nbsp; por {ult_op}
                          </div>
                        </div>""", unsafe_allow_html=True)

                        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
                        st.markdown("""
                        <style>
                        .btn-vep-conf > button {
                            background: linear-gradient(135deg,#DC2626,#991B1B) !important;
                            color:#fff !important; border:none !important;
                            border-radius:10px !important; height:46px !important;
                            font-size:13px !important; font-weight:800 !important;
                            box-shadow: 0 4px 0 rgba(100,10,10,0.35) !important;
                        }
                        .btn-voltar > button { height:46px !important; }
                        </style>""", unsafe_allow_html=True)
                        cv1, cv2 = st.columns(2)
                        with cv1:
                            st.markdown('<div class="btn-vep-conf">', unsafe_allow_html=True)
                            if st.button("✓  Sim, voltar etapa", use_container_width=True, key="vep_btn_confirmar"):
                                # Apaga o último registro dessa etapa para esse pedido
                                rows_del = _get("registros",
                                    f"pedido=eq.{vep_num}&etapa_idx=eq.{ult_idx}&select=id&order=id.desc&limit=1")
                                if isinstance(rows_del, list) and rows_del:
                                    reg_id = rows_del[0].get("id")
                                    if reg_id:
                                        _delete("registros", f"id=eq.{reg_id}")
                                # Se estava marcado como concluído, volta para aberto
                                if vep_base == "concluido":
                                    _patch("pedidos_base", f"numero=eq.{vep_num}", {"status": "aberto"})
                                # Reseta estado
                                st.session_state.vep_info    = None
                                st.session_state.vep_confirm = False
                                st.session_state.vep_busca_num = ""
                                st.toast(f"↩️ Pedido #{vep_num} · etapa {ult_lbl} removida", icon="↩️")
                                st.rerun()
                            st.markdown('</div>', unsafe_allow_html=True)
                        with cv2:
                            if st.button("✕  Cancelar", use_container_width=True, key="vep_btn_cancelar"):
                                st.session_state.vep_confirm = False
                                st.rerun()
    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

    # ── Expander: Zerar Pedido ───────────────────────────────────────────────
    with st.expander("🔄 Zerar Pedido (Voltar à Estaca Zero)", expanded=False):

        st.markdown("""<div style="background:#FEF2F2;border:1.5px solid #FCA5A5;border-radius:10px;
                    padding:11px 16px;font-size:12px;font-weight:700;color:#991B1B;">
            ⚠️ Esta ação <strong>apaga todos os registros e sessões</strong> do pedido no Sistema B.
            O pedido voltará ao estado inicial — como se nunca tivesse sido trabalhado.
        </div>""", unsafe_allow_html=True)

        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

        for _k, _v in [
            ("zer_num",     ""),
            ("zer_info",    None),
            ("zer_confirm", False),
        ]:
            if _k not in st.session_state:
                st.session_state[_k] = _v

        col_zi, col_zb = st.columns([3, 1])
        with col_zi:
            zer_input = st.text_input(
                "Número do pedido",
                placeholder="Ex: 48944",
                label_visibility="collapsed",
                key="zer_input_num"
            )
        with col_zb:
            st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
            if st.button("🔍  Buscar", use_container_width=True, key="zer_btn_buscar"):
                n = zer_input.strip()
                if n:
                    _info = buscar_status_completo_pedido(n)
                    st.session_state.zer_num     = n
                    st.session_state.zer_info    = _info
                    st.session_state.zer_confirm = False
                    st.rerun()

        zer_info = st.session_state.zer_info
        zer_num  = st.session_state.zer_num

        if zer_info is not None:
            zer_base = zer_info.get("base_status", "nao_encontrado")
            zer_cli  = zer_info.get("cliente", "")
            zer_etps = zer_info.get("etapas", [])

            if zer_base == "nao_encontrado":
                st.markdown(f"""<div style="background:#FFFBEB;border:2px solid #F59E0B;border-radius:12px;
                            padding:14px 20px;text-align:center;margin-top:8px;">
                  <div style="font-size:20px;margin-bottom:4px;">❓</div>
                  <div style="font-size:13px;font-weight:800;color:#92400E;">
                    Pedido <span style="font-family:monospace;">#{zer_num}</span> não encontrado.</div>
                </div>""", unsafe_allow_html=True)
            else:
                # Conta etapas feitas
                etps_feitas  = [e for e in zer_etps if e.get("feita")]
                etps_and     = [e for e in zer_etps if e.get("em_andamento")]
                n_feitas     = len(etps_feitas)
                n_and        = len(etps_and)

                etapas_html = ""
                for e in zer_etps:
                    if e.get("feita"):
                        etapas_html += f'<span style="background:#E8F2EC;color:#4A7C59;font-size:10px;font-weight:800;padding:2px 10px;border-radius:20px;margin-right:4px;">✓ {e["label"]}</span>'
                    elif e.get("em_andamento"):
                        etapas_html += f'<span style="background:#FFF8E6;color:#B45309;font-size:10px;font-weight:800;padding:2px 10px;border-radius:20px;margin-right:4px;">⏱ {e["label"]}</span>'
                    else:
                        etapas_html += f'<span style="background:#F5F5F5;color:#9C9490;font-size:10px;font-weight:800;padding:2px 10px;border-radius:20px;margin-right:4px;">○ {e["label"]}</span>'
                if not etapas_html:
                    etapas_html = '<span style="font-size:11px;color:#9C9490;font-weight:600;">Nenhuma etapa registrada</span>'

                cor_st = "#4A7C59" if zer_base == "aberto" else "#C8566A"
                bg_st  = "#F0F7F3" if zer_base == "aberto" else "#FFF0F2"
                lbl_st = "Aberto"  if zer_base == "aberto" else "Concluído"

                st.markdown(f"""<div style="background:#fff;border:1.5px solid #EDE9E4;border-radius:14px;
                            padding:14px 20px;box-shadow:0 2px 10px rgba(0,0,0,0.05);margin-top:8px;">
                  <div style="display:flex;align-items:center;gap:12px;margin-bottom:10px;flex-wrap:wrap;">
                    <span style="font-family:monospace;font-size:18px;font-weight:800;color:#1A1714;">#{zer_num}</span>
                    <span style="flex:1;font-size:13px;font-weight:700;color:#5C5450;min-width:0;
                      overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">{zer_cli}</span>
                    <span style="background:{bg_st};color:{cor_st};font-size:11px;font-weight:800;
                      padding:4px 14px;border-radius:20px;flex-shrink:0;">{lbl_st}</span>
                  </div>
                  <div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;margin-bottom:10px;">
                    <span style="font-size:9px;font-weight:800;letter-spacing:1.5px;
                      text-transform:uppercase;color:#9C9490;margin-right:2px;">Etapas:</span>
                    {etapas_html}
                  </div>
                  <div style="background:#FEF2F2;border-radius:8px;padding:8px 12px;font-size:11px;
                              font-weight:700;color:#991B1B;text-align:center;">
                    Serão apagados: <strong>{n_feitas} registro(s)</strong> de etapas
                    {"e <strong>" + str(n_and) + " sessão(ões) ativa(s)</strong>" if n_and else ""}.
                    Status voltará para <strong>Aberto</strong>.
                  </div>
                </div>""", unsafe_allow_html=True)

                st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

                if not st.session_state.zer_confirm:
                    if st.button("🔄  Zerar este pedido", use_container_width=True, key="zer_btn_zerar"):
                        st.session_state.zer_confirm = True
                        st.rerun()
                else:
                    st.markdown(f"""<div style="background:#FEF2F2;border:2px solid #F87171;border-radius:12px;
                                padding:13px 20px;text-align:center;">
                      <div style="font-size:13px;font-weight:800;color:#991B1B;margin-bottom:3px;">
                        Tem certeza? Esta ação não pode ser desfeita.</div>
                      <div style="font-size:12px;color:#B91C1C;font-weight:700;">
                        Todos os registros do pedido <span style="font-family:monospace;">#{zer_num}</span>
                        serão apagados do Sistema B.
                      </div>
                    </div>""", unsafe_allow_html=True)

                    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
                    st.markdown("""
                    <style>
                    .btn-confirm-del > button {
                        background: linear-gradient(135deg,#DC2626,#991B1B) !important;
                        color:#fff !important; border:none !important;
                        border-radius:10px !important; height:48px !important;
                        font-size:13px !important; font-weight:800 !important;
                        box-shadow: 0 4px 0 rgba(100,10,10,0.40) !important;
                    }
                    .btn-confirm-del > button:hover { transform:translateY(-1px) !important; }
                    .btn-voltar > button { height:48px !important; }
                    </style>""", unsafe_allow_html=True)
                    cz1, cz2 = st.columns(2)
                    with cz1:
                        st.markdown('<div class="btn-confirm-del">', unsafe_allow_html=True)
                        if st.button("✓  Sim, zerar pedido", use_container_width=True, key="zer_btn_confirmar"):
                            # 1. Apaga todos os registros de etapas do pedido
                            _delete("registros", f"pedido=eq.{zer_num}")
                            # 2. Remove qualquer sessão ativa
                            _delete("sessoes_ativas", f"pedido=eq.{zer_num}")
                            # 3. Volta o status para aberto no pedidos_base
                            _patch("pedidos_base", f"numero=eq.{zer_num}", {"status": "aberto"})
                            # Reset de estado
                            st.session_state.zer_info    = None
                            st.session_state.zer_confirm = False
                            st.session_state.zer_num     = ""
                            st.toast(f"✅ Pedido #{zer_num} zerado com sucesso!", icon="🔄")
                            st.rerun()
                        st.markdown('</div>', unsafe_allow_html=True)
                    with cz2:
                        if st.button("✕  Cancelar", use_container_width=True, key="zer_btn_cancelar"):
                            st.session_state.zer_confirm = False
                            st.rerun()

    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════
    #  BLOCO — ADICIONAR PEDIDO MANUAL / VIA XLSX
    # ══════════════════════════════════════════════════════════════════
    with st.expander("➕ Adicionar Pedido", expanded=False):

        # Inicializa session_state
        for _k, _v in [
            ("novo_ped_num",""), ("novo_ped_cli",""), ("novo_ped_prod",""),
            ("novo_ped_qtd",0),  ("novo_ped_vr",0.0), ("novo_ped_obs",""),
            ("novo_ped_ok",False), ("novo_ped_erro",""),
            ("novo_ped_xlsx_preview", None),  # lista de dicts lidos do xlsx
        ]:
            if _k not in st.session_state: st.session_state[_k] = _v

        # ── CSS botões ────────────────────────────────────────────────
        st.markdown("""<style>
        .btn-novo-ped > button {
            background: linear-gradient(135deg,#3B7DD8,#1e3a8a) !important;
            color:#fff !important; border:none !important;
            border-radius:12px !important; height:48px !important;
            font-size:14px !important; font-weight:900 !important;
            box-shadow: 0 4px 0 rgba(20,40,100,0.35) !important;
        }
        .btn-novo-ped > button:hover { filter:brightness(1.08) !important; }
        /* Abas do bloco Adicionar Pedido */
        div[data-testid="stTabs"] button[role="tab"] {
            color: #5C5450 !important;
            font-weight: 800 !important;
            font-size: 13px !important;
        }
        div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
            color: #1A1714 !important;
        }
        div[data-testid="stTabs"] button[role="tab"]:hover {
            color: #1A1714 !important;
        }
        </style>""", unsafe_allow_html=True)

        # ── Diagnóstico de erro Supabase (visível só quando falha) ──────────
        _sb_err = st.session_state.get("_ultimo_erro_supabase")
        if _sb_err:
            with st.expander(f"🔴 Erro técnico Supabase (HTTP {_sb_err['status']}) — clique para ver", expanded=True):
                st.code(json.dumps(_sb_err["detail"], ensure_ascii=False, indent=2)
                        if isinstance(_sb_err["detail"], dict) else str(_sb_err["detail"]),
                        language="json")
                st.caption("Copie esse erro e envie ao desenvolvedor. Clique no botão abaixo para limpar.")
                if st.button("✕ Limpar diagnóstico", key="limpar_err_sb"):
                    st.session_state.pop("_ultimo_erro_supabase", None)
                    st.rerun()

        # ── Sucesso após cadastro ─────────────────────────────────────
        if st.session_state.novo_ped_ok:
            st.success("✅ Pedido(s) adicionado(s) com sucesso! Os operadores já têm acesso.")
            st.markdown('<div class="btn-novo-ped">', unsafe_allow_html=True)
            if st.button("➕ Adicionar mais pedidos", use_container_width=True, key="novo_ped_reset"):
                for _k in ["novo_ped_num","novo_ped_cli","novo_ped_prod","novo_ped_obs",
                           "novo_ped_ok","novo_ped_erro","novo_ped_xlsx_preview"]:
                    st.session_state[_k] = "" if isinstance(st.session_state[_k], str) else (
                        False if isinstance(st.session_state[_k], bool) else None)
                st.session_state.novo_ped_qtd = 0
                st.session_state.novo_ped_vr  = 0.0
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        else:
            # ── ABAS: XLSX | Manual ───────────────────────────────────
            _modo_add = st.radio("_m", ["📂 Importar XLSX", "✏️ Cadastro Manual"],
                horizontal=True, label_visibility="collapsed", key="modo_add_k")

            # ════════════════════════════════════════════════
            #  MODO 1 — IMPORTAR XLSX
            # ════════════════════════════════════════════════
            if _modo_add == "📂 Importar XLSX":
                st.markdown("""
                <div style="background:#F0F5FF;border:1.5px solid #3B7DD8;border-radius:10px;
                            padding:10px 14px;margin-bottom:12px;font-size:12px;font-weight:700;color:#1e3a8a;">
                  📂 Exporte o pedido do sistema interno como <strong>XLSX</strong> e faça upload aqui.
                  O sistema extrai as informações automaticamente e descarta o arquivo.
                </div>""", unsafe_allow_html=True)

                xlsx_file = st.file_uploader("Upload XLSX", type=["xlsx"],
                    label_visibility="collapsed", key="novo_ped_xlsx_upload")

                if xlsx_file is not None:
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(io.BytesIO(xlsx_file.read()))
                        ws = wb.active
                        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

                        # Mapeamento de colunas pelo cabeçalho (match exato preferido)
                        def _col(names):
                            # Tenta match exato primeiro
                            for n in names:
                                for i, h in enumerate(headers):
                                    if h.strip().lower() == n.lower(): return i
                            # Fallback: contém
                            for n in names:
                                for i, h in enumerate(headers):
                                    if n.lower() in h.lower(): return i
                            return None

                        idx_num  = _col(["Pedido"])
                        # Cliente: pega a primeira coluna cujo valor seja string com nome (não data, não int)
                        idx_cli  = None
                        for _i, _h in enumerate(headers):
                            if _h.strip().lower() == "cliente":
                                idx_cli = _i; break
                        idx_est  = _col(["Est. Alocado"])
                        idx_vr   = _col(["Vr. Alocado"])
                        idx_obs  = _col(["Observação","Observacao"])
                        idx_prod = _col(["Perfil","Produto"])

                        pedidos_xlsx = []
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            num = row[idx_num] if idx_num is not None else None
                            if not num: continue
                            num = str(int(num)) if isinstance(num, float) else str(num).strip()
                            # Extrai nome do cliente — percorre as colunas "Cliente" até achar uma string
                            import re
                            cli = ""
                            for _ci in range(len(headers)):
                                if headers[_ci].strip().lower() == "cliente" and row[_ci]:
                                    _raw = str(row[_ci]).strip()
                                    # Ignora colunas que são só números ou datas
                                    if re.match(r'^\d+$', _raw): continue
                                    if "datetime" in _raw or re.match(r'\d{4}-\d{2}', _raw): continue
                                    # Remove prefixo "CODIGO - "
                                    _match = re.findall(r'\d+ - (.+)', _raw)
                                    cli = _match[-1].strip() if _match else _raw
                                    if cli: break
                            est  = row[idx_est]  if idx_est  is not None else None
                            vr   = row[idx_vr]   if idx_vr   is not None else None
                            obs  = str(row[idx_obs]).strip()  if idx_obs  is not None and row[idx_obs] and str(row[idx_obs]) != "None" else ""
                            prod = str(row[idx_prod]).strip() if idx_prod is not None and row[idx_prod] and str(row[idx_prod]) != "None" else ""
                            try: est = int(float(est)) if est else None
                            except: est = None
                            try: vr = round(float(vr), 2) if vr else None
                            except: vr = None
                            pedidos_xlsx.append({"num": num, "cli": cli, "prod": prod,
                                                 "est": est, "vr": vr, "obs": obs})

                        if not pedidos_xlsx:
                            st.error("❌ Nenhum pedido encontrado no arquivo.")
                        else:
                            st.session_state.novo_ped_xlsx_preview = pedidos_xlsx
                    except Exception as _xe:
                        st.error(f"❌ Erro ao ler arquivo: {_xe}")

                # Preview dos pedidos lidos
                preview = st.session_state.novo_ped_xlsx_preview
                if preview:
                    st.markdown(f"""
                    <div style="font-size:11px;font-weight:800;color:#4A7C59;letter-spacing:1px;
                                text-transform:uppercase;margin:10px 0 6px;">
                      ✓ {len(preview)} pedido(s) lido(s) — confirme antes de importar
                    </div>""", unsafe_allow_html=True)

                    for _p in preview:
                        _vr_fmt = f"R$ {_p['vr']:,.2f}".replace(",","X").replace(".",",").replace("X",".") if _p["vr"] else "—"
                        st.markdown(f"""
                        <div style="background:#F7F5F2;border:1.5px solid #DDD8D2;border-radius:10px;
                                    padding:10px 14px;margin-bottom:6px;font-size:12px;">
                          <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;">
                            <span style="font-family:monospace;font-size:14px;font-weight:900;color:#1A1714;">#{_p['num']}</span>
                            <span style="font-weight:700;color:#5C5450;flex:1;">{_p['cli'] or '—'}</span>
                            <span style="background:#F0F5FF;color:#3B7DD8;font-size:10px;font-weight:800;
                                  padding:2px 10px;border-radius:20px;">{_p['est'] or '—'} pçs</span>
                            <span style="background:#F0F7F3;color:#4A7C59;font-size:10px;font-weight:800;
                                  padding:2px 10px;border-radius:20px;">{_vr_fmt}</span>
                          </div>
                          {f'<div style="font-size:10px;color:#9C9490;margin-top:4px;">📝 {_p["obs"]}</div>' if _p["obs"] else ""}
                          {f'<div style="font-size:10px;color:#9C9490;margin-top:2px;">🏷 {_p["prod"]}</div>' if _p["prod"] else ""}
                        </div>""", unsafe_allow_html=True)

                    if st.session_state.novo_ped_erro:
                        st.error(st.session_state.novo_ped_erro)

                    st.markdown('<div class="btn-novo-ped">', unsafe_allow_html=True)
                    if st.button("💾  Importar Pedido(s)", use_container_width=True, key="novo_ped_xlsx_salvar"):
                        _erros = []; _falhas = []; _ok = 0
                        for _p in preview:
                            _existe = _get("pedidos_base", f"numero=eq.{_p['num']}&select=numero")
                            if isinstance(_existe, list) and _existe:
                                _erros.append(f"#{_p['num']} já existe — ignorado")
                                continue
                            _inseriu = cadastrar_pedido_avulso(
                                numero=_p["num"], cliente=_p["cli"],
                                produto=_p["prod"], est_alocado=_p["est"], vr_alocado=_p["vr"]
                            )
                            if _inseriu:
                                _ok += 1
                            else:
                                _falhas.append(f"#{_p['num']} falhou ao salvar")
                        buscar_pedidos_base.clear()
                        buscar_pedidos_por_etapa.clear()
                        st.session_state.novo_ped_xlsx_preview = None
                        msgs = []
                        if _erros:  msgs.append("⚠️ Já existiam: " + ", ".join(_erros))
                        if _falhas: msgs.append("❌ Erro ao salvar: " + ", ".join(_falhas) + " — verifique a conexão com o banco.")
                        st.session_state.novo_ped_erro = "\n".join(msgs) if msgs else ""
                        if _ok > 0:
                            st.session_state.novo_ped_ok = True
                        elif not _falhas and not _erros:
                            st.session_state.novo_ped_erro = "❌ Nenhum pedido foi importado."
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)

            # ════════════════════════════════════════════════
            #  MODO 2 — CADASTRO MANUAL
            # ════════════════════════════════════════════════
            elif _modo_add == "✏️ Cadastro Manual":
                st.markdown("""
                <div style="background:#F0F5FF;border:1.5px solid #3B7DD8;border-radius:10px;
                            padding:10px 14px;margin-bottom:12px;font-size:12px;font-weight:700;color:#1e3a8a;">
                  ✏️ Preencha manualmente os dados do pedido.
                </div>""", unsafe_allow_html=True)

                c1, c2 = st.columns([1, 2])
                with c1:
                    st.markdown('<div style="font-size:11px;font-weight:800;color:#5C5450;margin-bottom:4px;letter-spacing:.5px;">Nº DO PEDIDO *</div>', unsafe_allow_html=True)
                    novo_num = st.text_input("_np_num", placeholder="Ex: 50999",
                        label_visibility="collapsed", key="novo_ped_num_input")
                with c2:
                    st.markdown('<div style="font-size:11px;font-weight:800;color:#5C5450;margin-bottom:4px;letter-spacing:.5px;">CLIENTE</div>', unsafe_allow_html=True)
                    novo_cli = st.text_input("_np_cli", placeholder="Nome do cliente",
                        label_visibility="collapsed", key="novo_ped_cli_input")

                st.markdown('<div style="font-size:11px;font-weight:800;color:#5C5450;margin-bottom:4px;letter-spacing:.5px;margin-top:8px;">PRODUTO / DESCRIÇÃO</div>', unsafe_allow_html=True)
                novo_prod = st.text_input("_np_prod", placeholder="Ex: Conjunto Renda Preta P/M/G",
                    label_visibility="collapsed", key="novo_ped_prod_input")

                c3, c4 = st.columns(2)
                with c3:
                    st.markdown('<div style="font-size:11px;font-weight:800;color:#5C5450;margin-bottom:4px;letter-spacing:.5px;margin-top:8px;">QTD DE PEÇAS</div>', unsafe_allow_html=True)
                    novo_qtd = st.number_input("_np_qtd", min_value=0, value=0, step=1,
                        label_visibility="collapsed", key="novo_ped_qtd_input")
                with c4:
                    st.markdown('<div style="font-size:11px;font-weight:800;color:#5C5450;margin-bottom:4px;letter-spacing:.5px;margin-top:8px;">VALOR (R$)</div>', unsafe_allow_html=True)
                    novo_vr = st.number_input("_np_vr", min_value=0.0, value=0.0, step=0.01,
                        format="%.2f", label_visibility="collapsed", key="novo_ped_vr_input")

                if st.session_state.novo_ped_erro:
                    st.error(st.session_state.novo_ped_erro)

                st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
                st.markdown('<div class="btn-novo-ped">', unsafe_allow_html=True)
                if st.button("💾  Salvar Pedido", use_container_width=True, key="novo_ped_salvar"):
                    _num = novo_num.strip()
                    if not _num:
                        st.session_state.novo_ped_erro = "❌ O número do pedido é obrigatório."
                        st.rerun()
                    elif not _num.isdigit():
                        st.session_state.novo_ped_erro = "❌ O número deve conter apenas dígitos."
                        st.rerun()
                    else:
                        _existe = _get("pedidos_base", f"numero=eq.{_num}&select=numero")
                        if isinstance(_existe, list) and _existe:
                            st.session_state.novo_ped_erro = f"❌ Pedido #{_num} já existe na base."
                            st.rerun()
                        else:
                            _inseriu = cadastrar_pedido_avulso(
                                numero=_num, cliente=novo_cli.strip(),
                                produto=novo_prod.strip(),
                                est_alocado=int(novo_qtd) if novo_qtd > 0 else None,
                                vr_alocado=float(novo_vr) if novo_vr > 0 else None,
                            )
                            buscar_pedidos_base.clear()
                            if _inseriu:
                                st.session_state.novo_ped_ok   = True
                                st.session_state.novo_ped_erro = ""
                            else:
                                st.session_state.novo_ped_erro = (
                                    f"❌ Falha ao salvar pedido #{_num} no banco. "
                                    "Verifique a conexão com o Supabase ou os logs da aplicação."
                                )
                            buscar_pedidos_por_etapa.clear()
                            st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

    avulsos = buscar_pedidos_avulsos()

    with st.expander(
        f"📋 Pedidos Adicionados Manualmente  {'— ' + str(len(avulsos)) + ' encontrado(s)' if avulsos else '— nenhum cadastrado'}",
        expanded=bool(avulsos)
    ):
        if not avulsos:
            st.markdown("""
            <div style="background:#F0F7F3;border:1.5px solid #4A7C59;border-radius:12px;
                        padding:18px 20px;text-align:center;">
                <div style="font-size:22px;margin-bottom:6px;">✅</div>
                <div style="font-size:13px;font-weight:700;color:#2d5a3d;">
                    Nenhum pedido adicionado manualmente.</div>
                <div style="font-size:11px;color:#4A7C59;margin-top:4px;font-weight:600;">
                    Todos os pedidos vieram da planilha do Programa A.</div>
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown("""
            <div style="background:#F0F5FF;border:1.5px solid #3B7DD8;border-radius:10px;
                        padding:12px 16px;font-size:12px;font-weight:700;color:#1e3a8a;margin-bottom:12px;">
                📋 Pedidos adicionados manualmente pelo gestor. O pedido permanece visível
                enquanto estiver <strong>em aberto ou em produção</strong>. Some quando for
                excluído ou quando a planilha reimportada já incluir o pedido.
            </div>
            """, unsafe_allow_html=True)

            if "confirm_excluir" not in st.session_state:
                st.session_state.confirm_excluir = {}

            for numero, cliente, status, importado_em in avulsos:
                regs_vinculados = _get("registros", f"pedido=eq.{numero}&select=id")
                tem_registros   = isinstance(regs_vinculados, list) and len(regs_vinculados) > 0
                cor_status      = "#4A7C59" if status == "aberto" else "#C8566A"
                lbl_status      = "aberto" if status == "aberto" else "concluído"

                col_info, col_btn = st.columns([4, 1])
                with col_info:
                    aviso_reg = (
                        f' &nbsp;·&nbsp; <span style="color:#C47B2A;font-size:10px;">'
                        f'⚠ {len(regs_vinculados)} registro(s) de produção serão removidos</span>'
                        if tem_registros else ""
                    )
                    cli_show = f" · {cliente}" if cliente else ""
                    st.markdown(f"""
                    <div style="background:#fff;border:1.5px solid #EDE9E4;border-radius:10px;
                                padding:11px 16px;display:flex;align-items:center;gap:10px;flex-wrap:wrap;">
                        <span style="font-family:monospace;font-size:14px;font-weight:800;
                                     color:#1A1714;">{numero}</span>
                        <span style="background:{cor_status}22;color:{cor_status};font-size:10px;
                                     font-weight:800;padding:2px 10px;border-radius:20px;
                                     text-transform:uppercase;">{lbl_status}</span>
                        {f'<span style="font-size:12px;font-weight:700;color:#5C5450;">{cliente}</span>' if cliente else ""}
                        <span style="font-size:11px;color:#9C9490;">
                            Cadastrado em: {importado_em or "—"}</span>
                        {aviso_reg}
                    </div>
                    """, unsafe_allow_html=True)

                with col_btn:
                    em_confirmacao = st.session_state.confirm_excluir.get(numero, False)

                    if not em_confirmacao:
                        st.markdown("""
                        <style>
                        .btn-del > button {
                            background:#FEF2F2 !important; color:#C8566A !important;
                            border:1.5px solid #FECACA !important; border-radius:10px !important;
                            font-size:12px !important; font-weight:800 !important;
                            height:46px !important;
                        }
                        .btn-del > button:hover {
                            background:#C8566A !important; color:#fff !important;
                            border-color:#C8566A !important;
                        }
                        </style>""", unsafe_allow_html=True)
                        st.markdown('<div class="btn-del">', unsafe_allow_html=True)
                        if st.button("🗑 Excluir", key=f"del_{numero}", use_container_width=True):
                            st.session_state.confirm_excluir[numero] = True
                            st.rerun()
                        st.markdown('</div>', unsafe_allow_html=True)
                    else:
                        st.markdown("""
                        <div style="font-size:10px;font-weight:800;color:#C8566A;
                                    text-align:center;margin-bottom:4px;">Confirmar?</div>
                        """, unsafe_allow_html=True)
                        st.markdown("""
                        <style>
                        .btn-sim > button {
                            background:#C8566A !important; color:#fff !important;
                            border:none !important; border-radius:8px !important;
                            font-size:11px !important; font-weight:800 !important;
                            height:38px !important;
                        }
                        .btn-voltar > button { height:38px !important; }
                        </style>""", unsafe_allow_html=True)
                        ca, cb = st.columns(2)
                        with ca:
                            st.markdown('<div class="btn-sim">', unsafe_allow_html=True)
                            if st.button("✓ Sim", key=f"sim_{numero}", use_container_width=True):
                                excluir_pedido_avulso(numero)
                                st.session_state.confirm_excluir.pop(numero, None)
                                st.toast(f"✅ Pedido {numero} excluído com sucesso.", icon="🗑️")
                                st.rerun()
                            st.markdown('</div>', unsafe_allow_html=True)
                        with cb:
                            st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
                            if st.button("✕", key=f"nao_{numero}", use_container_width=True):
                                st.session_state.confirm_excluir[numero] = False
                                st.rerun()
                            st.markdown('</div>', unsafe_allow_html=True)

                st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

    st.markdown("<br style='line-height:0.3'>", unsafe_allow_html=True)

    regs     = buscar()
    ped_comp = list({r[1] for r in regs if r[4] == 2})
    ops_ativ = list({r[2] for r in regs})
    avg      = media([r[5] for r in regs]) // 60 if regs else 0
    total_pecas_geral = sum(r[8] for r in regs if r[8] is not None)

    kpi_html = f"""
    <!DOCTYPE html><html><head>
    <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@700;800;900&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
    <style>
    *{{margin:0;padding:0;box-sizing:border-box;}}
    body{{background:transparent;font-family:Nunito,sans-serif;}}
    .grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;}}
    .card{{background:#fff;border-radius:16px;padding:18px 16px 14px;
           border:1.5px solid #EDE9E4;box-shadow:0 2px 12px rgba(0,0,0,0.05);position:relative;overflow:hidden;}}
    .card-icon{{position:absolute;top:-6px;right:4px;font-size:44px;opacity:0.07;line-height:1;}}
    .card-lbl{{font-size:9px;font-weight:800;letter-spacing:1.8px;text-transform:uppercase;color:#9C9490;margin-bottom:8px;}}
    .card-num{{font-family:"DM Mono",monospace;font-size:32px;font-weight:500;letter-spacing:-1px;}}
    .card-bar{{height:3px;border-radius:2px;margin-top:12px;opacity:0.3;}}
    </style></head><body>
    <div class="grid">
      <div class="card"><div class="card-icon">📦</div><div class="card-lbl">Pedidos Concluídos</div><div class="card-num" style="color:#C8566A;">{len(ped_comp)}</div><div class="card-bar" style="background:#C8566A;"></div></div>
      <div class="card"><div class="card-icon">👥</div><div class="card-lbl">Operadores Ativos</div><div class="card-num" style="color:#4A7C59;">{len(ops_ativ)}</div><div class="card-bar" style="background:#4A7C59;"></div></div>
      <div class="card"><div class="card-icon">⏱</div><div class="card-lbl">Tempo Médio</div><div class="card-num" style="color:#3B5EC6;">{avg}m</div><div class="card-bar" style="background:#3B5EC6;"></div></div>
      <div class="card"><div class="card-icon">👕</div><div class="card-lbl">Peças Separadas</div><div class="card-num" style="color:#C47B2A;">{total_pecas_geral}</div><div class="card-bar" style="background:#C47B2A;"></div></div>
    </div>
    </body></html>"""
    st.markdown(kpi_html, unsafe_allow_html=True)

    op_map = {}
    for r in regs:
        op = r[2]
        if op not in op_map:
            op_map[op] = {"p": set(), "pecas": 0, "tempo_total": 0, "tempos": []}
        op_map[op]["p"].add(r[1])
        op_map[op]["tempo_total"] += int(r[5] or 0)
        op_map[op]["tempos"].append(int(r[5] or 0))
        op_map[op]["pecas"] += int(r[8] or 0)

    todas_datas_regs = sorted({
        r[6].split(" ")[0] for r in regs if r[6] and " " in str(r[6])
    }, reverse=True) if regs else []
    todos_ops_regs = sorted({r[2] for r in regs if r[2]}) if regs else []

    # Padrão: dia mais recente com dados (não "Todos os dias")
    _dia_padrao = todas_datas_regs[0] if todas_datas_regs else "Todos os dias"
    opcoes_data = ["Todos os dias"] + todas_datas_regs
    _idx_padrao = opcoes_data.index(_dia_padrao) if _dia_padrao in opcoes_data else 0

    fc1, fc2 = st.columns(2)
    with fc1:
        filtro_data = st.selectbox("📅 Filtrar por dia", opcoes_data,
            index=_idx_padrao, key="admin_filtro_data")
    with fc2:
        opcoes_op = ["Todos os operadores"] + todos_ops_regs
        filtro_op = st.selectbox("👤 Filtrar por operador", opcoes_op, key="admin_filtro_op")

    regs_filtrados = regs
    if filtro_data != "Todos os dias":
        regs_filtrados = [r for r in regs_filtrados if str(r[6]).startswith(filtro_data)]
    if filtro_op != "Todos os operadores":
        regs_filtrados = [r for r in regs_filtrados if r[2] == filtro_op]

    tem_filtro = filtro_data != "Todos os dias" or filtro_op != "Todos os operadores"
    if tem_filtro:
        partes_filtro = []
        if filtro_data != "Todos os dias": partes_filtro.append(f"📅 {filtro_data}")
        if filtro_op   != "Todos os operadores": partes_filtro.append(f"👤 {filtro_op}")
        ped_filt = len({r[1] for r in regs_filtrados})
        st.markdown(
            f'<div style="background:#F0F7F3;border:1.5px solid #4A7C59;border-radius:10px;'
            f'padding:10px 16px;font-size:13px;font-weight:700;color:#2d5a3d;margin-bottom:4px;">'
            f'Filtro ativo: {" · ".join(partes_filtro)} &nbsp;→&nbsp; '
            f'<strong>{ped_filt} pedido(s)</strong> · {len(regs_filtrados)} registro(s)</div>',
            unsafe_allow_html=True
        )

    regs_para_tabela = regs_filtrados

    op_map = {}
    for r in regs_filtrados:
        op = r[2]
        if op not in op_map:
            op_map[op] = {"p": set(), "pecas": 0, "tempo_total": 0, "tempos": []}
        op_map[op]["p"].add(r[1])
        op_map[op]["tempo_total"] += int(r[5] or 0)
        op_map[op]["tempos"].append(int(r[5] or 0))
        op_map[op]["pecas"] += int(r[8] or 0)

    # Cruzar pausas_log filtradas com op_map para contar pausas por operador
    pausas_log_full = buscar_pausas_log()
    pausas_filtradas_op = pausas_log_full
    if filtro_data != "Todos os dias":
        pausas_filtradas_op = [p for p in pausas_filtradas_op if str(p[4] or "").startswith(filtro_data)]
    if filtro_op != "Todos os operadores":
        pausas_filtradas_op = [p for p in pausas_filtradas_op if p[2] == filtro_op]
    pausas_por_op = {}
    for p in pausas_filtradas_op:
        pausas_por_op[p[2]] = pausas_por_op.get(p[2], 0) + 1

    if op_map:
        # ── Ordenar por peças desc ─────────────────────────────────────────
        op_sorted = sorted(op_map.items(), key=lambda x: x[1]["pecas"], reverse=True)
        op_rows = ""
        medals = ["🥇", "🥈", "🥉"]
        for rank, (op, d) in enumerate(op_sorted):
            n_pedidos   = len(d["p"])
            total_pecas = d["pecas"]
            tempo_total = d["tempo_total"]
            tempo_medio = media(d["tempos"])
            horas_trab  = tempo_total / 3600 if tempo_total > 0 else 0
            eficiencia  = f"{round(total_pecas / horas_trab, 1)} pçs/h" if horas_trab > 0 and total_pecas > 0 else "—"
            prod_ped    = f"{round(total_pecas / n_pedidos, 1)} pçs/ped" if n_pedidos > 0 and total_pecas > 0 else "—"
            n_pausas_op = pausas_por_op.get(op, 0)
            pausas_badge = (
                f'<span style="background:#FFF0E6;color:#E07B3A;font-weight:800;'
                f'font-size:12px;padding:2px 10px;border-radius:100px;">{n_pausas_op}</span>'
                if n_pausas_op > 0
                else '<span style="color:#C0BAB4;font-weight:700;font-size:12px;">0</span>'
            )
            ini         = op[0].upper()
            medal       = medals[rank] if rank < 3 else ""
            op_rows += f"""<tr>
              <td style="padding:12px 16px;vertical-align:middle;">
                <div style="display:flex;align-items:center;gap:10px;">
                  <div style="width:34px;height:34px;border-radius:50%;flex-shrink:0;
                       background:linear-gradient(135deg,#D9617A,#9E3F52);
                       display:flex;align-items:center;justify-content:center;
                       font-size:14px;font-weight:900;color:#fff;">{ini}</div>
                  <span style="font-weight:800;font-size:13px;color:#1A1714;">{op}</span>
                  {f'<span style="font-size:14px;margin-left:2px;">{medal}</span>' if medal else ""}
                </div>
              </td>
              <td style="padding:12px 8px;text-align:center;vertical-align:middle;">
                <span style="background:#F5E8EB;color:#C8566A;font-weight:800;font-size:13px;padding:3px 12px;border-radius:100px;">{n_pedidos}</span>
              </td>
              <td style="padding:12px 8px;text-align:center;font-family:monospace;font-size:13px;color:#3B5EC6;font-weight:800;vertical-align:middle;">{total_pecas}</td>
              <td style="padding:12px 8px;text-align:center;font-family:monospace;font-size:12px;color:#4A7C59;font-weight:700;vertical-align:middle;">{fmt(tempo_total)}</td>
              <td style="padding:12px 8px;text-align:center;font-family:monospace;font-size:12px;color:#5C5450;font-weight:700;vertical-align:middle;">{fmt(tempo_medio)}</td>
              <td style="padding:12px 8px;text-align:center;vertical-align:middle;">{pausas_badge}</td>
              <td style="padding:12px 8px;text-align:center;font-size:11px;color:#C47B2A;font-weight:800;vertical-align:middle;">{eficiencia}</td>
              <td style="padding:12px 8px;text-align:center;font-size:11px;color:#7C3AED;font-weight:700;vertical-align:middle;">{prod_ped}</td>
            </tr>"""

        n_ops = len(op_map)
        op_table_height = 24 + 52 + n_ops * 52 + 16
        lbl_op = f"Desempenho por Operador · {filtro_data}" if filtro_data != "Todos os dias" else "Desempenho por Operador"
        components.html(f"""<!DOCTYPE html><html><head>
        <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@700;800;900&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
        <style>
        *{{margin:0;padding:0;box-sizing:border-box;}} body{{background:transparent;font-family:Nunito,sans-serif;}}
        .lbl{{font-size:9px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:#9C9490;margin-bottom:10px;}}
        .wrap{{background:#fff;border-radius:16px;border:1.5px solid #EDE9E4;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.05);}}
        table{{width:100%;border-collapse:collapse;}} thead tr{{background:#1A1714;}}
        th{{padding:11px 8px;font-size:8px;font-weight:800;letter-spacing:1.3px;text-transform:uppercase;text-align:center;}}
        th:first-child{{text-align:left;padding-left:16px;}}
        tbody tr{{border-bottom:1px solid #F2EEE9;transition:background .15s;}}
        tbody tr:last-child{{border-bottom:none;}} tbody tr:hover{{background:#FDFAF9;}}
        </style></head><body>
        <div class="lbl">{lbl_op}</div>
        <div class="wrap"><table><thead><tr>
          <th style="color:rgba(255,255,255,0.45);">Operador</th>
          <th style="color:rgba(255,255,255,0.45);">Pedidos</th>
          <th style="color:#7B9FE0;">Peças</th>
          <th style="color:#7AB895;">Tempo Total</th>
          <th style="color:rgba(255,255,255,0.35);">Tempo Médio</th>
          <th style="color:#E07B3A;">Pausas</th>
          <th style="color:#E5A96A;">Eficiência</th>
          <th style="color:#C4A4F0;">Produtividade</th>
        </tr></thead><tbody>{op_rows}</tbody></table></div>
        </body></html>""", height=op_table_height, scrolling=False)

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        # ── Ranking do Dia ─────────────────────────────────────────────────
        ranking_html = ""
        max_pcs_rank = op_sorted[0][1]["pecas"] if op_sorted else 1
        for rank, (op, d) in enumerate(op_sorted[:5]):
            pcs     = d["pecas"]
            bar_pct = int(pcs / max_pcs_rank * 100) if max_pcs_rank > 0 else 0
            medal_bg  = ["#FFD700","#C0C0C0","#CD7F32"]
            m_color   = medal_bg[rank] if rank < 3 else "#EDE9E4"
            m_txt     = ["🥇","🥈","🥉"][rank] if rank < 3 else f"#{rank+1}"
            ranking_html += f"""
            <div style="display:flex;align-items:center;gap:12px;margin-bottom:10px;">
              <div style="width:32px;height:32px;border-radius:50%;background:{m_color};
                   display:flex;align-items:center;justify-content:center;
                   font-size:16px;flex-shrink:0;">{m_txt}</div>
              <div style="flex:1;">
                <div style="display:flex;justify-content:space-between;margin-bottom:4px;">
                  <span style="font-size:13px;font-weight:800;color:#1A1714;">{op}</span>
                  <span style="font-family:monospace;font-size:13px;font-weight:700;color:#3B5EC6;">{pcs} peças</span>
                </div>
                <div style="background:#F2EEE9;border-radius:4px;height:6px;overflow:hidden;">
                  <div style="width:{bar_pct}%;height:100%;background:linear-gradient(90deg,#C8566A,#9E3F52);border-radius:4px;"></div>
                </div>
              </div>
            </div>"""

        ranking_height = 24 + min(len(op_sorted), 5) * 58 + 32
        components.html(f"""<!DOCTYPE html><html><head>
        <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@700;800;900&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
        <style>*{{margin:0;padding:0;box-sizing:border-box;}} body{{background:transparent;font-family:Nunito,sans-serif;}}
        .lbl{{font-size:9px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:#9C9490;margin-bottom:12px;}}
        .card{{background:#fff;border-radius:16px;border:1.5px solid #EDE9E4;padding:18px 20px;
               box-shadow:0 2px 12px rgba(0,0,0,0.05);}}
        </style></head><body>
        <div class="lbl">🏆 Ranking do Dia — Peças Produzidas</div>
        <div class="card">{ranking_html}</div>
        </body></html>""", height=ranking_height, scrolling=False)

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        # ── Produção por Hora ──────────────────────────────────────────────
        from collections import defaultdict
        pecas_por_hora = defaultdict(int)
        for r in regs_filtrados:
            hora_str = str(r[6] or "")
            if " " in hora_str:
                try:
                    hora = int(hora_str.split(" ")[1].split(":")[0])
                    pecas_por_hora[hora] += int(r[8] or 0)
                except Exception:
                    pass

        if pecas_por_hora:
            horas_existentes = sorted(pecas_por_hora.keys())
            h_min = max(min(horas_existentes) - 1, 0)
            h_max = min(max(horas_existentes) + 1, 23)
            horas_range = list(range(h_min, h_max + 1))
            max_pcs_hora = max(pecas_por_hora.values()) if pecas_por_hora else 1
            bars_html = ""
            for h in horas_range:
                pcs_h   = pecas_por_hora.get(h, 0)
                bar_h   = int(pcs_h / max_pcs_hora * 110) if max_pcs_hora > 0 else 0
                cor_bar = "#C8566A" if pcs_h == max_pcs_hora else "#3B5EC6"
                bars_html += f"""
                <div style="display:flex;flex-direction:column;align-items:center;gap:4px;flex:1;min-width:28px;">
                  <div style="font-family:monospace;font-size:10px;font-weight:700;color:#3B5EC6;min-height:16px;">{pcs_h if pcs_h > 0 else ""}</div>
                  <div style="width:100%;max-width:36px;height:{bar_h}px;background:{cor_bar};border-radius:4px 4px 0 0;"></div>
                  <div style="font-size:9px;font-weight:800;color:#9C9490;">{h:02d}h</div>
                </div>"""
            components.html(f"""<!DOCTYPE html><html><head>
            <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@700;800;900&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
            <style>*{{margin:0;padding:0;box-sizing:border-box;}} body{{background:transparent;font-family:Nunito,sans-serif;}}
            .lbl{{font-size:9px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:#9C9490;margin-bottom:12px;}}
            .card{{background:#fff;border-radius:16px;border:1.5px solid #EDE9E4;padding:16px 20px 12px;
                   box-shadow:0 2px 12px rgba(0,0,0,0.05);}}
            .bars{{display:flex;align-items:flex-end;gap:6px;height:130px;padding-top:20px;}}
            </style></head><body>
            <div class="lbl">📈 Produção por Hora</div>
            <div class="card"><div class="bars">{bars_html}</div></div>
            </body></html>""", height=180, scrolling=False)

    else:
        st.markdown("""<div style="background:#fff;border-radius:16px;border:1.5px solid #EDE9E4;
                    padding:40px;text-align:center;color:#9C9490;font-size:14px;font-weight:600;">
            Nenhum registro encontrado para o filtro selecionado.</div>""", unsafe_allow_html=True)

    st.markdown("<br style='line-height:0.4'>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════
    #  BLOCO — HISTÓRICO DE PAUSAS
    # ══════════════════════════════════════════════════════════════════
    pausas_log = buscar_pausas_log()

    # Filtra pelo mesmo dia e operador dos outros filtros
    pausas_filtradas = pausas_log
    if filtro_data != "Todos os dias":
        pausas_filtradas = [p for p in pausas_filtradas if str(p[4] or "").startswith(filtro_data)]
    if filtro_op != "Todos os operadores":
        pausas_filtradas = [p for p in pausas_filtradas if p[2] == filtro_op]

    n_pausas = len(pausas_filtradas)
    with st.expander(
        f"⏸ Histórico de Pausas — {n_pausas} pausa(s)" if n_pausas > 0
        else "⏸ Histórico de Pausas — nenhuma registrada",
        expanded=n_pausas > 0
    ):
        if not pausas_filtradas:
            components.html("""<!DOCTYPE html><html><body style="background:transparent;
            font-family:Nunito,sans-serif;margin:0;padding:0;">
            <div style="background:#FFF8F0;border:1.5px solid #E07B3A33;border-radius:12px;
                        padding:20px;text-align:center;">
              <div style="font-size:22px;margin-bottom:6px;">⏸</div>
              <div style="font-size:13px;font-weight:700;color:#B85C20;">
                Nenhuma pausa registrada para este filtro.</div>
              <div style="font-size:11px;color:#9C9490;margin-top:4px;font-weight:600;">
                As pausas aparecem aqui quando confirmadas com senha do gestor.</div>
            </div></body></html>""", height=110, scrolling=False)
        else:
            ETAPA_TAG = {
                0: '<span style="background:#EBF0FB;color:#3B5EC6;padding:2px 8px;border-radius:100px;font-size:10px;font-weight:800;">Separação</span>',
                1: '<span style="background:#E8F2EC;color:#4A7C59;padding:2px 8px;border-radius:100px;font-size:10px;font-weight:800;">Embalagem</span>',
                2: '<span style="background:#FBF2E6;color:#C47B2A;padding:2px 8px;border-radius:100px;font-size:10px;font-weight:800;">Conferência</span>',
            }
            pausa_rows = ""
            for p in pausas_filtradas[:100]:
                # p: (id, pedido, operador, etapa_idx, pausado_em, tempo_pausado_s, motivo)
                etapa_tag  = ETAPA_TAG.get(p[3], str(p[3]))
                pausado_em = p[4] if p[4] else "—"
                tempo_p    = fmt(p[5]) if p[5] else "—"
                motivo_p   = p[6] if p[6] else '<span style="color:#C0BAB4;font-style:italic;">—</span>'
                pausa_rows += f"""<tr>
                  <td style="padding:10px 14px;font-family:monospace;font-size:12px;font-weight:700;color:#1A1714;">{p[1]}</td>
                  <td style="padding:10px 8px;font-size:13px;font-weight:700;color:#1A1714;">{p[2]}</td>
                  <td style="padding:10px 8px;">{etapa_tag}</td>
                  <td style="padding:10px 8px;font-size:11px;color:#9C9490;text-align:center;">{pausado_em}</td>
                  <td style="padding:10px 8px;font-family:monospace;font-size:12px;font-weight:700;color:#E07B3A;text-align:center;">{tempo_p}</td>
                  <td style="padding:10px 8px;font-size:12px;color:#5C5450;font-weight:600;">{motivo_p}</td>
                </tr>"""

            n_p = min(len(pausas_filtradas), 100)
            pausa_height = 52 + n_p * 44 + 20
            components.html(f"""<!DOCTYPE html><html><head>
            <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@700;800;900&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
            <style>
            *{{margin:0;padding:0;box-sizing:border-box;}} body{{background:transparent;font-family:Nunito,sans-serif;}}
            .wrap{{background:#fff;border-radius:16px;border:1.5px solid #EDE9E4;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.05);}}
            table{{width:100%;border-collapse:collapse;}} thead tr{{background:#1A1714;}}
            th{{padding:11px 8px;font-size:8px;font-weight:800;letter-spacing:1.3px;text-transform:uppercase;color:rgba(255,255,255,0.45);text-align:center;}}
            th:first-child{{text-align:left;padding-left:14px;}} th:nth-child(2){{text-align:left;}} th:last-child{{text-align:left;}}
            tbody tr{{border-bottom:1px solid #F2EEE9;}} tbody tr:last-child{{border-bottom:none;}}
            tbody tr:hover{{background:#FFF8F0;}}
            </style></head><body>
            <div class="wrap"><table><thead><tr>
              <th>Pedido</th><th>Operador</th><th>Etapa</th>
              <th style="color:#D4A45A;">Pausado em</th>
              <th style="color:#E07B3A;">Tempo</th>
              <th style="color:rgba(255,255,255,0.35);">Motivo</th>
            </tr></thead><tbody>{pausa_rows}</tbody></table></div>
            </body></html>""", height=pausa_height, scrolling=False)

            # ── Resumo por operador ────────────────────────────────────────
            st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
            pausa_por_op = {}
            for p in pausas_filtradas:
                op_p = p[2]
                if op_p not in pausa_por_op:
                    pausa_por_op[op_p] = {"n": 0, "tempo": 0}
                pausa_por_op[op_p]["n"]     += 1
                pausa_por_op[op_p]["tempo"] += int(p[5] or 0)

            resumo_rows = ""
            for op_p, d_p in sorted(pausa_por_op.items(), key=lambda x: x[1]["n"], reverse=True):
                resumo_rows += f"""
                <div style="display:flex;justify-content:space-between;align-items:center;
                     padding:8px 16px;border-bottom:1px solid #F2EEE9;">
                  <span style="font-size:13px;font-weight:800;color:#1A1714;">{op_p}</span>
                  <div style="display:flex;gap:16px;align-items:center;">
                    <span style="background:#FFF0E6;color:#E07B3A;font-size:12px;font-weight:800;
                          padding:2px 10px;border-radius:100px;">{d_p['n']} pausa(s)</span>
                    <span style="font-family:monospace;font-size:12px;font-weight:700;color:#9C9490;">
                      {fmt(d_p['tempo'])} parado</span>
                  </div>
                </div>"""

            resumo_h = len(pausa_por_op) * 40 + 24
            components.html(f"""<!DOCTYPE html><html><head>
            <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@700;800;900&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
            <style>*{{margin:0;padding:0;box-sizing:border-box;}} body{{background:transparent;font-family:Nunito,sans-serif;}}
            .lbl{{font-size:9px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:#9C9490;margin-bottom:8px;}}
            .card{{background:#fff;border-radius:14px;border:1.5px solid #EDE9E4;overflow:hidden;
                   box-shadow:0 2px 12px rgba(0,0,0,0.05);}}
            </style></head><body>
            <div class="lbl">Resumo de Pausas por Operador</div>
            <div class="card">{resumo_rows}</div>
            </body></html>""", height=resumo_h, scrolling=False)

    st.markdown("<br style='line-height:0.4'>", unsafe_allow_html=True)

    # ── Linha de botões de ação: 3 colunas iguais ─────────────────────────
    st.markdown("""
    <style>
    .btn-warn > button {
        background:#FEF3C7 !important; color:#92400E !important;
        border:1.5px solid #F59E0B !important; border-radius:10px !important;
        font-size:13px !important; font-weight:800 !important; height:48px !important;
    }
    .btn-warn > button:hover { background:#F59E0B !important; color:#fff !important; }
    .btn-danger > button {
        background:#FEF2F2 !important; color:#C8566A !important;
        border:1.5px solid #FECACA !important; border-radius:10px !important;
        font-size:13px !important; font-weight:800 !important; height:48px !important;
    }
    .btn-danger > button:hover { background:#C8566A !important; color:#fff !important; }
    .btn-reset-dia > button {
        background: linear-gradient(135deg,#7C3AED,#5B21B6) !important;
        color: #fff !important; border: none !important;
        border-radius: 10px !important; height: 48px !important;
        font-size: 13px !important; font-weight: 800 !important;
        box-shadow: 0 4px 0 rgba(60,10,120,0.35) !important;
    }
    .btn-reset-dia > button:hover { transform: translateY(-1px) !important; }
    </style>""", unsafe_allow_html=True)

    ba, bb, bc = st.columns(3)
    with ba:
        st.markdown('<div class="btn-warn">', unsafe_allow_html=True)
        if st.button("⊘  Limpar PIPs", use_container_width=True, help="Remove sessões ativas fantasmas"):
            limpar_sessoes_ativas(); st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    with bb:
        st.markdown('<div class="btn-danger">', unsafe_allow_html=True)
        if st.button("🗑  Limpar dados", use_container_width=True):
            limpar(); st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    with bc:
        st.markdown('<div class="btn-reset-dia">', unsafe_allow_html=True)
        if st.button("🧹  Apagar hoje", use_container_width=True, key="btn_limpar_dia_inline"):
            st.session_state.confirm_limpar_dia = True
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Botão: Apagar tudo do dia de hoje ──────────────────────────────────
    hoje_str = now_br().strftime("%d/%m/%Y")

    def limpar_dia(data_str):
        rows = _get("registros", f"select=id&data=like.{data_str}%25")
        if isinstance(rows, list):
            for r in rows:
                _delete("registros", f"id=eq.{r['id']}")
        limpar_sessoes_ativas()
        todos_regs = _get("registros", "select=pedido")
        pedidos_com_reg = {r["pedido"] for r in todos_regs} if isinstance(todos_regs, list) else set()
        pedidos_base_rows = _get("pedidos_base", "select=numero")
        if isinstance(pedidos_base_rows, list):
            for p in pedidos_base_rows:
                if p["numero"] not in pedidos_com_reg:
                    _patch("pedidos_base", f"numero=eq.{p['numero']}", {"status": "aberto"})

    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    if "confirm_limpar_dia" not in st.session_state:
        st.session_state.confirm_limpar_dia = False

    if st.session_state.confirm_limpar_dia:
        import streamlit.components.v1 as _cv1t
        _cv1t.html(f"""<!DOCTYPE html><html><head>
        <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@700;800;900&display=swap" rel="stylesheet">
        </head><body style="background:transparent;font-family:Nunito,sans-serif;margin:0;padding:0;">
        <div style="background:#FEF2F2;border:2px solid #FCA5A5;border-radius:14px;
                    padding:16px 20px;text-align:center;">
          <div style="font-size:22px;margin-bottom:6px;">⚠️</div>
          <div style="font-size:14px;font-weight:800;color:#991B1B;margin-bottom:4px;">
            Apagar todos os registros de {hoje_str}?</div>
          <div style="font-size:12px;color:#B91C1C;font-weight:600;">
            Esta ação remove todos os registros e sessões do dia de hoje.<br>
            Ideal para reiniciar os testes. <strong>Não pode ser desfeita.</strong>
          </div>
        </div></body></html>""", height=130, scrolling=False)

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        st.markdown("""
        <style>
        .btn-confirm-del > button {
            background: linear-gradient(135deg,#DC2626,#991B1B) !important;
            color:#fff !important; border:none !important;
            border-radius:10px !important; height:48px !important;
            font-size:13px !important; font-weight:800 !important;
            box-shadow: 0 4px 0 rgba(100,10,10,0.40) !important;
        }
        .btn-confirm-del > button:hover { transform:translateY(-1px) !important; }
        .btn-voltar > button { height:48px !important; }
        </style>""", unsafe_allow_html=True)
        ca, cb = st.columns(2)
        with ca:
            st.markdown('<div class="btn-confirm-del">', unsafe_allow_html=True)
            if st.button("✓ Sim, apagar tudo de hoje", use_container_width=True, key="confirmar_del_dia"):
                limpar_dia(hoje_str)
                st.session_state.confirm_limpar_dia = False
                st.toast(f"✅ Todos os registros de {hoje_str} foram apagados!", icon="🧹")
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
        with cb:
            st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
            if st.button("✕ Cancelar", use_container_width=True, key="cancelar_del_dia"):
                st.session_state.confirm_limpar_dia = False
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

    tag_html = {
        0: '<span style="background:#EBF0FB;color:#3B5EC6;padding:3px 10px;border-radius:100px;font-size:10px;font-weight:800;">Separação</span>',
        1: '<span style="background:#E8F2EC;color:#4A7C59;padding:3px 10px;border-radius:100px;font-size:10px;font-weight:800;">Embalagem</span>',
        2: '<span style="background:#FBF2E6;color:#C47B2A;padding:3px 10px;border-radius:100px;font-size:10px;font-weight:800;">Conferência</span>',
    }

    # ── Filtro de dia do Histórico ──────────────────────────────────────────
    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

    datas_hist = sorted({
        r[6].split(" ")[0] for r in regs_para_tabela if r[6] and " " in str(r[6])
    }, reverse=True) if regs_para_tabela else []

    _dia_hist_padrao = datas_hist[0] if datas_hist else "Todos os dias"
    opcoes_hist_data = ["Todos os dias"] + datas_hist
    _idx_hist_padrao = opcoes_hist_data.index(_dia_hist_padrao) if _dia_hist_padrao in opcoes_hist_data else 0

    fh1, fh2 = st.columns([2, 1])
    with fh1:
        filtro_hist_data = st.selectbox(
            "📋 Histórico — filtrar por dia",
            opcoes_hist_data,
            index=_idx_hist_padrao,
            key="admin_hist_filtro_data"
        )
    with fh2:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        n_total_hist = len(regs_para_tabela)
        n_dia_hist = len([r for r in regs_para_tabela
                          if filtro_hist_data == "Todos os dias"
                          or str(r[6]).startswith(filtro_hist_data)])
        st.markdown(
            f'<div style="background:#F0F5FF;border:1.5px solid #3B5EC6;border-radius:10px;'
            f'padding:8px 14px;font-size:12px;font-weight:700;color:#3B5EC6;text-align:center;">'
            f'{n_dia_hist} registro(s)</div>',
            unsafe_allow_html=True
        )

    if filtro_hist_data != "Todos os dias":
        regs_hist_filtrados = [r for r in regs_para_tabela if str(r[6]).startswith(filtro_hist_data)]
    else:
        regs_hist_filtrados = regs_para_tabela

    if regs_hist_filtrados:
        # Monta set de (pedido, operador) que tiveram pausa no mesmo dia
        pedidos_com_pausa = set()
        pausas_hist = buscar_pausas_log()
        for p in pausas_hist:
            if filtro_hist_data == "Todos os dias" or str(p[4] or "").startswith(filtro_hist_data):
                pedidos_com_pausa.add((str(p[1]), str(p[2])))

        hist_rows = ""
        for r in regs_hist_filtrados[:200]:
            # r: (id, pedido, operador, etapa, etapa_idx, tempo_s, data_fim, inicio, qtd_pecas)
            fim_str    = r[6] if r[6] else "—"
            inicio_str = r[7] if r[7] else "—"
            qtd_str    = str(r[8]) if r[8] is not None else "—"
            teve_pausa = (str(r[1]), str(r[2])) in pedidos_com_pausa
            pausa_tag  = (
                '<span style="background:#FFF0E6;color:#E07B3A;font-size:10px;font-weight:800;'
                'padding:2px 8px;border-radius:100px;">Sim</span>'
                if teve_pausa else
                '<span style="color:#C0BAB4;font-size:10px;font-weight:700;">Não</span>'
            )
            hist_rows += f"""<tr>
              <td style="padding:11px 16px;font-family:monospace;font-size:12px;font-weight:700;color:#1A1714;">{r[1]}</td>
              <td style="padding:11px 10px;font-size:13px;font-weight:700;color:#1A1714;">{r[2]}</td>
              <td style="padding:11px 10px;">{tag_html.get(r[4], r[3])}</td>
              <td style="padding:11px 10px;font-family:monospace;font-size:12px;font-weight:700;color:#4A7C59;text-align:center;">{fmt(r[5])}</td>
              <td style="padding:11px 10px;font-size:11px;font-weight:700;color:#3B5EC6;text-align:center;">{qtd_str}</td>
              <td style="padding:11px 10px;text-align:center;">{pausa_tag}</td>
              <td style="padding:11px 10px;font-size:11px;color:#9C9490;text-align:center;">{inicio_str}</td>
              <td style="padding:11px 10px;font-size:11px;color:#9C9490;text-align:center;">{fim_str}</td>
            </tr>"""

        n_hist = min(len(regs_hist_filtrados), 200)
        hist_height = 56 + (n_hist * 46) + 20

        lbl_hist = f"Histórico de Pedidos · {filtro_hist_data}" if filtro_hist_data != "Todos os dias" else "Histórico de Pedidos"

        components.html(f"""<!DOCTYPE html><html><head>
        <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@700;800;900&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
        <style>
        *{{margin:0;padding:0;box-sizing:border-box;}} body{{background:transparent;font-family:Nunito,sans-serif;}}
        .lbl{{font-size:9px;font-weight:800;letter-spacing:2px;text-transform:uppercase;color:#9C9490;margin-bottom:10px;}}
        .wrap{{background:#fff;border-radius:16px;border:1.5px solid #EDE9E4;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.05);}}
        table{{width:100%;border-collapse:collapse;}} thead tr{{background:#1A1714;}}
        th{{padding:12px 10px;font-size:9px;font-weight:800;letter-spacing:1.5px;text-transform:uppercase;color:rgba(255,255,255,0.45);text-align:center;}}
        th:first-child{{text-align:left;padding-left:16px;}} th:nth-child(2){{text-align:left;}}
        tbody tr{{border-bottom:1px solid #F2EEE9;}} tbody tr:last-child{{border-bottom:none;}}
        tbody tr:hover{{background:#FDFAF9;}}
        </style></head><body>
        <div class="lbl">{lbl_hist}</div>
        <div class="wrap"><table><thead><tr>
          <th>Pedido</th><th>Operador</th><th>Etapa</th><th>Tempo</th>
          <th style="color:#7B9FE0;">Qtd Peças</th>
          <th style="color:#E07B3A;">Pausa</th>
          <th style="color:#A0C8E0;">Início</th>
          <th style="color:#A0C8E0;">Fim</th>
        </tr></thead><tbody>{hist_rows}</tbody></table></div>
        </body></html>""", height=hist_height, scrolling=False)

    st.markdown("<br>", unsafe_allow_html=True)

    if regs:
        st.markdown("""
        <style>
        .btn-pdf > button { background:linear-gradient(135deg,#C8566A,#9E3F52) !important; color:#fff !important; border:none !important;
            box-shadow:0 5px 0 rgba(100,20,35,0.40),0 8px 20px rgba(200,86,106,0.28) !important; font-weight:800 !important; height:54px !important; }
        .btn-pdf > button:hover { transform:translateY(-2px) !important; }
        .btn-xml > button { background:linear-gradient(135deg,#3B5EC6,#2a469e) !important; color:#fff !important; border:none !important;
            box-shadow:0 5px 0 rgba(20,30,100,0.40),0 8px 20px rgba(59,94,198,0.28) !important; font-weight:800 !important; height:54px !important; }
        .btn-xml > button:hover { transform:translateY(-2px) !important; }
        </style>
        """, unsafe_allow_html=True)

        ts = now_br().strftime("%Y%m%d_%H%M")

        buf_csv = io.StringIO()
        csv.writer(buf_csv).writerows(
            [["ID","Pedido","Operador","Etapa","EtapaIdx","Tempo(s)","Data Fim","Início","Qtd Peças"]] + list(regs))

        pdf_bytes = gerar_pdf(regs, op_map, ped_comp, ops_ativ, avg)

        # ── Gerar XLS ──
        df_xls = pd.DataFrame(list(regs), columns=["ID","Pedido","Operador","Etapa","EtapaIdx","Tempo(s)","Data Fim","Início","Qtd Peças"])
        buf_xls = io.BytesIO()
        with pd.ExcelWriter(buf_xls, engine="openpyxl") as writer:
            df_xls.to_excel(writer, index=False, sheet_name="Producao")
        buf_xls.seek(0)
        xls_bytes = buf_xls.read()

        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown('<div class="btn-voltar">', unsafe_allow_html=True)
            st.download_button("⬇  Exportar CSV", data=buf_csv.getvalue().encode(),
                file_name=f"vi_producao_{ts}.csv", mime="text/csv", use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        with c2:
            st.markdown('<div class="btn-xml">', unsafe_allow_html=True)
            st.download_button("📊  Exportar XLS", data=xls_bytes,
                file_name=f"vi_producao_{ts}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        with c3:
            st.markdown('<div class="btn-pdf">', unsafe_allow_html=True)
            st.download_button("📄  Exportar PDF", data=pdf_bytes,
                file_name=f"vi_relatorio_{ts}.pdf", mime="application/pdf", use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)


# ─────────────────────────────────────
#  TELA: OPERAÇÕES EM ANDAMENTO
# ─────────────────────────────────────
def fmt_elapsed(ini_ts):
    s = max(int(time.time()) - int(ini_ts), 0)
    h, r = divmod(s, 3600); m, s2 = divmod(r, 60)
    return f"{h:02d}:{m:02d}:{s2:02d}", s

def tela_operacoes():
    _auto_refresh_watcher()
    render_logo()

    # Garante que a chave existe antes de qualquer renderização condicional
    if "busca_pedido_painel" not in st.session_state:
        st.session_state["busca_pedido_painel"] = ""

    sessoes_raw = buscar_todas_sessoes_ativas()
    # Separa em andamento e pausadas — ambas aparecem no painel
    sessoes_ativas  = [s for s in sessoes_raw if int(s.get("iniciado_em", 1)) != 0]
    sessoes_pausadas = [s for s in sessoes_raw if int(s.get("iniciado_em", 1)) == 0]
    # Mostra ativas primeiro, depois pausadas
    sessoes = sessoes_ativas + sessoes_pausadas
    n_total = len(sessoes)

    ETAPA_COR  = ["#C8566A", "#3B7DD8", "#4A7C59"]
    ETAPA_BG   = ["#FFF0F2", "#F0F5FF", "#F0F7F3"]
    ETAPA_ICON = ["📦", "🗃️", "✅"]

    # ── Header com contador + botões ─────────────────────────────────────────
    col_h1, col_h2, col_h3 = st.columns([3, 1.2, 1.2])
    with col_h1:
        st.markdown(f"""
        <div style="margin-bottom:4px;">
          <div style="font-size:10px;font-weight:800;letter-spacing:2.5px;text-transform:uppercase;
              color:#9C9490;margin-bottom:2px;">Painel de Operações</div>
          <div style="font-size:22px;font-weight:900;color:#1A1714;letter-spacing:-0.3px;">
            Operações
            <span style="font-size:14px;font-weight:700;color:#C8566A;margin-left:6px;">
              {len(sessoes_ativas)} ativa{"s" if len(sessoes_ativas) != 1 else ""}
            </span>
            {(f'<span style="font-size:13px;font-weight:700;color:#7C3AED;margin-left:4px;">· {len(sessoes_pausadas)} pausada{"s" if len(sessoes_pausadas) != 1 else ""}</span>') if sessoes_pausadas else ""}
          </div>
        </div>""", unsafe_allow_html=True)
    with col_h2:
        st.markdown('<div class="btn-iniciar btn-sm">', unsafe_allow_html=True)
        if st.button("＋ Nova Op.", use_container_width=True):
            st.session_state.tela = "home"; st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    with col_h3:
        st.markdown('<div class="btn-voltar btn-sm">', unsafe_allow_html=True)
        if st.button("← Lobby", use_container_width=True):
            st.session_state.tela = "home"; st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

    # ── Campo de busca por número de pedido ──────────────────────────────────
    st.markdown("""
    <style>
    div[data-testid="stTextInput"] label { display:none !important; }
    div[data-testid="stTextInput"] input {
        border: 2px solid #E0DBD4 !important;
        border-radius: 12px !important;
        background: #fff !important;
        font-family: 'Nunito', sans-serif !important;
        font-size: 16px !important;
        font-weight: 700 !important;
        padding: 14px 18px !important;
        height: 52px !important;
        transition: border-color .2s !important;
    }
    div[data-testid="stTextInput"] input:focus {
        border-color: #C8566A !important;
        box-shadow: 0 0 0 4px rgba(200,86,106,0.10) !important;
    }
    </style>""", unsafe_allow_html=True)

    busca = st.text_input("_busca", placeholder="🔍  Buscar por número do pedido...",
                          key="busca_pedido_painel")
    busca = busca.strip()

    # ── Sem sessões ──────────────────────────────────────────────────────────
    if not sessoes:
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        components.html("""<!DOCTYPE html><html><body style="background:transparent;
        font-family:sans-serif;padding:0;margin:0;">
        <div style="background:#F0F7F3;border:1.5px solid #4A7C59;border-radius:16px;
                    padding:48px 24px;text-align:center;">
            <div style="font-size:40px;margin-bottom:12px;">✅</div>
            <div style="font-size:16px;font-weight:800;color:#2d5a3d;margin-bottom:6px;">
                Nenhuma operação em andamento</div>
            <div style="font-size:12px;color:#4A7C59;font-weight:600;">
                Use "＋ Nova Op." para iniciar um processo.</div>
        </div></body></html>""", height=170, scrolling=False)
        return

    # ── Filtragem por busca ──────────────────────────────────────────────────
    if busca:
        sessoes_visiveis = [s for s in sessoes if busca in str(s.get("pedido",""))]
    else:
        sessoes_visiveis = sessoes

    if busca and not sessoes_visiveis:
        st.markdown(f"""
        <div style="background:#FEF3C7;border:1.5px solid #F59E0B;border-radius:12px;
                    padding:16px 20px;text-align:center;margin:8px 0;">
          <div style="font-size:14px;font-weight:800;color:#92400E;">
            Nenhuma operação com pedido <code>{busca}</code> em andamento.</div>
        </div>""", unsafe_allow_html=True)
        return

    # ── Separador de resultados ──────────────────────────────────────────────
    if busca:
        st.markdown(f"""
        <div style="font-size:11px;font-weight:800;color:#4A7C59;
            letter-spacing:1.5px;text-transform:uppercase;margin:10px 0 6px;">
            ✓ {len(sessoes_visiveis)} resultado(s) para "{busca}"
        </div>""", unsafe_allow_html=True)
    else:
        st.markdown(f"""
        <div style="font-size:11px;font-weight:700;color:#9C9490;margin:10px 0 6px;">
            Todos os pedidos — role para encontrar o seu
        </div>""", unsafe_allow_html=True)

    # ── Pré-carrega qtd de peças de todos os pedidos visíveis de uma vez ────
    _pedidos_ids = list({s.get("pedido","") for s in sessoes_visiveis if s.get("pedido")})
    _qtd_map = {}
    if _pedidos_ids:
        _ids_filter = ",".join(str(p) for p in _pedidos_ids)
        _rows_est = _get("pedidos_base",
            f"select=numero,est_alocado&numero=in.({_ids_filter})")
        if isinstance(_rows_est, list):
            for _r in _rows_est:
                _n = str(_r.get("numero",""))
                _v = _r.get("est_alocado")
                _qtd_map[_n] = int(float(_v)) if _v else None

    # ── Cards das sessões ────────────────────────────────────────────────────
    for s in sessoes_visiveis:
        ped      = s.get("pedido", "")
        op       = s.get("operador", "")
        eta_idx  = int(s.get("etapa_idx", 0))
        ini      = int(s.get("iniciado_em", 0))
        tp       = int(s.get("tempo_pausado") or 0)   # segundos já trabalhados antes
        cor      = ETAPA_COR[eta_idx]
        bg       = ETAPA_BG[eta_idx]
        icon     = ETAPA_ICON[eta_idx]
        lbl      = ETAPAS_LBL[eta_idx]
        pausado      = (ini == 0)  # True se sessão está pausada
        # Tempo total acumulado
        total_s  = tp if pausado else tp + max(int(time.time()) - ini, 0)
        h_t, r_t = divmod(total_s, 3600); m_t, s_t = divmod(r_t, 60)
        elapsed_str = f"{h_t:02d}:{m_t:02d}:{s_t:02d}"
        ini_op   = (op[0]+op[1]).upper() if len(op) >= 2 else op[0].upper()
        uid      = f"{ped}_{eta_idx}"
        qtd_pecas_card = _qtd_map.get(str(ped))
        qtd_str_card   = f" · {qtd_pecas_card} pçs" if qtd_pecas_card else ""

        # Visual diferente para pausados
        card_border  = "#9C9490" if pausado else cor
        card_bg      = "#F7F5F2" if pausado else "#fff"
        av_bg        = "linear-gradient(135deg,#9C9490,#7a7470)" if pausado else f"linear-gradient(135deg,{cor},{cor}99)"
        timer_lbl_txt = "pausado" if pausado else "em andamento"
        timer_cor    = "#9C9490" if pausado else cor
        # JS: só conta se ativo
        js_timer = "" if pausado else f"""
        (function(){{
          var ini={ini}, paused={tp}, el=document.getElementById('tv_{uid}');
          function u(){{
            var e=paused+Math.floor(Date.now()/1000)-ini;
            var h=Math.floor(e/3600),m=Math.floor((e%3600)/60),s=e%60;
            el.textContent=String(h).padStart(2,'0')+':'+String(m).padStart(2,'0')+':'+String(s).padStart(2,'0');
          }}
          u(); setInterval(u,1000);
        }})();
        """

        components.html(f"""<!DOCTYPE html><html><head>
        <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@700;800;900&family=DM+Mono:wght@500&display=swap" rel="stylesheet">
        <style>
        *{{margin:0;padding:0;box-sizing:border-box;}}
        body{{background:transparent;font-family:Nunito,sans-serif;}}
        .card{{background:{card_bg};border:1.5px solid #EDE9E4;border-left:5px solid {card_border};
               border-radius:14px;padding:14px 16px;
               box-shadow:0 2px 14px rgba(0,0,0,0.06);
               display:flex;align-items:center;gap:14px;}}
        .av{{width:44px;height:44px;border-radius:50%;flex-shrink:0;
             background:{av_bg};
             display:flex;align-items:center;justify-content:center;
             font-size:15px;font-weight:900;color:#fff;
             box-shadow:0 3px 10px rgba(0,0,0,0.16);}}
        .name{{font-size:14px;font-weight:900;color:#1A1714;margin-bottom:3px;}}
        .badge{{display:inline-flex;align-items:center;gap:4px;background:{bg};color:{cor};
                font-size:10px;font-weight:800;padding:2px 9px;border-radius:20px;margin-right:6px;}}
        .ped{{font-family:monospace;font-size:13px;color:#5C5450;font-weight:700;}}
        .timer-wrap{{text-align:right;flex-shrink:0;}}
        .timer-lbl{{font-size:9px;font-weight:800;color:#9C9490;letter-spacing:1.5px;
                    text-transform:uppercase;margin-bottom:2px;}}
        .timer-val{{font-family:'DM Mono',monospace;font-size:24px;font-weight:500;
                    color:{timer_cor};letter-spacing:-1px;line-height:1;}}
        </style></head>
        <body>
        <div class="card">
          <div class="av">{ini_op}</div>
          <div style="flex:1;min-width:0;">
            <div class="name">{op}</div>
            <div style="display:flex;align-items:center;flex-wrap:wrap;gap:4px;">
              <span class="badge">{icon} {lbl}</span>
              <span class="ped">#{ped}</span>
              {(f'<span style="display:inline-flex;align-items:center;gap:3px;background:{bg};border:1.5px solid {cor}33;border-radius:20px;padding:1px 9px;margin-left:4px;"><span style="font-family:monospace;font-size:12px;font-weight:900;color:{cor};">{qtd_pecas_card}</span><span style="font-size:10px;font-weight:700;color:{cor}99;">pçs</span></span>') if qtd_pecas_card else ""}
              {'<span style="display:inline-flex;align-items:center;gap:3px;background:#F3F0FF;border:1.5px solid #7C3AED33;border-radius:20px;padding:2px 10px;margin-left:2px;font-size:10px;font-weight:900;color:#7C3AED;">⏸ PAUSADO</span>' if pausado else ""}
            </div>
          </div>
          <div class="timer-wrap">
            <div class="timer-lbl">{timer_lbl_txt}</div>
            <div class="timer-val" id="tv_{uid}">{elapsed_str}</div>
          </div>
        </div>
        <script>{js_timer}</script>
        </body></html>""", height=80, scrolling=False)

        # ── Botões de ação: Finalizar | Pausar | Pausar p/ amanhã ──────────
        # Estados de pausa por card (uid)
        _k_modo   = f"painel_pausa_modo_{uid}"    # None | "pausar" | "amanha"
        _k_senha  = f"painel_pausa_senha_{uid}"
        _k_erro   = f"painel_pausa_erro_{uid}"
        for _k, _v in [(_k_modo, None), (_k_senha, ""), (_k_erro, False)]:
            if _k not in st.session_state:
                st.session_state[_k] = _v

        modo_pausa = st.session_state[_k_modo]

        if modo_pausa is None:
            if pausado:
                # ── Card pausado: ▶ Retomar (sem senha) + Finalizar ────────
                st.markdown(f"""<style>
                .btn-retomar-{uid} > button {{
                    background: linear-gradient(135deg,#7C3AED,#5b21b6) !important;
                    color:#fff !important; border:none !important;
                    border-radius:10px !important; height:50px !important;
                    font-size:15px !important; font-weight:900 !important;
                    box-shadow:0 4px 0 rgba(90,30,180,0.35) !important;
                    letter-spacing:.3px !important;
                }}
                .btn-retomar-{uid} > button:hover {{ filter:brightness(1.08) !important; }}
                .btn-finalizar > button {{ height:50px !important; }}
                </style>""", unsafe_allow_html=True)
                col_r, col_f = st.columns([3, 2])
                with col_r:
                    st.markdown(f'<div class="btn-retomar-{uid}">', unsafe_allow_html=True)
                    if st.button(f"▶  Retomar · #{ped} · {op}",
                                 use_container_width=True, key=f"retomar_{uid}"):
                        novo_ini = int(time.time())
                        _patch("sessoes_ativas",
                               f"pedido=eq.{ped}&etapa_idx=eq.{eta_idx}",
                               {"iniciado_em": novo_ini})
                        buscar_todas_sessoes_ativas.clear()
                        st.toast(f"▶ {op} · Pedido #{ped} retomado!", icon="▶️")
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
                with col_f:
                    st.markdown('<div class="btn-finalizar">', unsafe_allow_html=True)
                    if st.button(f"■ Finalizar{qtd_str_card}",
                                 use_container_width=True, key=f"fin_{uid}"):
                        salvar(ped, op, ETAPAS[eta_idx], eta_idx, max(tp, 1), qtd_pecas_card)
                        remover_sessao_ativa(ped, eta_idx)
                        if eta_idx == 2:
                            marcar_concluido(ped)
                        st.session_state["busca_pedido_painel"] = ""
                        st.toast(f"✅ {op} · Pedido #{ped} finalizado em {fmt(tp)}!", icon="🎉")
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
            else:
                # ── Card ativo: Finalizar + ⏸ Pausar (com senha) ───────────
                st.markdown("""<style>
                .btn-finalizar > button { height:46px !important; }
                .btn-pausar > button {
                    background:#fff !important; color:#E07B3A !important;
                    border:2px solid #E07B3A !important; border-radius:10px !important;
                    height:46px !important; font-size:12px !important; font-weight:800 !important;
                }
                </style>""", unsafe_allow_html=True)
                col_f, col_p = st.columns([3, 2])
                with col_f:
                    st.markdown('<div class="btn-finalizar">', unsafe_allow_html=True)
                    if st.button(f"■ Finalizar · #{ped}{qtd_str_card} · {op}",
                                 use_container_width=True, key=f"fin_{uid}"):
                        tempo = max(tp + (int(time.time()) - ini), 1)
                        salvar(ped, op, ETAPAS[eta_idx], eta_idx, tempo, qtd_pecas_card)
                        remover_sessao_ativa(ped, eta_idx)
                        if eta_idx == 2:
                            marcar_concluido(ped)
                        st.session_state["busca_pedido_painel"] = ""
                        st.toast(f"✅ {op} · Pedido #{ped} finalizado em {fmt(tempo)}!", icon="🎉")
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
                with col_p:
                    st.markdown('<div class="btn-pausar">', unsafe_allow_html=True)
                    if st.button("⏸ Pausar", use_container_width=True, key=f"pausar_{uid}"):
                        st.session_state[_k_modo]  = "pausar"
                        st.session_state[_k_senha] = ""
                        st.session_state[_k_erro]  = False
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)

        else:
            # ── Bloco de confirmação de pausa com senha ─────────────────
            cor_modo = "#E07B3A"
            bg_modo  = "#FFF8F0"

            components.html(f"""<!DOCTYPE html><html><head>
            <link href="https://fonts.googleapis.com/css2?family=Nunito:wght@700;800;900&display=swap" rel="stylesheet">
            </head><body style="background:transparent;font-family:Nunito,sans-serif;margin:0;">
            <div style="background:{bg_modo};border:2px solid {cor_modo};border-radius:12px;padding:12px 16px;">
              <div style="font-size:13px;font-weight:900;color:{cor_modo};margin-bottom:4px;">⏸ Pausar atividade</div>
              <div style="font-size:11px;font-weight:700;color:#5C5450;">O tempo ficará salvo. O operador retoma quando voltar.</div>
              <div style="font-size:11px;font-weight:700;color:#9C9490;margin-top:6px;">
                Pedido <span style="font-family:monospace;font-weight:900;">#{ped}</span>
                &nbsp;·&nbsp; {op} &nbsp;·&nbsp; Tempo: {fmt(tp + max(int(time.time()) - ini, 0))}
              </div>
            </div></body></html>""", height=96, scrolling=False)

            st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

            # Campo de senha
            senha_input = st.text_input(
                "_senha_pausa", placeholder="🔑 Senha do gestor...",
                type="password", label_visibility="collapsed",
                key=f"senha_input_{uid}"
            )
            # Campo de motivo
            motivo_input_painel = st.text_input(
                "_motivo_pausa_painel", placeholder="📝 Motivo da pausa (opcional)...",
                label_visibility="collapsed",
                key=f"motivo_input_{uid}"
            )
            if st.session_state[_k_erro]:
                st.markdown('<div style="color:#DC2626;font-size:11px;font-weight:700;'
                            'text-align:center;margin-top:2px;">❌ Senha incorreta</div>',
                            unsafe_allow_html=True)

            st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
            st.markdown(f"""<style>
            .btn-conf-pausa-{uid} > button {{
                background: {cor_modo} !important; color:#fff !important;
                border:none !important; border-radius:10px !important;
                height:44px !important; font-size:13px !important; font-weight:800 !important;
            }}
            .btn-canc-pausa-{uid} > button {{
                background: #fff !important; color:#5C5450 !important;
                border:2px solid #DDD8D2 !important; border-radius:10px !important;
                height:44px !important; font-size:13px !important; font-weight:800 !important;
            }}
            </style>""", unsafe_allow_html=True)
            cc1, cc2 = st.columns(2)
            with cc1:
                st.markdown(f'<div class="btn-conf-pausa-{uid}">', unsafe_allow_html=True)
                if st.button("✓  Confirmar", use_container_width=True, key=f"conf_pausa_{uid}"):
                    if senha_input == ADMIN_SENHA:
                        tempo_atual = tp + max(int(time.time()) - ini, 0)
                        pausar_para_amanha(ped, eta_idx, op, tempo_atual)
                        _patch("sessoes_ativas",
                               f"pedido=eq.{ped}&etapa_idx=eq.{eta_idx}",
                               {"iniciado_em": 0})
                        registrar_pausa_log(ped, eta_idx, op, tempo_atual, motivo_input_painel)
                        buscar_pausas_log.clear()
                        st.session_state[_k_modo]  = None
                        st.session_state[_k_erro]  = False
                        st.toast(f"⏸ Pedido #{ped} · {op} · pausado! Tempo salvo: {fmt(tempo_atual)}", icon="⏸")
                        st.rerun()
                    else:
                        st.session_state[_k_erro] = True
                        st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)
            with cc2:
                st.markdown(f'<div class="btn-canc-pausa-{uid}">', unsafe_allow_html=True)
                if st.button("✕  Cancelar", use_container_width=True, key=f"canc_pausa_{uid}"):
                    st.session_state[_k_modo] = None
                    st.session_state[_k_erro] = False
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

        st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)


# ─────────────────────────────────────
#  ROUTER
# ─────────────────────────────────────
{
    "home":        tela_home,
    "producao":    tela_producao,
    "operacoes":   tela_operacoes,
    "admin_login": tela_admin_login,
    "admin":       tela_admin,
}.get(st.session_state.tela, tela_home)()
